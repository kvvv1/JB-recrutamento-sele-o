import json
import logging
import math
import os
import re
import sqlite3
import tempfile
import time
from collections import defaultdict
from datetime import datetime, timedelta, date
from functools import wraps
from io import BytesIO
from pathlib import Path
from shutil import which
from typing import Optional
import inspect
from flask import (
    Flask,
    abort,
    after_this_request,
    current_app,
    flash,
    g,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_paginate import Pagination, get_page_parameter
from flask_socketio import SocketIO
from gevent import monkey
from markupsafe import Markup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from passlib.hash import pbkdf2_sha256
from PyPDF2 import PdfMerger, PdfReader
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

logger = logging.getLogger(__name__)

WEASYPRINT_AVAILABLE = False
WEASYPRINT_IMPORT_ERROR: Optional[Exception] = None
PDFKIT_IMPORTED = False
PDFKIT_IMPORT_ERROR: Optional[Exception] = None
WKHTMLTOPDF_CMD: Optional[str] = None
PDFKIT_CONFIGURATION = None

try:
    from weasyprint import HTML
    import pydyf  # type: ignore

    pdf_init_params = inspect.signature(pydyf.PDF.__init__).parameters
    if len(pdf_init_params) < 3:
        raise RuntimeError(
            "pydyf instalado é incompatível (versão muito antiga para o WeasyPrint atual)."
        )

    WEASYPRINT_AVAILABLE = True
except Exception as exc:  # noqa: BLE001
    WEASYPRINT_IMPORT_ERROR = exc
    logger.warning(
        "WeasyPrint indisponível: %s. Instale as dependências GTK+ ou mantenha o fallback.",
        exc,
    )

try:
    import pdfkit  # type: ignore

    PDFKIT_IMPORTED = True
    WKHTMLTOPDF_CMD = which("wkhtmltopdf")
    if WKHTMLTOPDF_CMD is None:
        logger.warning(
            "pdfkit importado, mas o executável 'wkhtmltopdf' não foi encontrado no PATH."
        )
    else:
        try:
            PDFKIT_CONFIGURATION = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_CMD)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Falha ao configurar pdfkit: %s", exc)
            PDFKIT_CONFIGURATION = None
except ImportError as exc:
    PDFKIT_IMPORT_ERROR = exc
    logger.warning("pdfkit indisponível: %s.", exc)

DEFAULT_PDFKIT_OPTIONS = {
    "page-size": "A4",
    "margin-top": "0.75in",
    "margin-right": "0.75in",
    "margin-bottom": "0.75in",
    "margin-left": "0.75in",
    "encoding": "UTF-8",
    "no-outline": None,
    "enable-local-file-access": None,
}


def _fix_static_urls_for_pdfkit(rendered_html: str) -> str:
    """Converte URLs /static/... em caminhos file:// absolutos para o wkhtmltopdf."""
    static_root = Path(current_app.static_folder).resolve()

    def repl(match: re.Match) -> str:
        attr = match.group(1)
        rel_path = match.group(2)  # começa com /static/
        rel_without_prefix = rel_path[len("/static/") :]
        abs_path = static_root / rel_without_prefix
        # wkhtmltopdf no Windows aceita file:///C:/...
        return f'{attr}="file:///{abs_path.as_posix()}"'

    # Substitui em atributos src="/static/..." e href="/static/..."
    pattern = r'(?:\s)(src|href)="(/static/[^"]+)"'
    return re.sub(pattern, lambda m: " " + repl(m), rendered_html)


def generate_pdf_from_html(rendered_html: str, output_path: Path) -> str:
    """
    Gera um PDF a partir de HTML utilizando o melhor backend disponível.

    Retorna o nome do backend utilizado ou levanta RuntimeError em caso de falha.
    """

    errors = []

    if WEASYPRINT_AVAILABLE:
        try:
            HTML(string=rendered_html, base_url=current_app.static_folder).write_pdf(
                output_path
            )
            return "weasyprint"
        except Exception as exc:  # noqa: BLE001
            logger.warning("Falha ao gerar PDF com WeasyPrint: %s", exc, exc_info=True)
            errors.append(f"WeasyPrint: {exc}")

    if PDFKIT_IMPORTED:
        if WKHTMLTOPDF_CMD is None or PDFKIT_CONFIGURATION is None:
            errors.append("pdfkit: executável 'wkhtmltopdf' não encontrado/sem configuração")
        else:
            try:
                fixed_html = _fix_static_urls_for_pdfkit(rendered_html)
                pdfkit.from_string(
                    fixed_html,
                    str(output_path),
                    configuration=PDFKIT_CONFIGURATION,
                    options=DEFAULT_PDFKIT_OPTIONS,
                )
                return "pdfkit"
            except Exception as exc:  # noqa: BLE001
                logger.warning("Falha ao gerar PDF com pdfkit: %s", exc, exc_info=True)
                errors.append(f"pdfkit: {exc}")

    if WEASYPRINT_IMPORT_ERROR:
        errors.append(f"WeasyPrint indisponível: {WEASYPRINT_IMPORT_ERROR}")
    if PDFKIT_IMPORT_ERROR:
        errors.append(f"pdfkit indisponível: {PDFKIT_IMPORT_ERROR}")

    detail = " | ".join(errors) if errors else "Nenhum backend de PDF disponível."
    raise RuntimeError(f"Não foi possível gerar o PDF. {detail}")


def ensure_parent_directory(path: Path) -> None:
    """Garante que o diretório pai do arquivo exista."""

    path.parent.mkdir(parents=True, exist_ok=True)



# Funções utilitárias para obter conexões com o SQL Server


def _get_connection_string(env_var_name):

    conn_str = os.environ.get(env_var_name)

    if not conn_str:

        raise RuntimeError(

            f"Defina a variável de ambiente {env_var_name} com a string de conexão do SQL Server."

        )

    return conn_str



def get_sql_server_connection():
    connection_string = "DRIVER={ODBC Driver 18 for SQL Server};SERVER=SRVDB01;DATABASE=SistemaRS;UID=sa;PWD=jblimpeza2015;TrustServerCertificate=yes;Encrypt=yes;"
    return pyodbc.connect(connection_string)


def get_jbc_connection():
    connection_string = "DRIVER={ODBC Driver 18 for SQL Server};SERVER=SRVDB01;DATABASE=Jbc ;UID=sa;PWD=jblimpeza2015;TrustServerCertificate=yes;Encrypt=yes;"
    return pyodbc.connect(connection_string)



def format_date(date_value):

    """Format date to DD/MM/YYYY."""

    if not date_value:

        return "Data não disponível"

    try:

        if isinstance(date_value, (datetime.date, datetime.datetime)):

            return date_value.strftime('%d/%m/%Y')

        if isinstance(date_value, str) and '-' in date_value:

            return datetime.datetime.strptime(date_value, '%Y-%m-%d').strftime('%d/%m/%Y')

        return "Formato de data inesperado" 

    except Exception as e:

        print(f"Erro ao formatar data: {e} | Valor recebido: {date_value}")

        return "Data inválida"       





# Função para registrar logs das fichas salvas

def salvar_ficha_log(dados):

    log_dir = "logs"

    if not os.path.exists(log_dir):

        os.makedirs(log_dir)  # Cria a pasta se não existir

    

    with open(os.path.join(log_dir, "fichas_salvas.txt"), "a", encoding="utf-8") as f:

        f.write(f"\n{'='*50}\n")

        f.write(f"Data/Hora: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        for chave, valor in dados.items():

            f.write(f"{chave}: {valor}\n")

        f.write(f"{'='*50}\n")



app = Flask(__name__, static_folder='static')

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-me')

app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

app.config['SESSION_REFRESH_EACH_REQUEST'] = True  

app.config['SESSION_COOKIE_SECURE'] = False  

app.config['SESSION_COOKIE_HTTPONLY'] = True  

app.config['SESSION_COOKIE_SAMESITE'] = 'Lax' 



@app.before_request

def make_session_permanent():

    session.permanent = True  



socketio = SocketIO(app, async_mode='gevent', cors_allowed_origins='*')

DATABASE = 'app.db'



BRASILIA_TZ = pytz.timezone('America/Sao_Paulo')



login_manager = LoginManager()

login_manager.init_app(app)

login_manager.login_view = 'login'



UPLOAD_DIR = os.path.join('static', 'uploads')

if not os.path.exists(UPLOAD_DIR):

    os.makedirs(UPLOAD_DIR)









def get_dates_for_period(period):

    today = datetime.now().strftime('%Y-%m-%d')

    if period == 'HOJE':

        return today, today

    elif period == '3DIAS':

        return (datetime.now() - timedelta(days=3)).strftime('%Y-%m-%d'), today

    elif period == 'SEMANA':

        return (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'), today

    elif period == 'MES':

        return (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'), today

    elif period == 'ANO':

        return (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d'), today

    else:

        return today, today



def determinar_situacao(form_data):

    admitido = (form_data.get('admitido') or '').strip().lower()

    avaliacao_rh = (form_data.get('avaliacao_rh') or '').strip().lower()

    sindicancia = (form_data.get('sindicancia') or '').strip().lower()

    avaliacao_gerencia = (form_data.get('avaliacao_gerencia') or '').strip().lower()



    # 1. Admitido (prioridade máxima)

    if admitido == 'sim':

        return 'Admitido'

    # 1.5. Em processo de admissão (prioridade alta)

    if admitido == 'em processo de admissão':

        return 'Em processo de admissão'



    # 2. Reprovado RH

    if avaliacao_rh == 'reprovado':

        return 'Reprovado RH'



    # 3. Reprovado Sindicância

    if avaliacao_rh == 'aprovado' and sindicancia == 'reprovado':

        return 'Reprovado Sindicância'



    # 4. Reprovado Gerência

    if avaliacao_rh == 'aprovado' and sindicancia == 'aprovado' and avaliacao_gerencia == 'reprovado':

        return 'Reprovado Gerência'



    # 5. Em Verificação (quando sindicância está em verificação e RH aprovado)

    if avaliacao_rh == 'aprovado' and sindicancia == 'em verificação':

        return 'Em Verificação'



    # 6. Em Conversa (Gestor)

    if avaliacao_rh == 'aprovado' and sindicancia == 'aprovado' and avaliacao_gerencia == 'em conversa':

        return 'Em Conversa'



    # 7. Aprovado Gerência

    if avaliacao_rh == 'aprovado' and sindicancia == 'aprovado' and avaliacao_gerencia == 'aprovado':

        return 'Aprovado Gerência'



    # 8. Aprovado Sindicância (aguardando avaliação gerência)

    if avaliacao_rh == 'aprovado' and sindicancia == 'aprovado' and (avaliacao_gerencia == '' or avaliacao_gerencia is None):

        return 'Aprovado Sindicância'



    # 9. Aprovado RH (aguardando sindicância)

    if avaliacao_rh == 'aprovado' and (sindicancia == '' or sindicancia is None):

        return 'Aprovado RH'



    # 10. Não Avaliado (tudo em branco ou não se encaixa nas regras)

    return 'Não Avaliado'



def ticket_to_dict(ticket):

    return {

        "id": ticket.id,

        "name": ticket.name,

        "ticket_number": ticket.ticket_number,

        "cpf": ticket.cpf,

        "dp_start_time": ticket.dp_start_time.strftime('%d/%m/%Y %H:%M:%S') if ticket.dp_start_time else '',

        "dp_process_start_time": ticket.dp_process_start_time.strftime('%d/%m/%Y %H:%M:%S') if ticket.dp_process_start_time else '',

        "dp_completed_time": ticket.dp_completed_time.strftime('%d/%m/%Y %H:%M:%S') if ticket.dp_completed_time else '',

        "dp_process_by": ticket.dp_process_by,  # Nome do operador, se quiser

        # Coloque outros campos se precisar

    }



def sincronizar_banco(cursor, db):

    """

    Sincroniza os campos 'situacao' e 'recrutador' da tabela registration_form,

    corrigindo apenas quando há divergência com a lógica de negócio.

    Se recrutador não existe mesmo na ficha, mantém 'Não definido'.

    """

    cursor.execute("""

        SELECT id, recrutador, avaliacao_rh, avaliacao_gerencia, sindicancia, admitido, situacao

        FROM registration_form

    """)

    registros = cursor.fetchall()

    total_corrigidos = 0



    for row in registros:

        id_ficha, recrutador, avaliacao_rh, avaliacao_gerencia, sindicancia, admitido, situacao_atual = row



        # Determina a situação correta conforme as regras

        dados = {

            "avaliacao_rh": avaliacao_rh,

            "avaliacao_gerencia": avaliacao_gerencia,

            "sindicancia": sindicancia,

            "admitido": admitido

        }

        situacao_certa = determinar_situacao(dados)



        atualizar = False

        campos_update = []

        valores_update = []



        # Atualiza situação se necessário

        if situacao_atual != situacao_certa:

            campos_update.append("situacao = ?")

            valores_update.append(situacao_certa)

            atualizar = True



        # Atualiza recrutador só se for vazio, NULL ou "Não definido", mas apenas para garantir o texto

        if not recrutador or str(recrutador).strip() == "" or str(recrutador).strip().lower() == "não definido":

            # Deixa como "Não definido"

            if str(recrutador).strip().lower() != "não definido":

                campos_update.append("recrutador = ?")

                valores_update.append("Não definido")

                atualizar = True



        if atualizar:

            valores_update.append(id_ficha)

            update_query = f"UPDATE registration_form SET {', '.join(campos_update)} WHERE id = ?"

            cursor.execute(update_query, valores_update)

            total_corrigidos += 1



    db.commit()

    print(f"[SINCRONIZADOR BANCO] {total_corrigidos} registros corrigidos/sincronizados.")







def atualizar_situacao_em_massa(cursor, db):

    situacao_map = [

        # 1. Admitido (maior prioridade)

        {

            'situacao': 'Admitido',

            'conditions': "LOWER(admitido) = 'sim'"

        },

        # 2. Reprovado RH

        {

            'situacao': 'Reprovado RH',

            'conditions': "LOWER(avaliacao_rh) = 'reprovado'"

        },

        # 3. Reprovado Sindicância

        {

            'situacao': 'Reprovado Sindicância',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND LOWER(sindicancia) = 'reprovado'"

        },

        # 4. Reprovado Gerência

        {

            'situacao': 'Reprovado Gerência',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND LOWER(sindicancia) = 'aprovado' AND LOWER(avaliacao_gerencia) = 'reprovado'"

        },

        # 5. Em Verificação

        {

            'situacao': 'Em Verificação',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND LOWER(sindicancia) = 'em verificação'"

        },

        # 6. Em Conversa

        {

            'situacao': 'Em Conversa',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND LOWER(sindicancia) = 'aprovado' AND LOWER(avaliacao_gerencia) = 'em conversa'"

        },

        # 7. Aprovado Gerência

        {

            'situacao': 'Aprovado Gerência',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND LOWER(sindicancia) = 'aprovado' AND LOWER(avaliacao_gerencia) = 'aprovado'"

        },

        # 8. Aprovado Sindicância (aguardando avaliação gerência)

        {

            'situacao': 'Aprovado Sindicância',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND LOWER(sindicancia) = 'aprovado' AND (avaliacao_gerencia IS NULL OR avaliacao_gerencia = '')"

        },

        # 9. Aprovado RH (aguardando sindicância)

        {

            'situacao': 'Aprovado RH',

            'conditions': "LOWER(avaliacao_rh) = 'aprovado' AND (sindicancia IS NULL OR sindicancia = '')"

        },

        # 10. Não Avaliado

        {

            'situacao': 'Não Avaliado',

            'conditions': "(avaliacao_rh IS NULL OR avaliacao_rh = '') AND (avaliacao_gerencia IS NULL OR avaliacao_gerencia = '') AND (sindicancia IS NULL OR sindicancia = '') AND (admitido IS NULL OR admitido = '')"

        }

    ]



    for item in situacao_map:

        situacao = item['situacao']

        conditions = item['conditions']

        update_query = f"""

            UPDATE registration_form

            SET situacao = ?

            WHERE {conditions} AND (situacao IS NULL OR situacao != ?)

        """

        cursor.execute(update_query, (situacao, situacao))

    db.commit()
def atualizar_situacao_em_massa(cursor, db):

    """

    Atualiza o campo 'situacao' na tabela 'registration_form' segundo a lógica de determinar_situacao().

    Primeiro trata os admitidos = 'Sim'; depois, para admitido = 'Não' (ou outros valores), aplica

    as regras de Reprovado RH, Reprovado Sindicância, etc., na ordem de prioridade.

    """



    start_time = time.time()

    total_atualizadas = 0



    try:

        # Define o nível de isolamento para evitar deadlocks

        cursor.execute("SET TRANSACTION ISOLATION LEVEL READ COMMITTED")

        cursor.execute("BEGIN TRANSACTION")



        # ---------------------------------------------------------

        # 1. Admitido (maior prioridade) → admitido = 'Sim'

        # ---------------------------------------------------------

        update_admitido = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) = 'sim'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_admitido, ('Admitido', 'Admitido'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Admitido'")

        # ---------------------------------------------------------

        # 1.5. Em processo de admissão (prioridade alta) → admitido = 'Em processo de admissão'

        # ---------------------------------------------------------

        update_em_processo = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) = 'em processo de admissão'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_em_processo, ('Em processo de admissão', 'Em processo de admissão'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Em processo de admissão'")



        # ---------------------------------------------------------

        # 2. Reprovado RH → admitido ≠ 'Sim' e admitido ≠ 'Em processo de admissão' e avaliacao_rh = 'Reprovado'

        # ---------------------------------------------------------

        update_reprovado_rh = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'reprovado'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_reprovado_rh, ('Reprovado RH', 'Reprovado RH'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Reprovado RH'")



        # ---------------------------------------------------------

        # 3. Reprovado Sindicância → admitido ≠ 'Sim' e avaliacao_rh = 'Aprovado' e sindicancia = 'Reprovado'

        # ---------------------------------------------------------

        update_reprovado_sindicancia = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND LOWER(sindicancia) = 'reprovado'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_reprovado_sindicancia, ('Reprovado Sindicância', 'Reprovado Sindicância'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Reprovado Sindicância'")



        # ---------------------------------------------------------

        # 4. Reprovado Gerência → admitido ≠ 'Sim' e avaliacao_rh = 'Aprovado' e sindicancia = 'Aprovado' e avaliacao_gerencia = 'Reprovado'

        # ---------------------------------------------------------

        update_reprovado_gerencia = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND LOWER(sindicancia) = 'aprovado'

              AND LOWER(avaliacao_gerencia) = 'reprovado'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_reprovado_gerencia, ('Reprovado Gerência', 'Reprovado Gerência'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Reprovado Gerência'")



        # ---------------------------------------------------------

        # 5. Em Verificação → admitido ≠ 'Sim' e avaliacao_rh = 'Aprovado' e sindicancia = 'Em Verificação'

        # ---------------------------------------------------------

        update_em_verificacao = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND LOWER(sindicancia) = 'em verificação'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_em_verificacao, ('Em Verificação', 'Em Verificação'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Em Verificação'")



        # ---------------------------------------------------------

        # 6. Em Conversa → admitido ≠ 'Sim' e avaliacao_rh = 'Aprovado' e sindicancia = 'Aprovado' e avaliacao_gerencia = 'Em Conversa'

        # ---------------------------------------------------------

        update_em_conversa = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND LOWER(sindicancia) = 'aprovado'

              AND LOWER(avaliacao_gerencia) = 'em conversa'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_em_conversa, ('Em Conversa', 'Em Conversa'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Em Conversa'")



        # ---------------------------------------------------------

        # 7. Aprovado Gerência → admitido ≠ 'Sim', avaliacao_rh = 'Aprovado', sindicancia = 'Aprovado', avaliacao_gerencia = 'Aprovado'

        # ---------------------------------------------------------

        update_aprovado_gerencia = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND LOWER(sindicancia) = 'aprovado'

              AND LOWER(avaliacao_gerencia) = 'aprovado'

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_aprovado_gerencia, ('Aprovado Gerência', 'Aprovado Gerência'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Aprovado Gerência'")



        # ---------------------------------------------------------

        # 8. Aprovado Sindicância → admitido ≠ 'Sim', avaliacao_rh = 'Aprovado', sindicancia = 'Aprovado', e (avaliacao_gerencia é NULL ou '')

        # ---------------------------------------------------------

        update_aprovado_sindicancia = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND LOWER(sindicancia) = 'aprovado'

              AND (avaliacao_gerencia IS NULL OR avaliacao_gerencia = '')

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_aprovado_sindicancia, ('Aprovado Sindicância', 'Aprovado Sindicância'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Aprovado Sindicância'")



        # ---------------------------------------------------------

        # 9. Aprovado RH → admitido ≠ 'Sim', avaliacao_rh = 'Aprovado', e (sindicancia é NULL ou '')

        # ---------------------------------------------------------

        update_aprovado_rh = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND LOWER(avaliacao_rh) = 'aprovado'

              AND (sindicancia IS NULL OR sindicancia = '')

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_aprovado_rh, ('Aprovado RH', 'Aprovado RH'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Aprovado RH'")



        # ---------------------------------------------------------

        # 10. Não Avaliado → admitido ≠ 'Sim' e nenhum dos campos de avaliação está preenchido

        # ---------------------------------------------------------

        update_nao_avaliado = """

            UPDATE registration_form WITH (ROWLOCK)

            SET situacao = ?

            WHERE LOWER(admitido) <> 'sim'
              AND LOWER(admitido) <> 'em processo de admissão'
              AND ( (avaliacao_rh IS NULL OR avaliacao_rh = '')

                    AND (avaliacao_gerencia IS NULL OR avaliacao_gerencia = '')

                    AND (sindicancia IS NULL OR sindicancia = '') )

              AND (situacao IS NULL OR situacao <> ?)

        """

        cursor.execute(update_nao_avaliado, ('Não Avaliado', 'Não Avaliado'))

        total_atualizadas += cursor.rowcount

        print(f"Atualizou {cursor.rowcount} registros para situação 'Não Avaliado'")



        # ---------------------------------------------------------

        # Commit geral

        # ---------------------------------------------------------

        db.commit()



    except Exception as e:

        db.rollback()

        print(f"Erro ao atualizar situações: {e}")

        raise



    elapsed_time = time.time() - start_time

    print(f"[OK] Atualizacao de situacao concluida em {elapsed_time:.2f} segundos – {total_atualizadas} registros alterados.")















# Função para obter a quantidade de tickets por categoria para SQL Server

def get_ticket_counts_by_category(db, start_date, end_date, category='ALL'):

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']

    cursor = db.cursor()

    

    if category != 'ALL':

        cursor.execute("SELECT COUNT(*) FROM tickets WHERE category = ? AND created_at BETWEEN ? AND ?", 

                       (category, start_date, end_date))

        counts = {category: cursor.fetchone()[0]}

    else:

        counts = {}

        for cat in categories:

            cursor.execute("SELECT COUNT(*) FROM tickets WHERE category = ? AND created_at BETWEEN ? AND ?", 

                           (cat, start_date, end_date))

            counts[cat] = cursor.fetchone()[0]



    cursor.close()

    return counts



# Função para calcular o desempenho dos recrutadores para SQL Server

def get_recruiter_performance(db, start_date, end_date):

    cursor = db.cursor()

    

    # Obter a lista de recrutadores distintos

    cursor.execute("SELECT DISTINCT recruiter FROM tickets")

    recruiters = cursor.fetchall()

    

    # Calcular o número de tickets para cada recrutador no intervalo de datas

    performance = {}

    for recruiter in recruiters:

        recruiter_name = recruiter[0]

        cursor.execute("SELECT COUNT(*) FROM tickets WHERE recruiter = ? AND created_at BETWEEN ? AND ?", 

                       (recruiter_name, start_date, end_date))

        performance[recruiter_name] = cursor.fetchone()[0]

    

    cursor.close()

    return performance



# Função para calcular o tempo médio de espera e atendimento por categoria no SQL Server

def get_average_times_by_category(db, start_date, end_date, time_type):

    query = f'SELECT AVG({time_type}) FROM tickets WHERE created_at BETWEEN ? AND ? AND category = ?'

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']

    

    cursor = db.cursor()

    average_times = {}

    for category in categories:

        cursor.execute(query, (start_date, end_date, category))

        average_times[category] = cursor.fetchone()[0] or 0

    

    cursor.close()

    return average_times



# Função para comparar tickets emitidos e concluídos por semana no SQL Server

def get_tickets_comparison(db, start_date, end_date):

    # Consulta para tickets emitidos por semana

    query_issued = '''

        SELECT DATEPART(wk, created_at) as week, COUNT(*) 

        FROM tickets 

        WHERE created_at BETWEEN ? AND ? 

        GROUP BY DATEPART(wk, created_at)

    '''

    

    # Consulta para tickets concluídos por semana

    query_completed = '''

        SELECT DATEPART(wk, created_at) as week, COUNT(*) 

        FROM tickets 

        WHERE status = 'CONCLUIDO' AND created_at BETWEEN ? AND ? 

        GROUP BY DATEPART(wk, created_at)

    '''



    cursor = db.cursor()

    cursor.execute(query_issued, (start_date, end_date))

    tickets_issued = cursor.fetchall()



    cursor.execute(query_completed, (start_date, end_date))

    tickets_completed = cursor.fetchall()



    cursor.close()



    return {

        "tickets_issued": [row[1] for row in tickets_issued],

        "tickets_completed": [row[1] for row in tickets_completed]

    }





# Função auxiliar para garantir que nenhum valor no dicionário seja None

def safe_dict(data):

    if isinstance(data, dict):

        return {str(k) if k is not None else 'undefined': (v if v is not None else 0) for k, v in data.items()}

    return data



# Decorador para verificar se o usuário é admin

def admin_required(f):

    @wraps(f)

    def decorated_function(*args, **kwargs):

        if not current_user.is_authenticated or not current_user.is_admin:

            # Se o usuário não é admin ou não está autenticado, retorna erro 403 (Proibido)

            abort(403)

        return f(*args, **kwargs)

    return decorated_function



def save_registration_form(ticket_data):

    db = get_sql_server_connection()

    cursor = db.cursor()

    

    # Mapeamento dos dados para o formato correto do registration_form no SQL Server

    cursor.execute('''

        INSERT INTO registration_form (

            nome_completo, cpf, cep, endereco, numero, complemento, bairro, cidade, telefone, avaliacao_rh, admitido

        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)

    ''', (

        ticket_data['name'],

        ticket_data['cpf'],

        ticket_data['cep'],

        ticket_data['rua'],

        ticket_data['numero'],

        ticket_data['complemento'],

        ticket_data['bairro'],

        ticket_data['cidade'],

        ticket_data['telefones'],

        'APROVADO',  # Valor especificado

        'SIM'        # Valor especificado

    ))

    

    db.commit()

    cursor.close()



@login_manager.user_loader

def load_user(user_id):

    db = get_sql_server_connection()

    cursor = db.cursor()

    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))

    user_row = cursor.fetchone()

    

    if user_row:

        # Acessando os valores pelo índice da tupla no SQL Server

        return User(user_row[0], user_row[1], user_row[2], user_row[3], user_row[4], user_row[5])

    

    cursor.close()

    return None



class User(UserMixin):

    def __init__(self, id, username, name, email, password, is_admin):

        self.id = id

        self.username = username

        self.name = name  # Aqui é onde o nome é atribuído

        self.email = email

        self.password = password

        self.is_admin = is_admin





    def is_active(self):

        return True



    def is_authenticated(self):

        return True



    def is_anonymous(self):

        return False



    def get_id(self):

        return str(self.id)





@app.route('/')

def index():

    if current_user.is_authenticated:

        return redirect(url_for('home'))

    else:

        return redirect(url_for('login'))



@app.route('/home')

@login_required

def home():

    return render_template('home.html')
def login():

    if current_user.is_authenticated:

        return redirect(url_for('home'))



    if request.method == 'POST':

        username = request.form.get('username')

        password = request.form.get('password')



        db = get_sql_server_connection()

        cursor = db.cursor()



        cursor.execute('SELECT id, username, name, email, password, is_admin FROM users WHERE username = ?', (username,))

        result = cursor.fetchone()



        if result:

            user_id, db_username, name, email, db_password, is_admin = result

            cursor.close()

            db.close()



            # Verificar a senha com tratamento para diferentes formatos de hash

            try:

                senha_valida = False

                

                # Verificar se é um hash pbkdf2_sha256

                if db_password.startswith('$pbkdf2-sha256$'):

                    senha_valida = pbkdf2_sha256.verify(password, db_password)

                # Verificar se é um hash gerado com werkzeug (Flask padrão)

                elif db_password.startswith('pbkdf2:sha256:'):

                    senha_valida = check_password_hash(db_password, password)

                # Caso seja uma senha em texto puro (não recomendado, mas para compatibilidade)

                elif db_password == password:

                    senha_valida = True

                    # Avisar no log que precisa atualizar o hash

                    print(f"ATENÇÃO: Senha em texto puro detectada para o usuário {username}. Recomendado atualizar para hash seguro.")

                

                if senha_valida:

                    user = User(user_id, db_username, name, email, db_password, is_admin)

                    login_user(user, remember=True)

                    

                    # Registrar o login

                    log_login(user_id, username)

                    

                    # Verificar se existem fichas perdidas para recuperação

                    check_for_lost_forms(user_id)

                    

                    # Redirecionar para a página solicitada originalmente ou home

                    next_page = request.args.get('next')

                    return redirect(next_page or url_for('home'))

                else:

                    flash('Login inválido. Verifique seu nome de usuário e senha.', 'danger')

            except Exception as e:

                print(f"Erro na verificação de senha: {str(e)}")

                flash('Erro na autenticação. Por favor, contacte o administrador.', 'danger')

        else:

            cursor.close()

            db.close()

            flash('Login inválido. Verifique seu nome de usuário e senha.', 'danger')



    return render_template('login.html')



# Função para registrar o login do usuário

def log_login(user_id, username):

    try:

        db = get_sql_server_connection()

        cursor = db.cursor()

        

        # Registrar o login no log de atividades

        cursor.execute('''

            INSERT INTO user_logs (user_id, action, details, timestamp)

            VALUES (?, ?, ?, CURRENT_TIMESTAMP)

        ''', (user_id, 'LOGIN', f"Login bem-sucedido para o usuário {username}"))

        

        db.commit()

        cursor.close()

        db.close()

    except Exception as e:

        print(f"Erro ao registrar log de login: {e}")



# Função para verificar se existem fichas perdidas que precisam ser recuperadas

def check_for_lost_forms(user_id):

    try:

        # Verificar se há arquivos de backup ou erro que podem ser recuperados

        error_dir = os.path.join('static', 'errors')

        backup_dir = os.path.join('static', 'backups')

        

        # Não fazer nada se os diretórios não existirem

        if not os.path.exists(error_dir) and not os.path.exists(backup_dir):

            return

            

        error_files = []

        backup_files = []

        

        # Contar arquivos de erro não processados

        if os.path.exists(error_dir):

            error_files = [f for f in os.listdir(error_dir) if f.endswith('.json') and not f.endswith('.processed')]

            

        # Contar arquivos de backup não processados

        if os.path.exists(backup_dir):

            backup_files = [f for f in os.listdir(backup_dir) if f.endswith('.json') and not f.endswith('.processed')]

            

        total_lost = len(error_files) + len(backup_files)

        

        if total_lost > 0:

            # Se houver fichas perdidas, exibir mensagem para o usuário

            message = f"Atenção: Existem {total_lost} fichas que podem precisar de recuperação. "

            

            # Verificar se o usuário é administrador

            db = get_sql_server_connection()

            cursor = db.cursor()

            cursor.execute('SELECT is_admin FROM users WHERE id = ?', (user_id,))

            is_admin = cursor.fetchone()[0]

            cursor.close()

            db.close()

            

            if is_admin:

                message += "Acesse a área de <a href='/admin/recovery'>Recuperação de Fichas</a> para verificá-las."

            else:

                message += "Peça ao administrador para verificar o sistema de recuperação de fichas."

            

            flash(Markup(message), 'warning')

    except Exception as e:

        print(f"Erro ao verificar fichas perdidas: {e}")



@app.route('/logout')

@login_required

def logout():

    logout_user()

    session.clear()

    flash('Desconectado com sucesso!', 'success')

    return redirect(url_for('login'))



@app.route('/account_settings', methods=['GET', 'POST'])

@login_required

def account_settings():

    db = get_sql_server_connection()  # Conecta ao banco de dados

    cursor = db.cursor()



    if request.method == 'POST':

        # Obtenha os dados do formulário

        name = request.form['name']

        email = request.form['email']

        password = request.form['password']

        confirm_password = request.form['confirm_password']



        # Verifica se as senhas coincidem

        if password and password != confirm_password:

            flash('As senhas não coincidem. Por favor, tente novamente.', 'danger')

            cursor.close()

            return redirect(url_for('account_settings'))



        # Atualiza o nome e e-mail no banco de dados

        cursor.execute('UPDATE users SET name = ?, email = ? WHERE id = ?', (name, email, current_user.id))



        # Atualiza a senha se for fornecida

        if password:

            hashed_password = generate_password_hash(password)

            cursor.execute('UPDATE users SET password = ? WHERE id = ?', (hashed_password, current_user.id))



        # Salva as alterações no banco de dados

        db.commit()

        cursor.close()

        flash('Configurações atualizadas com sucesso!', 'success')

        return redirect(url_for('account_settings'))



    cursor.close()

    return render_template('account_settings.html')



@app.route('/admin/dashboard')

@login_required

@admin_required

def admin_dashboard():

    db = get_sql_server_connection()

    cursor = db.cursor()



    cursor.execute('SELECT * FROM users')

    users = cursor.fetchall()



    cursor.close()

    db.close()



    return render_template('admin_dashboard.html', users=users)



def gerar_excel(dados, data_inicio, data_fim, tipo='resumo_geral', logo_path='static/logo.png'):

    wb = Workbook()

    ws = wb.active

    ws.title = "Relatório"



    # Estilos

    header_font = Font(bold=True, color="FFFFFF", name="Poppins", size=12)

    title_font = Font(bold=True, size=14, name="Poppins")

    cell_font = Font(name="Poppins", size=11)

    align_center = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    thin_border = Border(

        left=Side(border_style="thin"),

        right=Side(border_style="thin"),

        top=Side(border_style="thin"),

        bottom=Side(border_style="thin")

    )



    # Logo

    try:

        img = Image(logo_path)

        img.width = 160

        img.height = 80

        ws.add_image(img, "A1")

    except Exception as e:

        print("Erro ao adicionar logo:", e)



    # Título

    titulo = f"Relatório: {tipo.replace('_', ' ').title()} ({data_inicio} a {data_fim})"

    ws.merge_cells('A5:F5')

    ws['A5'] = titulo

    ws['A5'].font = title_font

    ws['A5'].alignment = align_center



    # Cabeçalho e dados

    colunas = []

    linhas = []



    if tipo == 'resumo_geral':

        colunas = ['Mês', 'Ano', 'Recrutador', 'Total', 'Aprovados', 'Reprovados']

        linhas = [[row.Mes, row.Ano, row.recrutador, row.Total_Candidatos, row.Total_Aprovados, row.Total_Reprovados] for row in dados]



    elif tipo == 'por_atendente':

        colunas = ['Guichê', 'Total Atendimentos', 'Categoria']

        linhas = [[row.guiche, row.Total_Atendimentos, row.category] for row in dados]



    elif tipo == 'entrevistas_agendamentos':

        colunas = ['Mês', 'Ano', 'Categoria', 'Quantidade']

        linhas = [[row.Mes, row.Ano, row.category, row.Quantidade] for row in dados]



    elif tipo == 'detalhado':

        colunas = ['ID', 'Nome', 'Email', 'Telefone', 'Situação', 'Criado em', 'Recrutador']

        linhas = [[row.id, row.name, row.email, row.telefone, row.situacao, row.created_at.strftime('%d/%m/%Y %H:%M'), row.recrutador] for row in dados]



    # Inserir cabeçalhos

    ws.append(colunas)

    for col_num, col_cell in enumerate(ws[7], 1):  # linha 7

        col_cell.font = header_font

        col_cell.fill = header_fill

        col_cell.alignment = align_center

        col_cell.border = thin_border

        ws.column_dimensions[col_cell.column_letter].width = 20



    # Inserir linhas

    for linha in linhas:

        ws.append(linha)

        for cell in ws[ws.max_row]:

            cell.alignment = align_center

            cell.font = cell_font

            cell.border = thin_border



    # Ajuste de altura da linha do título

    ws.row_dimensions[5].height = 30



    # Exportar para memória

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output





@app.route('/relatorio_excel', methods=['POST'])

def gerar_relatorio_excel():

    data_inicio = request.form['data_inicio']

    data_fim = request.form['data_fim']

    tipo = request.form.get('tipo', 'resumo_geral')



    conn = get_sql_server_connection()

    cursor = conn.cursor()



    if tipo == 'resumo_geral':

        query = """

        SELECT 

            MONTH(created_at) AS Mes,

            YEAR(created_at) AS Ano,

            COUNT(*) AS Total_Candidatos,

            SUM(CASE WHEN situacao = 'Aprovado' THEN 1 ELSE 0 END) AS Total_Aprovados,

            SUM(CASE WHEN situacao = 'Reprovado' THEN 1 ELSE 0 END) AS Total_Reprovados,

            recrutador

        FROM registration_form

        WHERE created_at BETWEEN ? AND ?

        GROUP BY MONTH(created_at), YEAR(created_at), recrutador

        ORDER BY Ano, Mes;

        """



    elif tipo == 'por_atendente':

        query = """

        SELECT 

            guiche,

            COUNT(*) AS Total_Atendimentos,

            category

        FROM tickets

        WHERE created_at BETWEEN ? AND ?

            AND category IN ('Entrevista', 'Agendado')

        GROUP BY guiche, category

        ORDER BY guiche;

        """



    elif tipo == 'entrevistas_agendamentos':

        query = """

        SELECT 

            MONTH(created_at) AS Mes,

            YEAR(created_at) AS Ano,

            category,

            COUNT(*) AS Quantidade

        FROM tickets

        WHERE created_at BETWEEN ? AND ?

            AND category IN ('Entrevista', 'Agendado')

        GROUP BY MONTH(created_at), YEAR(created_at), category

        ORDER BY Ano, Mes;

        """



    elif tipo == 'detalhado':

        query = """

        SELECT

            id, name, email, telefone, situacao, created_at, recrutador

        FROM registration_form

        WHERE created_at BETWEEN ? AND ?

        ORDER BY created_at DESC;

        """



    else:

        flash("Tipo de relatório inválido para exportação!", "danger")

        return redirect(url_for('admin_dashboard'))



    cursor.execute(query, (data_inicio, data_fim))

    dados = cursor.fetchall()



    # Gera o Excel com base no tipo

    arquivo_excel = gerar_excel(dados, data_inicio, data_fim, tipo)



    return send_file(

        arquivo_excel,

        as_attachment=True,

        download_name=f"relatorio_{tipo}_{data_inicio}_a_{data_fim}.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )





@app.route('/add_user', methods=['POST'])

@login_required

@admin_required

def add_user():

    db = get_sql_server_connection()

    cursor = db.cursor()



    try:

        # Verifique se o usuário atual tem permissões de administrador

        if not current_user.is_admin:

            flash('Você não tem permissão para realizar esta ação.', 'danger')

            return redirect(url_for('home'))



        # Obtenha os dados do formulário

        username = request.form.get('username')

        name = request.form.get('name')

        email = request.form.get('email')

        password = generate_password_hash(request.form.get('password'))

        is_admin = int('is_admin' in request.form)  # Converte booleano para int



        # Determina o próximo ID (id + 1)

        cursor.execute('SELECT ISNULL(MAX(id), 0) + 1 FROM users')

        next_id = cursor.fetchone()[0]



        # Adiciona o novo usuário ao banco de dados com o ID incrementado

        cursor.execute('''

            INSERT INTO users (id, username, name, email, password, is_admin) 

            VALUES (?, ?, ?, ?, ?, ?)

        ''', (next_id, username, name, email, password, is_admin))

        

        # Confirma a transação

        db.commit()



        flash('Usuário adicionado com sucesso!', 'success')

    except Exception as e:

        db.rollback()  # Reverte em caso de erro

        flash(f'Erro ao adicionar usuário: {e}', 'danger')

    finally:

        cursor.close()



    return redirect(url_for('manage_users'))







@app.template_filter('getattr')

def getattr_filter(obj, attr):

    try:

        return getattr(obj, attr)

    except AttributeError:

        return 'Não informado'
@app.route('/visualizar_ficha/<int:id>', methods=['GET'])
def visualizar_ficha(id):

    candidato = get_candidato_by_id(id)  # Substitua pela lógica de busca do candidato

    if not candidato:

        return "Candidato não encontrado", 404

    return render_template('partials/_ficha_visualizacao.html', candidato=candidato)
@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):

    db = get_sql_server_connection()

    cursor = db.cursor()



    if not current_user.is_admin:

        cursor.close()

        return redirect(url_for('home'))



    # Busca o usuário com o id fornecido diretamente no banco de dados

    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))

    user = cursor.fetchone()



    if not user:

        cursor.close()

        flash('Usuário não encontrado.', 'danger')

        return redirect(url_for('manage_users'))



    if request.method == 'POST':

        # Atualiza os dados do usuário

        username = request.form['username']

        name = request.form['name']

        email = request.form['email']

        is_admin = int('is_admin' in request.form)  # Converte booleano para int



        # Atualiza os dados no banco

        cursor.execute('UPDATE users SET username = ?, name = ?, email = ?, is_admin = ? WHERE id = ?',

                       (username, name, email, is_admin, user_id))

        db.commit()

        cursor.close()



        flash('Usuário atualizado com sucesso!', 'success')

        return redirect(url_for('manage_users'))



    cursor.close()

    return render_template('edit_user.html', user=user)



@app.route('/view_logs')

@login_required

def view_logs():

    if not current_user.is_admin:

        flash("Acesso negado. Apenas administradores podem acessar os logs.", "danger")

        return redirect(url_for('home'))



    db = get_sql_server_connection()

    cursor = db.cursor()

    cursor.execute('SELECT * FROM user_logs')

    logs = cursor.fetchall()

    cursor.close()



    return render_template('view_logs.html', logs=logs)



@app.route('/delete_user/<int:user_id>', methods=['POST'])

@login_required

def delete_user(user_id):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Verifica se o usuário existe

    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))

    user = cursor.fetchone()



    if not user:

        cursor.close()

        flash('Usuário não encontrado.', 'danger')

        return redirect(url_for('manage_users'))



    # Deleta o usuário do banco de dados

    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))

    db.commit()

    cursor.close()



    flash('Usuário excluído com sucesso!', 'success')

    return redirect(url_for('manage_users'))



@app.route('/painel')

@login_required

def painel():

    db = get_sql_server_connection()

    today = datetime.now().date()



    # Tickets em espera do dia, ordenados pela data e hora de criação

    waiting_tickets = db.execute(

        'SELECT * FROM dbo.tickets WHERE status = ? AND CONVERT(DATE, created_at) = ? ORDER BY created_at',

        ('ESPERA', today)

    ).fetchall()



    # Tickets chamados do dia, ordenados pela data e hora de chamado

    called_tickets = db.execute(

        'SELECT * FROM dbo.tickets WHERE status = ? AND CONVERT(DATE, called_at) = ? ORDER BY called_at',

        ('CHAMADO', today)

    ).fetchall()



    # Tickets concluídos do dia, ordenados pela data e hora de conclusão

    concluded_tickets = db.execute(

        'SELECT * FROM dbo.tickets WHERE status = ? AND CONVERT(DATE, concluded_at) = ? ORDER BY concluded_at',

        ('CONCLUIDO', today)

    ).fetchall()



    # Resumo dos tickets do dia

    total_tickets = db.execute(

        'SELECT COUNT(*) FROM dbo.tickets WHERE CONVERT(DATE, created_at) = ?', 

        (today,)

    ).fetchone()[0]

    total_concluido = db.execute(

        'SELECT COUNT(*) FROM dbo.tickets WHERE status = ? AND CONVERT(DATE, concluded_at) = ?', 

        ('CONCLUIDO', today)

    ).fetchone()[0]



    # Contagem de tickets em espera por categoria

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']

    counts = {

        category: db.execute(

            'SELECT COUNT(*) FROM dbo.tickets WHERE category = ? AND status = ? AND CONVERT(DATE, created_at) = ?', 

            (category, 'ESPERA', today)

        ).fetchone()[0]

        for category in categories

    }



    # Incluindo campos adicionais nos tickets de espera, chamados e concluídos

    def convert_ticket_to_dict(ticket):

        return {

            'ticket_number': ticket.ticket_number,

            'name': ticket.name,

            'category': ticket.category,

            'priority': ticket.priority,

            'id': ticket.id,

            'created_at': ticket.created_at,

            'called_at': ticket.called_at if ticket.called_at else 'Não chamado',

            'concluded_at': ticket.concluded_at if ticket.concluded_at else 'Não concluído',

            'stage': ticket.stage,

            'guiche': ticket.guiche,

            'especificacao': getattr(ticket, 'especificacao', ''),

            'agendado_com': ticket.recruiter if ticket.category == 'Agendado' else None

        }



    waiting_tickets = [convert_ticket_to_dict(ticket) for ticket in waiting_tickets]

    called_tickets = [convert_ticket_to_dict(ticket) for ticket in called_tickets]

    concluded_tickets = [convert_ticket_to_dict(ticket) for ticket in concluded_tickets]



    # Dados do formulário

    cpf = request.args.get('cpf')

    form_data = {

        'name': '',

        'cep': '',

        'rua': '',

        'numero': '',

        'complemento': '',

        'bairro': '',

        'cidade': '',

        'telefones': ''

    }



    if cpf:

        # Busca dados internos

        internal_data = db.execute('SELECT * FROM dbo.registration_form WHERE cpf = ?', (cpf,)).fetchone()

        if internal_data:

            form_data.update({

                'name': internal_data.nome_completo,

                'cep': internal_data.cep,

                'rua': internal_data.endereco,

                'numero': internal_data.numero,

                'complemento': internal_data.complemento,

                'bairro': internal_data.bairro,

                'cidade': internal_data.cidade,

                'telefones': internal_data.telefones

            })



        # Conexão externa para busca adicional

        conn = get_jbc_connection()

        cursor = conn.cursor()

        cursor.execute('''SELECT pe_nome, pe_cep, pe_logradouro_end, pe_numero_end, pe_complemento_end, 

                   pe_bairro_end, pe_cidade_end 

            FROM dbo.fo_pessoa 

            WHERE pe_cpf = ?''', cpf)

        external_data = cursor.fetchone()

        conn.close()



        if external_data:

            form_data.update({

                'name': external_data[0] or form_data['name'],

                'cep': external_data[1] or form_data['cep'],

                'rua': external_data[2] or form_data['rua'],

                'numero': external_data[3] or form_data['numero'],

                'complemento': external_data[4] or form_data['complemento'],

                'bairro': external_data[5] or form_data['bairro'],

                'cidade': external_data[6] or form_data['cidade']

            })



    return render_template(

        'painel.html', 

        waiting_tickets=waiting_tickets, 

        called_tickets=called_tickets, 

        concluded_tickets=concluded_tickets, 

        total_tickets=total_tickets, 

        total_concluido=total_concluido, 

        counts=counts,

        form_data=form_data

    )



@app.route('/get_ticket/<int:ticket_id>', methods=['GET'])

@login_required

def get_ticket(ticket_id):

    # Conectando ao banco de dados SQL Server

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Buscando o ticket pelo ID

    cursor.execute('SELECT * FROM dbo.tickets WHERE id = ?', (ticket_id,))

    ticket = cursor.fetchone()

    cursor.close()



    # Verificando se o ticket foi encontrado

    if ticket:

        # Estrutura de dados com os valores do ticket para enviar ao front-end

        ticket_data = {

            'id': ticket[0],

            'name': ticket[1],

            'category': ticket[2],

            'priority': ticket[10],

            'cpf': ticket[12],

            'cep': ticket[13],

            'rua': ticket[14],

            'numero': ticket[15],

            'complemento': ticket[16],

            'bairro': ticket[17],

            'cidade': ticket[18],

            'telefones': ticket[19],

            'especificacao': ticket[21] if len(ticket) > 21 else None,

            'data_nasc': ticket[23],

            'recruiter': ticket[22] if len(ticket) > 22 else None  # Adicionando o recrutador

            

        }



        # Enviando os dados do ticket para o front-end em formato JSON

        return jsonify(ticket_data)

    else:

        # Se o ticket não for encontrado, retorna um erro 404

        return jsonify({'error': 'Ticket não encontrado'}), 404





@app.route('/update_ticket', methods=['POST'])

@login_required

def update_ticket():

    # Captura os dados do formulário

    ticket_id = request.form.get('ticket_id')

    name = request.form.get('name')

    category = request.form.get('category')

    priority = request.form.get('priority')

    cpf = request.form.get('cpf')

    cep = request.form.get('cep')

    rua = request.form.get('rua')

    numero = request.form.get('numero')

    complemento = request.form.get('complemento')

    bairro = request.form.get('bairro')

    cidade = request.form.get('cidade')

    telefones = request.form.get('telefones')

    especificacao = request.form.get('especificacao', '')

    data_nasc = request.form.get('data_nasc')  # Novo campo





    recruiter = None  # Inicializa recrutador como None

    if category == "Agendado":

        recruiter = request.form.get('recruiter')  # Captura o recrutador se a categoria for Agendado



    db = get_sql_server_connection()  # Conexão ao banco de dados SQL Server

    cursor = db.cursor()



    # Verifica se o ticket existe

    cursor.execute('SELECT * FROM dbo.tickets WHERE id = ?', (ticket_id,))

    ticket = cursor.fetchone()



    if ticket:

        # Mapeia manualmente os índices das colunas retornadas pelo banco

        ticket_data = {

            'id': ticket[0],  # Substitua os índices de acordo com a ordem das colunas

            'name': ticket[1],

            'category': ticket[2],

            'ticket_number': ticket[3],

            'cpf': ticket[12],

            'cep': ticket[13],

            'rua': ticket[14],

            'numero': ticket[15],

            'complemento': ticket[16],

            'bairro': ticket[17],

            'cidade': ticket[18],

            'telefones': ticket[19],

            'priority': ticket[10],

            'especificacao': ticket[22],

            'data_nasc': ticket[23],  # Novo campo

            'recruiter': ticket[21]

        }



        # Verifica se a categoria foi alterada e gera um novo número de ticket

        if ticket_data['category'] != category:

            ticket_number = generate_ticket_number(category)  # Função que gera o número do ticket

        else:

            ticket_number = ticket_data['ticket_number']



        # Atualiza os dados do ticket, incluindo o recrutador, se for Agendado

        cursor.execute('''

            UPDATE dbo.tickets

            SET name = ?, category = ?, priority = ?, cpf = ?, cep = ?, rua = ?, numero = ?, complemento = ?, 

                bairro = ?, cidade = ?, telefones = ?, especificacao = ?, data_nasc = ?, ticket_number = ?, recruiter = ?

            WHERE id = ?

        ''', (name, category, priority, cpf, cep, rua, numero, complemento, bairro, cidade, telefones, especificacao, data_nasc, ticket_number, recruiter, ticket_id))



        db.commit()

        cursor.close()



        flash('Ticket atualizado com sucesso!', 'success')

    else:

        cursor.close()

        flash('Ticket não encontrado.', 'danger')



    return redirect(url_for('painel'))  # Certifique-se de que a rota 'painel' exista



@app.route('/export_pdf/<cpf>')

@login_required

def export_pdf(cpf):

    db = get_sql_server_connection()

    cursor = db.cursor()

    try:

        cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

        row = cursor.fetchone()

        if not row:

            flash("Candidato não encontrado.", "danger")

            return redirect(url_for('banco_rs'))



        columns = [column[0] for column in cursor.description]

        candidato = dict(zip(columns, row))



        # Data de nascimento para exibição

        if candidato.get('data_nasc'):

            try:

                candidato['data_nasc'] = candidato['data_nasc'].strftime('%d/%m/%Y')

            except AttributeError:

                candidato['data_nasc'] = "Data inválida"



        # Currículo: só aceitar PDF existente

        curriculo_path: Optional[Path] = None

        if candidato.get('curriculo'):

            p = Path('static') / 'uploads' / str(candidato['curriculo'])

            if p.suffix.lower() == '.pdf' and p.exists():

                curriculo_path = p

            else:

                flash("Currículo não é PDF ou não foi encontrado. Será ignorado no anexo.", "warning")



        # Render do HTML da ficha (mantém estilização do template atual)

        logo_url = 'https://jbconservadora.com.br/wp-content/uploads/2020/09/logo-final-jb.png'

        rendered_html = render_template('candidato_template.html', form_data=candidato, logo_url=logo_url)



        # Geração do PDF da ficha via backend disponível

        out_dir = Path('static') / 'temp'

        out_dir.mkdir(parents=True, exist_ok=True)

        ficha_pdf_path = out_dir / f'Ficha_{cpf}.pdf'



        try:

            backend = generate_pdf_from_html(rendered_html, ficha_pdf_path)

            logger.info("Ficha %s gerada com backend %s", cpf, backend)

        except RuntimeError as exc:

            logger.error("Falha ao gerar PDF da ficha %s: %s", cpf, exc, exc_info=True)

            flash(str(exc), "danger")

            return redirect(url_for('banco_rs'))



        if not ficha_pdf_path.exists() or ficha_pdf_path.stat().st_size == 0:

            flash("Erro ao gerar a ficha do candidato.", "danger")

            return redirect(url_for('banco_rs'))



        # Merge com currículo (se existir)

        merger = PdfMerger()

        try:

            with ficha_pdf_path.open('rb') as f:

                merger.append(f)

            if curriculo_path:

                try:

                    with curriculo_path.open('rb') as f:

                        PdfReader(f)  # valida PDF

                    with curriculo_path.open('rb') as f:

                        merger.append(f)

                except Exception as exc:  # noqa: BLE001

                    logger.warning("Falha ao anexar currículo %s: %s", curriculo_path, exc, exc_info=True)

                    flash("O currículo não pôde ser anexado ao PDF final.", "warning")



            final_pdf_path = out_dir / f'Candidato_{cpf}.pdf'

            with final_pdf_path.open('wb') as f:

                merger.write(f)

        finally:

            merger.close()



        if not final_pdf_path.exists() or final_pdf_path.stat().st_size == 0:

            flash("Erro ao mesclar os PDFs.", "danger")

            return redirect(url_for('banco_rs'))



        return send_file(

            str(final_pdf_path),

            mimetype='application/pdf',

            as_attachment=True,

            download_name=f'Candidato_{cpf}.pdf',

        )



    except Exception as exc:  # noqa: BLE001

        logger.error("Erro inesperado no export_pdf para %s: %s", cpf, exc, exc_info=True)

        flash(f"Erro ao gerar o PDF do candidato: {exc}", "danger")

        return redirect(url_for('banco_rs'))

    finally:

        cursor.close()

        db.close()



@app.route('/submit_form', methods=['POST'])

@login_required
def submit_form():

    db = None

    cursor = None

    cpf = None

    novo_id = None



    try:

        # Captura os dados antes mesmo de estabelecer conexão para casos de erro de conexão

        data = request.form

        curriculo = request.files.get('curriculo')

        

        # Extrai o CPF para logs de rastreamento

        cpf = data.get('cpf', '').upper().strip() if data.get('cpf') else 'CPF_NAO_INFORMADO'

        

        # Log inicial para rastreamento

        print(f"[{datetime.now()}] Iniciando salvamento para CPF: {cpf}")

        

        # Estabelece a conexão com o banco

        try:

            db = get_sql_server_connection()

            if not db:

                raise Exception("Falha ao conectar ao banco de dados SQL Server")

                

            cursor = db.cursor()

            if not cursor:

                raise Exception("Falha ao criar cursor para o banco de dados")

        except Exception as conn_error:

            print(f"[{datetime.now()}] ERRO DE CONEXÃO: {str(conn_error)}")

            # Salva em arquivo de log local para recuperação posterior

            _salvar_erro_local(cpf, "ERRO_CONEXAO", str(conn_error), data)

            return jsonify({"success": False, "message": "Erro de conexão com o banco de dados. Os dados foram salvos localmente e serão processados assim que possível."}), 500



        # Gerenciamento do arquivo de currículo

        curriculo_filename = None

        if curriculo and curriculo.filename:

            try:

                curriculo_filename = secure_filename(curriculo.filename)

                UPLOAD_DIR = os.path.join('static', 'uploads')

                if not os.path.exists(UPLOAD_DIR):

                    os.makedirs(UPLOAD_DIR)

                curriculo.save(os.path.join(UPLOAD_DIR, curriculo_filename))

                print(f"[{datetime.now()}] Currículo salvo: {curriculo_filename}")

            except Exception as upload_error:

                print(f"[{datetime.now()}] Erro ao salvar currículo: {str(upload_error)}")

                # Continue mesmo se o upload falhar



        # Captura os dados do formulário com validação rigorosa

        form_data = {

            'cpf': cpf,

            'nome_completo': data.get('name', '').upper().strip() if data.get('name') else '',

            'estado_civil': data.get('estado_civil', '').upper().strip() if data.get('estado_civil') else '',

            'cep': data.get('cep', '').strip(),

            'endereco': data.get('endereco', '').upper().strip() if data.get('endereco') else '',

            'numero': data.get('numero', '').strip(),

            'complemento': data.get('complemento', '').upper().strip() if data.get('complemento') else '',

            'bairro': data.get('bairro', '').upper().strip() if data.get('bairro') else '',

            'cidade': data.get('cidade', '').upper().strip() if data.get('cidade') else '',

            'telefone': data.get('telefone', '').strip(),

            'telefone_recado': data.get('telefone_recado', '').strip(),

            'email': data.get('email', '').lower().strip() if data.get('email') else '', 

            'estado_nasc': data.get('estado_nasc', '').upper().strip() if data.get('estado_nasc') else '',

            'cidade_nasc': data.get('cidade_nasc', '').upper().strip() if data.get('cidade_nasc') else '',

            'data_nasc': data.get('data_nasc', '').strip(),

            'idade': data.get('idade', '').strip(),

            'numero_filhos': data.get('numero_filhos', '').strip() or None,  # Ajuste para evitar valor padrão incorreto

            'fumante': data.get('fumante', '').upper().strip() if data.get('fumante') else '',

            'bebida': data.get('bebida', '').upper().strip() if data.get('bebida') else '',

            'alergia': data.get('alergia', '').upper().strip() if data.get('alergia') else '',

            'medicamento_constante': data.get('medicamento_constante', '').upper().strip() if data.get('medicamento_constante') else '',

            'qual_medicamento': data.get('qual_medicamento', '').upper().strip() if data.get('qual_medicamento') else '',

            'genero': data.get('genero', '').upper().strip() if data.get('genero') else '',

            'peso': data.get('peso', '').strip(),

            'cor_pele': data.get('cor_pele', '').upper().strip() if data.get('cor_pele') else '',

            'tatuagem': data.get('tatuagem', '').upper().strip() if data.get('tatuagem') else '',

            'perfil': data.get('perfil', '').upper().strip() if data.get('perfil') else '',

            'cargo_indicado': ','.join([item.upper().strip() for item in data.getlist('cargo_indicado')]) if data.getlist('cargo_indicado') else '',

            'cursos_realizados': data.get('cursos_realizados', '').upper().strip() if data.get('cursos_realizados') else '',

            'regioes_preferencia': ','.join([item.upper().strip() for item in data.getlist('regioes_preferencia')]) if data.getlist('regioes_preferencia') else '',

            'disponibilidade_horario': ','.join([item.upper().strip() for item in data.getlist('disponibilidade_horario')]) if data.getlist('disponibilidade_horario') else '',

            'empregos_informais': data.get('empregos_informais', '').upper().strip() if data.get('empregos_informais') else '',

            'avaliacao_rh': data.get('avaliacao_rh', '').strip(),

            'assinatura_rh': data.get('assinatura_rh', '').strip(),

            'avaliacao_gerencia': data.get('avaliacao_gerencia', '').strip(),

            'conhecimento_digitacao': data.get('conhecimento_digitacao', '').strip(),

            'assinatura_gerencia': data.get('assinatura_gerencia', '').strip(),

            'curriculo': curriculo_filename,

            'observacoes': data.get('observacoes', '').upper().strip() if data.get('observacoes') else '',

            'pcd': data.get('pcd', '').strip(),

            'escolaridade': data.get('escolaridade', '').upper().strip() if data.get('escolaridade') else '',

            'motivo_reprovacao_rh': data.get('motivo_reprovacao_rh', '').upper().strip() if data.get('motivo_reprovacao_rh') else '',

            'admitido': data.get('admitido', '').strip(),

            'empresa1': data.get('empresa1', '').upper().strip() if data.get('empresa1') else '',

            'cidade1': data.get('cidade1', '').upper().strip() if data.get('cidade1') else '',

            'funcao1': data.get('funcao1', '').upper().strip() if data.get('funcao1') else '',

            'data_admissao1': data.get('data_admissao1', '').strip(),

            'data_saida1': data.get('data_saida1', '').strip(),

            'motivo_saida1': data.get('motivo_saida1', '').upper().strip() if data.get('motivo_saida1') else '',

            'salario1': data.get('salario1', '').strip(),

            'empresa2': data.get('empresa2', '').upper().strip() if data.get('empresa2') else '',

            'cidade2': data.get('cidade2', '').upper().strip() if data.get('cidade2') else '',

            'funcao2': data.get('funcao2', '').upper().strip() if data.get('funcao2') else '',

            'data_admissao2': data.get('data_admissao2', '').strip(),

            'data_saida2': data.get('data_saida2', '').strip(),

            'motivo_saida2': data.get('motivo_saida2', '').upper().strip() if data.get('motivo_saida2') else '',

            'salario2': data.get('salario2', '').strip(),

            'empresa3': data.get('empresa3', '').upper().strip() if data.get('empresa3') else '',

            'cidade3': data.get('cidade3', '').upper().strip() if data.get('cidade3') else '',

            'funcao3': data.get('funcao3', '').upper().strip() if data.get('funcao3') else '',

            'data_admissao3': data.get('data_admissao3', '').strip(),

            'data_saida3': data.get('data_saida3', '').strip(),

            'motivo_saida3': data.get('motivo_saida3', '').upper().strip() if data.get('motivo_saida3') else '',

            'salario3': data.get('salario3', '').strip(),

            'atividades_empresa1': data.get('atividades_empresa1', '').upper().strip() if data.get('atividades_empresa1') else '',

            'atividades_empresa2': data.get('atividades_empresa2', '').upper().strip() if data.get('atividades_empresa2') else '',

            'atividades_empresa3': data.get('atividades_empresa3', '').upper().strip() if data.get('atividades_empresa3') else '',

            'tempo_permanencia1_anos': data.get('tempo_permanencia1_anos', '').strip(),

            'tempo_permanencia1_meses': data.get('tempo_permanencia1_meses', '').strip(),

            'tempo_permanencia1_dias': data.get('tempo_permanencia1_dias', '').strip(),

            'tempo_permanencia2_anos': data.get('tempo_permanencia2_anos', '').strip(),

            'tempo_permanencia2_meses': data.get('tempo_permanencia2_meses', '').strip(),

            'tempo_permanencia2_dias': data.get('tempo_permanencia2_dias', '').strip(),

            'tempo_permanencia3_anos': data.get('tempo_permanencia3_anos', '').strip(),

            'tempo_permanencia3_meses': data.get('tempo_permanencia3_meses', '').strip(),

            'tempo_permanencia3_dias': data.get('tempo_permanencia3_dias', '').strip()

        }



        # Adicionando lógica para salvar os checkboxes como booleanos

        # Mapeamento de valores para colunas booleanas

        checkbox_mappings = {

            'disponibilidade_horario': {

                '44H (HORARIO COMERCIAL)': 'horario_comercial',

                '12X36 DIA': 'horario_12x36_dia',

                '12X36 NOITE': 'horario_12x36_noite',

                'FEIRISTA': 'horario_feirista'

            },

            'cargo_indicado': {

                'MANUTENÇÃO': 'cargo_manutencao',

                'ADMINISTRATIVO': 'cargo_administrativo',

                'JOVEM APRENDIZ LIMPEZA': 'cargo_jovem_aprendiz_limpeza',

                'JOVEM APRENDIZ ADMINISTRATIVO': 'cargo_jovem_aprendiz_administrativo',

                'MANOBRISTA': 'cargo_manobrista',

                'RECEPCIONISTA': 'cargo_recepcionista',

                'PORTARIA': 'cargo_portaria',

                'ASG': 'cargo_asg',

                'ZELADORIA': 'cargo_zeladoria',

                'MENSAGEIRO': 'cargo_mensageiro',

                'COPEIRA': 'cargo_copeira',

                'SUPERVISOR PREDIAL': 'cargo_supervisor_predial'

            },

            'regioes_preferencia': {

                'REGIÃO BARREIRO': 'regiao_barreiro',

                'REGIÃO CENTRO SUL': 'regiao_centro_sul',

                'REGIÃO LESTE': 'regiao_leste',

                'REGIÃO NORDESTE': 'regiao_nordeste',

                'REGIÃO NOROESTE': 'regiao_noroeste',

                'REGIÃO NORTE': 'regiao_norte',

                'REGIÃO OESTE': 'regiao_oeste',

                'REGIÃO PAMPULHA': 'regiao_pampulha',

                'REGIÃO VENDA NOVA': 'regiao_venda_nova',

                'REGIÃO BETIM/CONTAGEM': 'regiao_betim_contagem',

                'REGIÃO NOVA LIMA': 'regiao_nova_lima',

                'OUTROS MUNICÍPIOS': 'regiao_outros_municipios'

            }

        }



        # Processar os checkboxes e salvar como booleanos

        for field, mapping in checkbox_mappings.items():

            selected_values = data.getlist(field)

            for value, column in mapping.items():

                form_data[column] = 1 if value in selected_values else 0



        # Determinação da situação

        if form_data.get('admitido') == 'Sim':

            form_data['situacao'] = 'Admitido'

        else:

            form_data['situacao'] = 'Não Avaliado'



        # Backup dos dados do formulário em arquivo local antes de tentar salvar no banco

        # Isso garante que mesmo em falha total, os dados possam ser recuperados

        _salvar_backup_formulario(form_data)

        print(f"[{datetime.now()}] Backup local criado para CPF: {cpf}")



        # Verificar colunas existentes na tabela para evitar erros

        try:

            cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'registration_form'")

            column_names = [row[0] for row in cursor.fetchall()]

            if not column_names:

                raise Exception("Erro ao obter colunas da tabela 'registration_form'")

        except Exception as col_error:

            print(f"[{datetime.now()}] ERRO AO OBTER COLUNAS: {str(col_error)}")

            _salvar_erro_local(cpf, "ERRO_COLUNAS", str(col_error), form_data)

            return jsonify({"success": False, "message": "Erro ao verificar estrutura da tabela. Os dados foram salvos localmente e serão processados assim que possível."}), 500

        

        # Filtrar o dicionário para incluir apenas colunas que existem na tabela

        form_data_filtered = {k: v for k, v in form_data.items() if k in column_names}

        

        # Verificar se o CPF já está cadastrado com retry para casos de problemas temporários

        for tentativa in range(3):  # Tenta até 3 vezes

            try:

                cursor.execute('SELECT id FROM registration_form WHERE cpf = ?', (cpf,))

                existing_entry = cursor.fetchone()

                break

            except Exception as select_error:

                print(f"[{datetime.now()}] ERRO NA VERIFICAÇÃO DE CPF (Tentativa {tentativa+1}/3): {str(select_error)}")

                if tentativa < 2:  # Se não for a última tentativa

                    time.sleep(1)  # Espera 1 segundo antes de tentar novamente

                else:

                    _salvar_erro_local(cpf, "ERRO_VERIFICACAO_CPF", str(select_error), form_data)

                    return jsonify({"success": False, "message": "Erro ao verificar existência de CPF. Os dados foram salvos localmente e serão processados assim que possível."}), 500



        if not existing_entry:

            # Remove `id` antes da inserção para permitir que o SQL Server gere automaticamente

            form_data_filtered.pop('id', None)



            # Remove `last_updated` se estiver na lista, pois será gerado pelo banco

            colunas_insert = list(form_data_filtered.keys())

            if 'last_updated' in colunas_insert:

                colunas_insert.remove('last_updated')



            # Log detalhado para depuração

            print(f"[{datetime.now()}] INSERÇÃO - Colunas: {colunas_insert}")

            

            # Prepara a query com os placeholders

            placeholders = ', '.join(['?' for _ in colunas_insert])

            valores = [form_data_filtered[col] for col in colunas_insert]

            

            # Log detalhado para depuração

            print(f"[{datetime.now()}] INSERÇÃO - Quantidade de valores: {len(valores)}")



            # Query de inserção

            insert_query = f'''

                INSERT INTO registration_form ({colunas_insert}, created_at, last_updated)

                VALUES ({placeholders}, GETDATE(), GETDATE())

            '''

            

            # Executa a inserção com retry para casos de falha temporária

            for tentativa in range(3):  # Tenta até 3 vezes

                try:

                    print(f"[{datetime.now()}] Tentativa {tentativa+1}/3 de execução da query de inserção")

                    cursor.execute(insert_query, valores)

                    

                    # Recupera o ID gerado automaticamente

                    cursor.execute("SELECT SCOPE_IDENTITY()")

                    novo_id = cursor.fetchone()[0]

                    

                    # Log para depuração

                    print(f"[{datetime.now()}] SUCESSO - Novo ID gerado: {novo_id}")

                    

                    # Teste de verificação para confirmar que o registro foi realmente inserido

                    cursor.execute('SELECT id FROM registration_form WHERE cpf = ?', (cpf,))

                    verify_entry = cursor.fetchone()

                    

                    if verify_entry:

                        print(f"[{datetime.now()}] VERIFICAÇÃO BEM-SUCEDIDA - ID na tabela: {verify_entry[0]}")

                        break

                    else:

                        raise Exception("Registro não encontrado após inserção")

                        

                except Exception as insert_error:

                    print(f"[{datetime.now()}] ERRO NA INSERÇÃO (Tentativa {tentativa+1}/3): {str(insert_error)}")

                    if tentativa < 2:  # Se não for a última tentativa

                        time.sleep(1)  # Espera 1 segundo antes de tentar novamente

                    else:

                        raise  # Re-lança a exceção se for a última tentativa



            # LOG DE SUCESSO NO BANCO

            try:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "SUCESSO", "Novo cadastro realizado com sucesso! ID: {novo_id}"))

            except Exception as log_error:

                print(f"Erro ao registrar log de sucesso: {log_error}")

                # Continua mesmo se o log falhar, já que o registro principal foi criado



        # Se admitido = "Sim", duplicar no banco de admitidos

        if form_data.get('admitido') == 'Sim':

            try:

                cursor.execute('SELECT nome_completo FROM registration_form WHERE cpf = ?', (cpf,))

                candidato = cursor.fetchone()

                if candidato:

                    # Verifica se já existe na tabela de admitidos

                    cursor.execute('SELECT id FROM admitidos WHERE nome_completo = ?', (candidato[0],))

                    admitido_existente = cursor.fetchone()

                    

                    if not admitido_existente:

                        # Obter o próximo ID para a tabela admitidos

                        cursor.execute('SELECT ISNULL(MAX(id), 0) + 1 FROM admitidos')

                        novo_id_admitidos = cursor.fetchone()[0]

                        

                        cursor.execute('''

                            INSERT INTO admitidos (id, nome_completo, admitido, data_admissao)

                            VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                        ''', (novo_id_admitidos, candidato[0], admitido))

            except Exception as e:

                print(f"Erro ao registrar admitido: {e}")

                # Continue mesmo se esta parte falhar, pois o registro principal já foi criado



        # Registro de log também em arquivo (opcional)

        try:

            salvar_ficha_log(form_data)

            print(f"Log em arquivo salvo com sucesso")

        except Exception as log_error:

            print(f"Erro ao salvar log em arquivo: {log_error}")

            # Continua mesmo com erro de log



        # Commit final das alterações

        db.commit()

        print(f"Commit finalizado com sucesso para CPF: {cpf}")

        

        # Depois do commit bem-sucedido, remove o backup local se existir

        _remover_backup_formulario(cpf)

        

        return jsonify({"success": True, "message": "Cadastro realizado com sucesso!"})



    except Exception as e:

        # Realiza rollback se possível

        if db:

            try:

                db.rollback()

                print(f"Rollback realizado com sucesso")

            except Exception as rollback_error:

                print(f"Erro ao fazer rollback: {rollback_error}")

                

        erro_msg = str(e)

        print(f"ERRO GRAVE ao inserir no banco de dados: {erro_msg}")

        

        # Armazena detalhes técnicos adicionais para depuração

        import traceback

        traceback_str = traceback.format_exc()

        print(f"Traceback completo: {traceback_str}")



        # Salva o erro em backup local para tentativa posterior

        _salvar_erro_local(cpf, "ERRO_GERAL", f"{erro_msg}\n{traceback_str}", form_data if 'form_data' in locals() else request.form.to_dict())



        # Tenta registrar o erro no banco de dados

        try:

            if cursor and cpf:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "ERRO", erro_msg[:950]))  # Limita tamanho da mensagem para evitar erros

                db.commit()

                print(f"Log de erro registrado no banco de dados")

        except Exception as log_error:

            print(f"Erro ao registrar erro no banco: {log_error}")



        return jsonify({

            "success": False, 

            "message": "Erro ao realizar o cadastro! Os dados foram salvos localmente e serão processados em breve.", 

            "error": erro_msg

        }), 400



    finally:

        # Garante fechamento de recursos

        if cursor:

            try:

                cursor.close()

                print(f"Cursor fechado com sucesso")

            except:

                pass

        if db:

            try:

                db.close()

                print(f"Conexão fechada com sucesso")

            except:

                pass



# Funções auxiliares para backup e recuperação de dados

def _salvar_backup_formulario(dados):

    """Salva um backup dos dados do formulário em arquivo local"""

    try:

        backup_dir = os.path.join('static', 'backups')

        if not os.path.exists(backup_dir):

            os.makedirs(backup_dir)

            

        cpf = dados.get('cpf', 'sem_cpf')

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        filename = f"{cpf}_{timestamp}.json"

        

        with open(os.path.join(backup_dir, filename), 'w', encoding='utf-8') as f:

            json.dump(dados, f, ensure_ascii=False, indent=4)

            

        return True

    except Exception as e:

        print(f"Erro ao salvar backup: {e}")

        return False
def _remover_backup_formulario(cpf):

    """Remove backups existentes para um CPF após salvamento bem-sucedido"""

    try:

        backup_dir = os.path.join('static', 'backups')

        if not os.path.exists(backup_dir):

            return

            

        for filename in os.listdir(backup_dir):

            if filename.startswith(f"{cpf}_") and filename.endswith('.json'):

                os.remove(os.path.join(backup_dir, filename))

                

    except Exception as e:

        print(f"Erro ao remover backup: {e}")



def _salvar_erro_local(cpf, tipo_erro, mensagem, dados):

    """Salva detalhes de erro em arquivo local para recuperação posterior"""

    try:

        error_dir = os.path.join('static', 'errors')

        if not os.path.exists(error_dir):

            os.makedirs(error_dir)

            

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        filename = f"{cpf}_{tipo_erro}_{timestamp}.json"

        

        erro_data = {

            'cpf': cpf,

            'tipo_erro': tipo_erro,

            'mensagem': mensagem,

            'timestamp': timestamp,

            'dados_formulario': dados

        }

        

        with open(os.path.join(error_dir, filename), 'w', encoding='utf-8') as f:

            json.dump(erro_data, f, ensure_ascii=False, indent=4)

            

        return True

    except Exception as e:

        print(f"Erro ao salvar registro de erro: {e}")

        return False



@app.route('/banco_rs', methods=['GET'])

@login_required

def banco_rs():

    db = get_sql_server_connection()

    cursor = db.cursor()



    sincronizar_banco(cursor, db)

    # Atualiza situações em massa

    atualizar_situacao_em_massa(cursor, db)



    # Mapeamento de nomes técnicos para nomes amigáveis

    filtro_nomes_amigaveis = {

        "nome": "Nome Completo",

        "cpf": "CPF",

        "genero": "Gênero",

        "estado_civil": "Estado Civil",

        "data_nasc_inicio": "Data de Nascimento (Início)",

        "data_nasc_fim": "Data de Nascimento (Fim)",

        "fumante": "Fumante",

        "bebida": "Consome Bebida",

        "alergia": "Alergia",

        "medicamento": "Medicamento de Uso Contínuo",

       

        "avaliacao_rh": "Avaliação RH",

        "avaliacao_gerencia": "Avaliação Gerência",

        "avaliacao_geral": "Avaliação Geral",

        "admitido": "Admitido",

        "situacao": "Situação"

    }



    # Mapeamento de IDs de recrutadores para nomes amigáveis

    recrutador_nomes = {

        "samira.barbosa": "Samira Barbosa",

        "nara.rodrigues": "Nara Rodrigues",

        "wilson.monteiro": "Wilson Monteiro",

        "vivian.wanderley": "Vivian Wanderley",

        "grasielle.mapa": "Grasielle Mapa",

        "guilherme.vieira": "Guilherme Vieira"

    }



    # Capturando os filtros do formulário

    form_data = request.args.to_dict(flat=True)  # Captura os filtros



    # Desempacotando filtros individuais

    nome_completo = form_data.get('nome', '')

    cpf = form_data.get('cpf', '')

    genero = form_data.get('genero', '')

    estado_civil = form_data.get('estado_civil', '')

    data_nasc_inicio = form_data.get('data_nasc_inicio', '')

    data_nasc_fim = form_data.get('data_nasc_fim', '')

    fumante = form_data.get('fumante', '')

    bebida = form_data.get('bebida', '')

    alergia = form_data.get('alergia', '')

    medicamento = form_data.get('medicamento', '')

    pcd = form_data.get('pcd', '')

    tatuagem = form_data.get('tatuagem', '')

    cidade = form_data.get('cidade', '')

    regioes_preferencia = request.args.getlist('regioes_preferencia')

    disponibilidade_horario = request.args.getlist('disponibilidade_horario')

    recrutador = form_data.get('recrutador', '')

    sindicancia = form_data.get('sindicancia', '')

    cargos_indicados = request.args.getlist('cargo_indicado')

    avaliacao_rh = form_data.get('avaliacao_rh', '')

    avaliacao_gerencia = form_data.get('avaliacao_gerencia', '')

    avaliacao_geral = form_data.get('avaliacao_geral', '')

    situacao = form_data.get('situacao', '')  # Obtém o valor da situação

    assinatura_gerencia = form_data.get('assinatura_gerencia', '')  # Obtém o valor da situação



    

    # Inicializar consulta base

    # Captura o parâmetro de ordenação

    sort = request.args.get('sort', 'date_desc')  # Padrão: ordem decrescente por data



    # Ajusta a cláusula ORDER BY com base no parâmetro

    if sort == 'date_asc':

        order_by_clause = 'ORDER BY created_at ASC'

    elif sort == 'date_desc':

        order_by_clause = 'ORDER BY created_at DESC'

    else:

        order_by_clause = 'ORDER BY created_at DESC'  # Padrão

    

    # Atualize a query para incluir a cláusula ORDER BY

    query = f'''

        SELECT *, 

               ISNULL(cidade, '') + '/' + ISNULL(bairro, '') AS cidade_bairro,

               ROW_NUMBER() OVER ({order_by_clause}) AS row_num

        FROM registration_form

        WHERE 1=1

    '''



    total_query = '''

        SELECT COUNT(*)

        FROM registration_form 

        WHERE 1=1

    '''



    params = []

    total_params = []



    # Aplicando filtros

    if nome_completo:

        query += ' AND LTRIM(RTRIM(nome_completo)) LIKE ?'

        total_query += ' AND LTRIM(RTRIM(nome_completo)) LIKE ?'

        params.append(f'%{nome_completo}%')

        total_params.append(f'%{nome_completo}%')



    if cpf:

        query += ' AND cpf = ?'

        total_query += ' AND cpf = ?'

        params.append(cpf)

        total_params.append(cpf)



    if genero:

        query += ' AND LOWER(genero) = LOWER(?)'

        total_query += ' AND LOWER(genero) = LOWER(?)'

        params.append(genero)

        total_params.append(genero)



    if estado_civil:

        query += ' AND estado_civil = ?'

        total_query += ' AND estado_civil = ?'

        params.append(estado_civil)

        total_params.append(estado_civil)



    if sindicancia and not avaliacao_geral:

        if sindicancia == 'Em Verificação':

            query += " AND sindicancia = ?"

            total_query += " AND sindicancia = ?"

            params.append('Em Verificação')

            total_params.append('Em Verificação')

        elif sindicancia == 'Aprovado':

            query += " AND sindicancia = ?"

            total_query += " AND sindicancia = ?"

            params.append('Aprovado')

            total_params.append('Aprovado')

        elif sindicancia == 'Reprovado':

            query += " AND sindicancia = ?"

            total_query += " AND sindicancia = ?"

            params.append('Reprovado')

            total_params.append('Reprovado')





    if data_nasc_inicio and data_nasc_fim:

        query += ' AND data_nasc BETWEEN ? AND ?'

        total_query += ' AND data_nasc BETWEEN ? AND ?'

        params.extend([data_nasc_inicio, data_nasc_fim])

        total_params.extend([data_nasc_inicio, data_nasc_fim])



    if fumante:

        query += ' AND fumante = ?'

        total_query += ' AND fumante = ?'

        params.append(fumante)

        total_params.append(fumante)



    if cidade:

        query += ' AND UPPER(cidade) = ?'

        total_query += ' AND UPPER(cidade) = ?'

        params.append(cidade.upper())

        total_params.append(cidade.upper())



    if bebida:

        query += ' AND bebida = ?'

        total_query += ' AND bebida = ?'

        params.append(bebida)

        total_params.append(bebida)



    if alergia:

        query += ' AND alergia = ?'

        total_query += ' AND alergia = ?'

        params.append(alergia)

        total_params.append(alergia)



    if medicamento:

        query += ' AND medicamento_uso_constante = ?'

        total_query += ' AND medicamento_uso_constante = ?'

        params.append(medicamento)

        total_params.append(medicamento)



    if pcd:

        query += ' AND pcd = ?'

        total_query += ' AND pcd = ?'

        params.append(pcd)

        total_params.append(pcd)



    if tatuagem:

        query += ' AND tatuagem = ?'

        total_query += ' AND tatuagem = ?'

        params.append(tatuagem)

        total_params.append(tatuagem)



    if regioes_preferencia:

        regioes_query = ' OR '.join(['regioes_preferencia LIKE ?' for _ in regioes_preferencia])

        query += f' AND ({regioes_query})'

        total_query += f' AND ({regioes_query})'

        for regiao in regioes_preferencia:

            params.append(f'%{regiao}%')

            total_params.append(f'%{regiao}%')



    if disponibilidade_horario:

        horarios_query = ' OR '.join(['disponibilidade_horario LIKE ?' for _ in disponibilidade_horario])

        query += f' AND ({horarios_query})'

        total_query += f' AND ({horarios_query})'

        for horario in disponibilidade_horario:

            params.append(f'%{horario}%')

            total_params.append(f'%{horario}%')



    if recrutador:

        if recrutador.lower() == 'não definido':

            # Busca fichas SEM recrutador definido (vazio ou NULL)

            query += ' AND (recrutador IS NULL OR RTRIM(LTRIM(recrutador)) = \'\')'

            total_query += ' AND (recrutador IS NULL OR RTRIM(LTRIM(recrutador)) = \'\')'

        else:

            query += ' AND recrutador LIKE ?'

            total_query += ' AND recrutador LIKE ?'

            recrutador_formatado = f'%{recrutador.replace(".", "%")}%'

            params.append(recrutador_formatado)

            total_params.append(recrutador_formatado)


    if cargos_indicados:
        cargos_query = ' AND '.join(['cargo_indicado LIKE ?' for _ in cargos_indicados])
        query += f' AND ({cargos_query})'
        total_query += f' AND ({cargos_query})'
        for cargo in cargos_indicados:
            params.append(f'%{cargo}%')
            total_params.append(f'%{cargo}%')

    if avaliacao_rh and not avaliacao_geral and not avaliacao_gerencia and not sindicancia:

        if avaliacao_rh == 'Aprovado':

            query += ' AND avaliacao_rh = ?'

            total_query += ' AND avaliacao_rh = ?'

            params.append('Aprovado')

            total_params.append('Aprovado')

        elif avaliacao_rh == 'Reprovado':

            query += ' AND avaliacao_rh = ?'

            total_query += ' AND avaliacao_rh = ?'

            params.append('Reprovado')

            total_params.append('Reprovado')







    if avaliacao_gerencia and not avaliacao_geral:

        query += ' AND avaliacao_gerencia = ?'

        total_query += ' AND avaliacao_gerencia = ?'

        params.append(avaliacao_gerencia)

        total_params.append(avaliacao_gerencia)

    

    if assinatura_gerencia:

        query += ' AND assinatura_gerencia LIKE ?'

        total_query += ' AND assinatura_gerencia LIKE ?'

        params.append(f'%{assinatura_gerencia}%')

        total_params.append(f'%{assinatura_gerencia}%')



    if avaliacao_geral:

             if avaliacao_geral == 'Aprovado':

                 query += '''

                     AND avaliacao_rh = 'Aprovado'

                     AND avaliacao_gerencia = 'Aprovado'

                     AND sindicancia = 'Aprovado'

                 '''

                 total_query += '''

                     AND avaliacao_rh = 'Aprovado'

                     AND avaliacao_gerencia = 'Aprovado'

                     AND sindicancia = 'Aprovado'

                 '''

             elif avaliacao_geral == 'Reprovado':

                 query += '''

                     AND (

                         avaliacao_rh = 'Reprovado'

                         OR avaliacao_gerencia = 'Reprovado'

                         OR sindicancia = 'Reprovado'

                     )

                 '''

                 total_query += '''

                     AND (

                         avaliacao_rh = 'Reprovado'

                         OR avaliacao_gerencia = 'Reprovado'

                         OR sindicancia = 'Reprovado'

                     )

                 '''

             elif avaliacao_geral == 'Em Conversa':

                 query += '''

                     AND (

                         avaliacao_rh = 'Em Conversa'

                         OR avaliacao_gerencia = 'Em Conversa'

                         OR sindicancia = 'Em Conversa'

                     )

                 '''

                 total_query += '''

                     AND (

                         avaliacao_rh = 'Em Conversa'

                         OR avaliacao_gerencia = 'Em Conversa'

                         OR sindicancia = 'Em Conversa'

                     )

                 '''

             elif avaliacao_geral == 'Não Avaliado':

                 query += '''

                     AND (

                         (avaliacao_rh IS NULL OR avaliacao_rh = '')

                         AND (avaliacao_gerencia IS NULL OR avaliacao_gerencia = '')

                         AND (sindicancia IS NULL OR sindicancia = '')

                     )

                 '''

                 total_query += '''

                     AND (

                         (avaliacao_rh IS NULL OR avaliacao_rh = '')

                         AND (avaliacao_gerencia IS NULL OR avaliacao_gerencia = '')

                         AND (sindicancia IS NULL OR sindicancia = '')

                     )

                 '''





    # Filtro por Admitido
    admitido_filtro = form_data.get('admitido', '')
    if admitido_filtro:
        query += ' AND admitido = ?'
        total_query += ' AND admitido = ?'
        params.append(admitido_filtro)
        total_params.append(admitido_filtro)

    # Filtro por Situação
    if situacao:
        query += ' AND situacao = ?'
        total_query += ' AND situacao = ?'
        params.append(situacao)
        total_params.append(situacao)

    # Parâmetros de paginação

    items_per_page = request.args.get('items_per_page', default=10, type=int)

    page = request.args.get('page', default=1, type=int)

    start_row = (page - 1) * items_per_page + 1

    end_row = start_row + items_per_page - 1



    # Finalizando a consulta

    query = f'''

        SELECT *

        FROM (

            {query}

        ) AS temp

        WHERE row_num BETWEEN ? AND ?

    '''

    params.extend([start_row, end_row])



    # Executar consulta
    cursor.execute(query, params)
    candidatos = cursor.fetchall()
    # Converter resultados
    candidatos_dict = []
    for candidato in candidatos:

        candidato_dict = dict(zip([column[0] for column in cursor.description], candidato))



        # Determinar situação atualizada

        candidato_dict['situacao'] = determinar_situacao(candidato_dict)



        # Ajustar a coluna cidade_bairro

        cidade_bairro = candidato_dict.get('cidade_bairro', '')

        if not cidade_bairro or cidade_bairro == '/':

            cidade_bairro = "Não informado"

        candidato_dict['cidade_bairro'] = cidade_bairro

    

        # Formatar data

        created_at = candidato_dict.get('created_at')

        if created_at:

            try:

                # Se o dado for um objeto datetime

                if isinstance(created_at, datetime):

                    candidato_dict['created_at'] = created_at.strftime('%d/%m/%Y %H:%M')

                # Se o dado for uma string

                elif isinstance(created_at, str):

                    # Tenta diferentes formatos para conversão

                    formatos = ['%Y-%m-%d %H:%M:%S', '%b %d %Y %I:%M%p', '%d/%m/%Y %H:%M']

                    for formato in formatos:

                        try:

                            candidato_dict['created_at'] = datetime.strptime(created_at, formato).strftime('%d/%m/%Y %H:%M')

                            break

                        except ValueError:

                            continue

                    else:

                        candidato_dict['created_at'] = "Formato de data inválido"

                else:

                    candidato_dict['created_at'] = "Não disponível"

            except Exception as e:

                candidato_dict['created_at'] = "Erro ao processar data"

        else:

            candidato_dict['created_at'] = "Não disponível"

    

        # Mapear o ID do recrutador para o nome amigável

        recrutador_id = candidato_dict.get('recrutador')

        if recrutador_id in recrutador_nomes:

            candidato_dict['recrutador'] = recrutador_nomes[recrutador_id]

    

        candidatos_dict.append(candidato_dict)



    # Total de registros com filtros

    count_query = 'SELECT COUNT(*) FROM registration_form WHERE 1=1'

    count_query += query.split('WHERE 1=1')[1].split(') AS temp')[0]

    cursor.execute(count_query, params[:-2])  # Remove parâmetros de paginação para contagem

    total = cursor.fetchone()[0]

    total_pages = (total + items_per_page - 1) // items_per_page



    # Cálculo de intervalo de páginas visíveis

    page_range = 2  # Número de páginas visíveis antes e depois da página atual

    start_page = max(1, page - page_range)

    end_page = min(total_pages, page + page_range)



    # Índices dos candidatos na página atual

    start_candidate = start_row

    end_candidate = min(total, end_row)  # Garante que não ultrapasse o total



    pagination = {

        'page': page,

        'total_pages': total_pages,

        'has_prev': page > 1,

        'has_next': page < total_pages,

        'prev_num': page - 1 if page > 1 else None,

        'next_num': page + 1 if page < total_pages else None,

        'start_page': start_page,

        'end_page': end_page,

        'visible_pages': list(range(start_page, end_page + 1))  # Lista de páginas visíveis

    }



    # Parâmetros que não devem aparecer como filtros

    parametros_excluidos = {'page', 'items_per_page'}



    # Filtrar os parâmetros ativos, excluindo os indesejados

    filtros_ativos = {
        k: v for k, v in form_data.items() if v and k not in parametros_excluidos
    }

    campos_multivalorados = {
        'regioes_preferencia': regioes_preferencia,
        'disponibilidade_horario': disponibilidade_horario,
        'cargo_indicado': cargos_indicados
    }

    for campo, valores in campos_multivalorados.items():
        if valores:
            filtros_ativos[campo] = ', '.join(valores)





    # Links para remoção de filtros, garantindo que os parâmetros excluídos sejam ignorados

    filtro_remocao_links = {

        k: url_for(

            'banco_rs',

            **{key: value for key, value in form_data.items() if key != k and key not in parametros_excluidos}

        )

        for k in filtros_ativos

    }



    current_args = request.args.to_dict()

    # Remove o campo de ordenação para adicionar o novo valor

    args_date_asc = current_args.copy()

    args_date_asc['sort'] = 'date_asc'



    args_date_desc = current_args.copy()

    args_date_desc['sort'] = 'date_desc'





    # Renderizando o template

    return render_template(

        'banco_rs.html',

        candidatos=candidatos_dict,

        pagination=pagination,

        filtros_ativos=filtros_ativos,

        filtro_remocao_links=filtro_remocao_links,

        start_candidate=start_candidate,

        end_candidate=end_candidate,

        total_candidates=total,

        current_args=current_args,

        args_date_asc=args_date_asc,

        args_date_desc=args_date_desc,

        recrutador_nomes=recrutador_nomes

    )



@app.route('/inscription', methods=['GET', 'POST'])

@login_required

def inscription():

    if request.method == 'POST':

        form_data = request.form.to_dict()

        return redirect(url_for('inscription'))



    return render_template('view_or_fill_inscription.html', form_data={})



@app.route('/view_or_fill_inscription/<int:id>', methods=['GET', 'POST'])

@login_required

def view_or_fill_inscription(id):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Busca o ticket associado ao ID fornecido

    cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

    ticket = cursor.fetchone()

    if not ticket:

        cursor.close()

        return f"Ticket não encontrado para o ID: {id}", 404



    cpf = ticket[13]  # Supondo que 'cpf' está na coluna 13 no SQL Server

    

    # Inicializa o dicionário form_data

    form_data = {

        'nome_completo': '',

        'endereco': '',

        'numero': '',

        'complemento': '',

        'cep': '',

        'estado': '',

        'cidade': '',

        'bairro': '',

        'cpf': cpf

    }



    # Busca a ficha de inscrição na tabela registration_form usando o CPF do ticket

    cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

    registration_form = cursor.fetchone()

    if registration_form:

        form_data.update({

            'nome_completo': registration_form[2],

            'cep': registration_form[5],

            'endereco': registration_form[6],

            'numero': registration_form[7],

            'complemento': registration_form[8],

            'bairro': registration_form[9],

            'cidade': registration_form[10],

            'estado': registration_form[11]

        })

    else:

        cursor.close()

        return f"Ficha não encontrada para o CPF: {cpf}", 404



    # Lógica adicional para buscar dados de um banco externo

    try:

        conn = get_jbc_connection()

        ext_cursor = conn.cursor()

        ext_cursor.execute('''

            SELECT pe_nome, pe_cep, pe_logradouro_end, pe_numero_end, pe_complemento_end, pe_bairro_end, pe_cidade_end 

            FROM dbo.fo_pessoa 

            WHERE pe_cpf = ?

        ''', (cpf,))

        external_data = ext_cursor.fetchone()

        conn.close()

    except Exception as e:

        external_data = None

        print(f"Erro ao conectar ao banco de dados externo: {e}")



    # Atualiza form_data com dados externos, se existirem

    if external_data:

        form_data.update({

            'nome_completo': external_data[0] if external_data[0] else '',           

            'endereco': external_data[1] if external_data[1] else '',       

            'numero': external_data[2] if external_data[2] else '',         

            'complemento': external_data[3] if external_data[3] else '',    

            'cep': external_data[4] if external_data[4] else '',            

            'estado': external_data[5] if external_data[5] else '',         

            'cidade': external_data[6] if external_data[6] else '',         

            'bairro': external_data[7] if external_data[7] else ''          

        })



    if request.method == 'POST':

        data = request.form

        try:

            # Verifica se o registro existe para determinar se é uma inserção ou atualização

            cursor.execute('SELECT cpf FROM registration_form WHERE cpf = ?', (data['cpf'],))

            existing_entry = cursor.fetchone()



            if existing_entry:

                # Atualiza registro existente

                update_query = '''

                    UPDATE registration_form

                    SET nome_completo = ?, cep = ?, endereco = ?, numero = ?, complemento = ?, bairro = ?, cidade = ?, estado_nasc = ?, 

                        telefone = ?, telefone_recado = ?, cidade_nasc = ?, data_nasc = ?, idade = ?, numero_filhos = ?, filhos = ?, 

                        estado_civil = ?, fumante = ?, bebida = ?, alergia = ?, medicamento_constante = ?, qual_medicamento = ?, 

                        medicamento_uso_constante = ?, genero = ?, peso = ?, cor_pele = ?, tatuagem = ?, perfil = ?, 

                        cargo_indicado = ?, identidade = ?, cursos_realizados = ?, regioes_preferencia = ?, disponibilidade_horario = ?, 

                        empresa1 = ?, cidade1 = ?, funcao1 = ?, data_admissao1 = ?, data_saida1 = ?, motivo_saida1 = ?, salario1 = ?, 

                        empresa2 = ?, cidade2 = ?, funcao2 = ?, data_admissao2 = ?, data_saida2 = ?, motivo_saida2 = ?, salario2 = ?, 

                        empresa3 = ?, cidade3 = ?, funcao3 = ?, data_admissao3 = ?, data_saida3 = ?, motivo_saida3 = ?, salario3 = ?, 

                        empregos_informais = ?, nome_entrevistador = ?, avaliacao_rh = ?, assinatura_rh = ?, avaliacao_gerencia = ?, 

                        assinatura_gerencia = ?, rota_trabalho = ?, curriculo = ?, situacao = ?, last_updated = GETDATE()

                    WHERE cpf = ?

                '''

                cursor.execute(update_query, (

                    data['nome_completo'], data.get('cep'), data.get('endereco'), data.get('numero'), data.get('complemento'),

                    data.get('bairro'), data.get('cidade'), data.get('estado'), data.get('telefone'), data.get('telefone_recado'), data.get('cidade_nasc'), 

                    data.get('data_nasc'), data.get('idade'), data.get('numero_filhos'), data.get('filhos'), data.get('estado_civil'),

                    data.get('fumante'), data.get('bebida'), data.get('alergia'), data.get('medicamento_constante'), data.get('qual_medicamento'), 

                    data.get('medicamento_constante'), data.get('genero'), data.get('peso'), data.get('cor_pele'), data.get('tatuagem'), 

                    data.get('perfil'), data.get('cargo_indicado'), data.get('identidade'), data.get('cursos_realizados'), data.get('regioes_preferencia'), 

                    data.get('disponibilidade_horario'), data.get('empresa1'), data.get('cidade1'), data.get('funcao1'), data.get('data_admissao1'), 

                    data.get('data_saida1'), data.get('motivo_saida1'), data.get('salario1'), data.get('empresa2'), data.get('cidade2'), 

                    data.get('funcao2'), data.get('data_admissao2'), data.get('data_saida2'), data.get('motivo_saida2'), data.get('salario2'), 

                    data.get('empresa3'), data.get('cidade3'), data.get('funcao3'), data.get('data_admissao3'), data.get('data_saida3'), 

                    data.get('motivo_saida3'), data.get('salario3'), data.get('empregos_informais'), data.get('nome_entrevistador'), 

                    data.get('avaliacao_rh'), data.get('assinatura_rh'), data.get('avaliacao_gerencia'), data.get('assinatura_gerencia'), 

                    data.get('rota_trabalho'), data.get('curriculo'), data.get('situacao'), data['cpf']

                ))

            else:

                # Insere novo registro

                insert_query = '''

                    INSERT INTO registration_form (

                        cpf, nome_completo, cep, endereco, numero, complemento, bairro, cidade, estado_nasc, telefone, telefone_recado, 

                        cidade_nasc, data_nasc, idade, numero_filhos, filhos, estado_civil, fumante, bebida, alergia, 

                        medicamento_constante, qual_medicamento, medicamento_uso_constante, genero, peso, cor_pele, tatuagem, 

                        perfil, cargo_indicado, identidade, cursos_realizados, regioes_preferencia, disponibilidade_horario, 

                        empresa1, cidade1, funcao1, data_admissao1, data_saida1, motivo_saida1, salario1, empresa2, cidade2, 

                        funcao2, data_admissao2, data_saida2, motivo_saida2, salario2, empresa3, cidade3, funcao3, data_admissao3, 

                        data_saida3, motivo_saida3, salario3, empregos_informais, nome_entrevistador, avaliacao_rh, assinatura_rh, 

                        avaliacao_gerencia, assinatura_gerencia, rota_trabalho, curriculo, situacao, created_at, last_updated

                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE())

                '''

                cursor.execute(insert_query, (

                    data['cpf'], data['nome_completo'], data.get('cep'), data.get('endereco'), data.get('numero'), data.get('complemento'),

                    data.get('bairro'), data.get('cidade'), data.get('estado'), data.get('telefone'), data.get('telefone_recado'), data.get('cidade_nasc'), data.get('data_nasc'), 

                    data.get('idade'), data.get('numero_filhos'), data.get('filhos'), data.get('estado_civil'), data.get('fumante'), data.get('bebida'), 

                    data.get('alergia'), data.get('medicamento_constante'), data.get('qual_medicamento'), data.get('medicamento_constante'), data.get('genero'), 

                    data.get('peso'), data.get('cor_pele'), data.get('tatuagem'), data.get('perfil'), data.get('cargo_indicado'), data.get('identidade'), 

                    data.get('cursos_realizados'), data.get('regioes_preferencia'), data.get('disponibilidade_horario'), data.get('empresa1'), data.get('cidade1'), 

                    data.get('funcao1'), data.get('data_admissao1'), data.get('data_saida1'), data.get('motivo_saida1'), data.get('salario1'), data.get('empresa2'), 

                    data.get('cidade2'), data.get('funcao2'), data.get('data_admissao2'), data.get('data_saida2'), data.get('motivo_saida2'), data.get('salario2'), 

                    data.get('empresa3'), data.get('cidade3'), data.get('funcao3'), data.get('data_admissao3'), data.get('data_saida3'), data.get('motivo_saida3'), 

                    data.get('salario3'), data.get('empregos_informais'), data.get('nome_entrevistador'), data.get('avaliacao_rh'), data.get('assinatura_rh'), 

                    data.get('avaliacao_gerencia'), data.get('assinatura_gerencia'), data.get('rota_trabalho'), data.get('curriculo'), data.get('situacao')

                ))



            db.commit()

            cursor.close()



        except Exception as e:

            db.rollback()

            cursor.close()

            print(f"Erro ao salvar no banco de dados: {e}")

            return f"Erro ao salvar no banco de dados: {str(e)}", 500



        return redirect(url_for('sistema_rs'))



    cursor.close()

    return render_template('view_or_fill_inscription.html', ticket=ticket, form_data=form_data)



@app.route('/admin/manage_users', methods=['GET', 'POST'])

@login_required

def manage_users():

    if not current_user.is_admin:

        flash('Acesso negado: Você não tem permissão para acessar esta página.', 'danger')

        return redirect(url_for('home'))



    db = get_sql_server_connection()

    cursor = db.cursor()



    # Se o método for POST, significa que estamos adicionando um usuário

    if request.method == 'POST':

        username = request.form['username']

        name = request.form['name']

        email = request.form['email']

        password = generate_password_hash(request.form['password'])

        is_admin = int('is_admin' in request.form)



        # Verificar se o username já existe

        cursor.execute('SELECT * FROM users WHERE username = ?', (username,))

        existing_user = cursor.fetchone()



        if existing_user:

            flash('O nome de usuário já está em uso. Por favor, escolha outro.', 'danger')

            return redirect(url_for('manage_users'))



        # Inserir o novo usuário

        cursor.execute('''

            INSERT INTO users (username, name, email, password, is_admin) 

            VALUES (?, ?, ?, ?, ?)

        ''', (username, name, email, password, is_admin))

        db.commit()



        flash('Usuário adicionado com sucesso!', 'success')

        return redirect(url_for('manage_users'))



    # Recupera a lista de usuários cadastrados

    cursor.execute('SELECT * FROM users')

    users = cursor.fetchall()

    cursor.close()



    return render_template('admin_dashboard.html', users=users)



@app.route('/manage_candidates', methods=['GET', 'POST'])

@login_required

def manage_candidates():

    db = get_sql_server_connection()

    cursor = db.cursor()



    if request.method == 'POST':

        candidate_id = request.form['candidate_id']

        new_status = request.form['status']

        rejection_reason = request.form.get('rejection_reason', '')



        if new_status == "REPROVADO" and not rejection_reason:

            flash("Motivo da reprovação é obrigatório para candidatos reprovados.", "danger")

            return redirect(url_for('manage_candidates'))



        cursor.execute('''

            UPDATE registration_form

            SET status = ?, rejection_reason = ?

            WHERE id = ?

        ''', (new_status, rejection_reason if new_status == "REPROVADO" else None, candidate_id))

        db.commit()



        flash("Situação do candidato atualizada com sucesso!", "success")

        return redirect(url_for('manage_candidates'))



    cursor.execute('SELECT * FROM registration_form')

    candidates = cursor.fetchall()

    cursor.close()

    

    return render_template('manage_candidates.html', candidates=candidates)
def user_logs():

    if not current_user.is_admin:

        flash('Acesso negado: Você não tem permissão para acessar esta página.', 'danger')

        return redirect(url_for('home'))



    db = get_sql_server_connection()

    cursor = db.cursor()



    search_query = request.args.get('search', '')

    user_id_filter = request.args.get('user_id', '')

    action_filter = request.args.get('action', '')



    query = '''

        SELECT logs.id, logs.user_id, logs.action, logs.created_at, users.name 

        FROM user_logs AS logs

        JOIN users ON logs.user_id = users.id

        WHERE 1=1

    '''

    params = []



    if user_id_filter:

        query += ' AND logs.user_id = ?'

        params.append(user_id_filter)

    if action_filter:

        query += ' AND logs.action LIKE ?'

        params.append(f'%{action_filter}%')

    if search_query:

        query += ' AND (users.name LIKE ? OR logs.action LIKE ?)'

        params.extend([f'%{search_query}%', f'%{search_query}%'])



    page = request.args.get(get_page_parameter(), type=int, default=1)

    per_page = 10

    start_row = (page - 1) * per_page + 1

    end_row = start_row + per_page - 1



    paginated_query = f'''

        SELECT * FROM (

            SELECT logs.id, logs.user_id, logs.action, logs.created_at, users.name, 

                   ROW_NUMBER() OVER (ORDER BY logs.created_at DESC) AS row_num

            FROM user_logs AS logs

            JOIN users ON logs.user_id = users.id

            WHERE 1=1 {query[len(query) - len(query.split('WHERE')[1]):]}

        ) AS temp_table

        WHERE row_num BETWEEN ? AND ?

    '''

    params.extend([start_row, end_row])



    cursor.execute(paginated_query, params)

    logs = cursor.fetchall()



    cursor.execute('SELECT COUNT(*) FROM user_logs')

    total = cursor.fetchone()[0]

    total_pages = (total // per_page) + (1 if total % per_page else 0)



    pagination = Pagination(page=page, total=total, per_page=per_page, css_framework='bootstrap4')



    cursor.close()



    return render_template(

        'user_logs.html', 

        logs=logs, 

        pagination=pagination, 

        search_query=search_query, 

        user_id_filter=user_id_filter, 

        action_filter=action_filter

    )



# Filtro para exibir apenas o primeiro e o segundo nome de uma string de nome completo

@app.template_filter('first_and_second_name')

def first_and_second_name(name):

    # Certifique-se de que `name` é uma string

    if not isinstance(name, str):

        name = str(name)

    parts = name.split()

    return ' '.join(parts[:2]) if parts else 'Desconhecido'



# Registra o filtro para uso nos templates Jinja

app.jinja_env.filters['first_and_second_name'] = first_and_second_name



# Filtro para formatar datas no estilo brasileiro com fuso horário de Brasília

@app.template_filter('format_brazilian_date')

def format_brazilian_date(value):

    if value:

        try:

            # Converte o valor para um objeto datetime, se for uma string

            if isinstance(value, str):

                dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')

            else:

                dt = value  # Se já for um objeto datetime, apenas usa o valor



            # Define o fuso horário de Brasília e ajusta a data

            brasilia_tz = pytz.timezone('America/Sao_Paulo')

            dt_brasilia = dt.astimezone(brasilia_tz)



            # Formata a data no padrão brasileiro

            return dt_brasilia.strftime('%d/%m/%Y %H:%M:%S')

        except Exception as e:

            return value  # Retorna o valor original em caso de erro

    return value



# Registra o filtro para uso nos templates Jinja

app.jinja_env.filters['format_brazilian_date'] = format_brazilian_date



@app.route('/verify_cpf', methods=['POST'])

@login_required

def verify_cpf():

    data = request.json

    cpf = data['cpf'].replace('.', '').replace('-', '')  # Remove os pontos e traços do CPF

    category = data['category'].lower()  # Transforma a categoria em minúsculas para garantir compatibilidade



    print(f"Verificando CPF: {cpf} na categoria: {category}")  # Log para depuração



    db = get_sql_server_connection()

    cursor = db.cursor()



    response_data = None



    try:

        # Verificar se o CPF existe na tabela 'tickets' (para todas as categorias)

        query_tickets = '''

            SELECT name, cep, rua, numero, complemento, bairro, cidade, telefones

            FROM tickets

            WHERE REPLACE(REPLACE(cpf, '.', ''), '-', '') = ?

        '''

        cursor.execute(query_tickets, (cpf,))

        ticket_data = cursor.fetchone()



        if ticket_data:

            print(f"CPF encontrado na tabela 'tickets' para a categoria {category}")  # Log para depuração



            # Preenche os dados do ticket se encontrado

            response_data = {

                'exists': True,

                'name': ticket_data[0],

                'cep': ticket_data[1],

                'rua': ticket_data[2],

                'numero': ticket_data[3],

                'complemento': ticket_data[4],

                'bairro': ticket_data[5],

                'cidade': ticket_data[6],

                'telefones': ticket_data[7],

                'external': False  # Indica que foi encontrado no banco interno de tickets

            }



            # Verificação adicional para a categoria "entrevista"

            if category == "entrevista":

                query_candidates = '''

                    SELECT name, created_at

                    FROM interview_candidates

                    WHERE REPLACE(REPLACE(cpf, '.', ''), '-', '') = ?

                '''

                cursor.execute(query_candidates, (cpf,))

                candidate = cursor.fetchone()



                if candidate:

                    query_situacao = '''

                        SELECT situacao

                        FROM registration_form

                        WHERE REPLACE(REPLACE(cpf, '.', ''), '-', '') = ?

                    '''

                    cursor.execute(query_situacao, (cpf,))

                    situacao = cursor.fetchone()



                    # Formatar a data e adicionar ao aviso

                    formatted_date = candidate[1] if candidate[1] else 'Data não disponível'

                    formatted_situacao = situacao[0] if situacao else 'Não Avaliado'



                    response_data.update({

                        'warning': f"Atenção! O candidato {candidate[0].upper()} já participou do processo no dia {formatted_date}. Situação: {formatted_situacao}."

                    })



            return jsonify(response_data)



        # Caso o CPF não seja encontrado na tabela 'tickets', verificar no banco externo

        conn = get_jbc_connection()

        ext_cursor = conn.cursor()



        query_external = '''

            SELECT pe_nome, pe_cep, pe_numero_end, pe_complemento_end 

            FROM dbo.fo_pessoa 

            WHERE REPLACE(REPLACE(pe_cpf, '.', ''), '-', '') = ?

        '''

        ext_cursor.execute(query_external, (cpf,))

        result = ext_cursor.fetchone()

        conn.close()



        if result:

            response_data = {

                'exists': True,

                'name': result[0],

                'cep': result[1],

                'numero': result[2],

                'complemento': result[3],

                'external': True

            }

        else:

            response_data = {

                'exists': False,

                'message': 'CPF não encontrado. Prosseguindo com o cadastro.'

            }



    except pyodbc.Error as e:

        print(f"Erro ao verificar CPF: {e}")

        response_data = {

            'exists': False,

            'error': f"Erro ao verificar CPF: {str(e)}"

        }

    finally:

        cursor.close()

        db.close()



    return jsonify(response_data)









def verify_internal_cpf(cursor, cpf, category):

    """

    Verifica o CPF no banco de dados interno e retorna os dados, se encontrados.

    """

    cursor.execute('''

        SELECT name, cep, rua, numero, complemento, bairro, cidade, telefones

        FROM tickets

        WHERE cpf = ?

    ''', (cpf,))

    ticket_data = cursor.fetchone()



    if ticket_data:

        response_data = {

            'exists': True,

            'name': ticket_data[0],

            'cep': ticket_data[1],

            'rua': ticket_data[2],

            'numero': ticket_data[3],

            'complemento': ticket_data[4],

            'bairro': ticket_data[5],

            'cidade': ticket_data[6],

            'telefones': ticket_data[7],

            'external': False

        }



        # Verifica "entrevista" se a categoria for correspondente

        if category == "entrevista":

            enhance_with_interview_data(cursor, cpf, response_data)



        return response_data



    return {'exists': False}





def verify_external_cpf(cpf, category):

    """

    Verifica o CPF no banco de dados externo e retorna os dados, se encontrados.

    """

    conn = get_jbc_connection()

    ext_cursor = conn.cursor()



    try:

        ext_cursor.execute('''

            SELECT pe_nome, pe_cep, pe_numero_end, pe_complemento_end 

            FROM dbo.fo_pessoa 

            WHERE REPLACE(REPLACE(pe_cpf, '.', ''), '-', '') = ?

        ''', (cpf,))

        result = ext_cursor.fetchone()



        if result:

            response_data = {

                'exists': True,

                'name': result[0],

                'cep': result[1],

                'numero': result[2],

                'complemento': result[3],

                'external': True

            }



            # Verifica "entrevista" se a categoria for correspondente

            if category == "entrevista":

                with get_sql_server_connection() as db:

                    cursor = db.cursor()

                    enhance_with_interview_data(cursor, cpf, response_data)



            return response_data



    except pyodbc.Error as e:

        print(f"Erro no banco de dados externo: {e}")

        return {'success': False, 'message': f"Erro no banco externo: {str(e)}"}



    finally:

        ext_cursor.close()

        conn.close()



    return {'exists': False}





def enhance_with_interview_data(cursor, cpf, response_data):

    """

    Adiciona informações adicionais ao dicionário de resposta com base na participação em entrevistas.

    """

    cursor.execute('SELECT name, created_at FROM interview_candidates WHERE cpf = ?', (cpf,))

    candidate = cursor.fetchone()



    if candidate:

        cursor.execute('SELECT situacao FROM registration_form WHERE cpf = ?', (cpf,))

        situacao = cursor.fetchone()



        response_data.update({

            'name': candidate[0],

            'created_at': candidate[1],

            'situacao': situacao[0] if situacao else 'Não Avaliado',

            'warning': 'Este candidato já participou de uma entrevista anteriormente.'  # Adiciona o aviso

        })
@app.route('/verify_cpf_modal', methods=['POST'])
@login_required
def verify_cpf_modal():

    data = request.json

    cpf = data['cpf'].replace('.', '').replace('-', '')  # Remove os pontos e traços do CPF



    db = get_sql_server_connection()

    cursor = db.cursor()



    # Verificar se o CPF existe na tabela 'tickets'

    cursor.execute('''

        SELECT name, cep, rua, numero, complemento, bairro, cidade, telefones

        FROM tickets

        WHERE REPLACE(REPLACE(cpf, ".", ""), "-", "") = ?

    ''', (cpf,))

    ticket_data = cursor.fetchone()



    if ticket_data:

        print("CPF encontrado na tabela 'tickets'")  # Log para depuração

        

        # Preenche os dados do ticket se encontrado

        response_data = {

            'exists': True,

            'name': ticket_data[0],

            'cep': ticket_data[1],

            'rua': ticket_data[2],

            'numero': ticket_data[3],

            'complemento': ticket_data[4],

            'bairro': ticket_data[5],

            'cidade': ticket_data[6],

            'telefones': ticket_data[7],

            'external': False  # Indica que foi encontrado no banco interno de tickets

        }

        cursor.close()

        return jsonify(response_data)



    # Caso o CPF não seja encontrado na tabela 'tickets', continuar com a lógica

    cursor.execute('''

        SELECT name, cep, numero, complemento, bairro, cidade

        FROM registration_form

        WHERE REPLACE(REPLACE(cpf, ".", ""), "-", "") = ?

    ''', (cpf,))

    registration_data = cursor.fetchone()



    if registration_data:

        response_data = {

            'exists': True,

            'name': registration_data[0],

            'cep': registration_data[1],

            'rua': registration_data[2],

            'numero': registration_data[3],

            'complemento': registration_data[4],

            'bairro': registration_data[5],

            'cidade': registration_data[6],

            'external': False

        }

        cursor.close()

        return jsonify(response_data)



    # Verificar se o CPF existe na tabela 'interview_candidates'

    cursor.execute('''

        SELECT name

        FROM interview_candidates

        WHERE REPLACE(REPLACE(cpf, ".", ""), "-", "") = ?

    ''', (cpf,))

    interview_data = cursor.fetchone()



    if interview_data:

        cursor.close()

        return jsonify({

            'exists': True,

            'name': interview_data[0],

            'warning': 'Este candidato já participou de uma entrevista anteriormente.'

        })



    # Se o CPF não for encontrado em nenhum dos bancos internos, verificar no banco externo

    conn = get_jbc_connection()

    ext_cursor = conn.cursor()



    ext_cursor.execute('''

        SELECT pe_nome, pe_cep, pe_numero_end, pe_complemento_end 

        FROM dbo.fo_pessoa 

        WHERE REPLACE(REPLACE(pe_cpf, '.', ''), '-', '') = ?

    ''', (cpf,))

    result = ext_cursor.fetchone()

    conn.close()



    if result:

        nome, cep, numero, complemento = result[0], result[1], result[2], result[3]

        return jsonify({

            'exists': True,

            'name': nome,

            'cep': cep,

            'numero': numero,

            'complemento': complemento,

            'external': True

        })



    # Se o CPF não for encontrado em nenhum dos bancos de dados

    cursor.close()

    return jsonify({'exists': False, 'message': 'CPF não encontrado. Prosseguindo com o cadastro.'})



@app.route('/create_ticket', methods=['POST'])

@login_required

def create_ticket():

    # Inicialize o cursor como None para evitar problemas no bloco finally

    cursor = None

    try:

        # Obtendo os dados do formulário

        category = request.form['category']

        name = request.form['name']

        cpf = request.form.get('cpf')

        data_nasc = request.form.get('data_nasc')  # Novo campo



        # Remove caracteres não numéricos do CPF (limpeza)

        if cpf:

            cpf = ''.join(filter(str.isdigit, cpf))



        # Validação do CPF

        if cpf and not cpf.isdigit():

            return f"CPF inválido: {cpf}. Deve conter apenas números.", 400



        # Validação de prioridade

        try:

            priority = int(request.form['priority'])

        except ValueError:

            return "Prioridade inválida. Deve ser um número.", 400



        # Validação da data de nascimento

        if data_nasc:

            try:

                data_nasc = datetime.strptime(data_nasc, '%Y-%m-%d').strftime('%Y-%m-%d')

            except ValueError:

                return "Data de nascimento inválida. Use o formato YYYY-MM-DD.", 400



        cep = request.form['cep']

        rua = request.form['rua']

        numero = request.form['numero']

        complemento = request.form['complemento']

        bairro = request.form['bairro']

        cidade = request.form['cidade']

        telefones = request.form['telefones']

        ticket_number = generate_ticket_number(category)



        # Campos opcionais

        especificacao = request.form.get('especificacao') if category == 'Outros' else None

        recruiter = request.form.get('recruiter') if category == 'Agendado' else None



        # Conexão com o banco de dados

        db = get_sql_server_connection()

        cursor = db.cursor()

        created_at = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')



        # Lógica para Entrevista

        if category == 'Entrevista' and cpf:

            # Verifica se o CPF já existe no formulário de registro

            cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

            existing_form = cursor.fetchone()



            # Verifica se o CPF já está na tabela de candidatos para entrevista

            cursor.execute('SELECT * FROM interview_candidates WHERE cpf = ?', (cpf,))

            existing_interview = cursor.fetchone()



            if existing_form:

                # Apenas avise (no frontend), mas continue o fluxo normalmente

                cpf_warning = f"Ficha já existe para este CPF: {cpf}."

            else:

                cpf_warning = None



            if not existing_interview:

                # Gera o próximo valor para id em interview_candidates

                cursor.execute("SELECT ISNULL(MAX(id), 0) + 1 FROM interview_candidates")

                next_id = cursor.fetchone()[0]



                # Insere o CPF na tabela de candidatos para entrevista

                cursor.execute(

                    'INSERT INTO interview_candidates (id, cpf, name, created_at) VALUES (?, ?, ?, ?)',

                    (next_id, cpf, name, created_at)

                )



        # Inserção do ticket com cálculo do próximo id

        cursor.execute("SELECT ISNULL(MAX(id), 0) + 1 FROM tickets")

        next_ticket_id = cursor.fetchone()[0]



        cursor.execute('''

            INSERT INTO tickets (id, name, category, ticket_number, created_at, priority, stage, updated_at, 

                                 status, cpf, cep, rua, numero, complemento, bairro, cidade, telefones, recruiter, especificacao, data_nasc) 

            OUTPUT Inserted.id

            VALUES (?, ?, ?, ?, ?, ?, 'RH', ?, 'ESPERA', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)

        ''', (next_ticket_id, name, category, ticket_number, created_at, priority, created_at, cpf, cep, rua, numero, complemento, bairro, cidade, telefones, recruiter, especificacao, data_nasc))



        # Recupera o ID do ticket diretamente da execução do comando

        ticket_id_row = cursor.fetchone()

        if ticket_id_row:

            ticket_id = ticket_id_row[0]

        else:

            raise ValueError("Erro ao gerar o ID do ticket. Verifique a configuração do banco de dados.")



        # Commit no banco de dados

        db.commit()



        # Emissão do evento para adicionar o ticket em tempo real

        socketio.emit('new_ticket', {

            'id': ticket_id,

            'name': name,

            'ticket_number': ticket_number,

            'created_at': created_at,

            'category': category,

            'cpf': cpf,

            'data_nasc': data_nasc  # Inclui a data de nascimento no evento

        }, namespace='/')



    except Exception as e:

        if db:

            db.rollback()

        print(f"Erro ao criar o ticket: {e}")

        return "Erro ao criar o ticket. Verifique o log do servidor.", 500

    finally:

        # Fecha o cursor apenas se ele foi inicializado

        if cursor:

            cursor.close()



    # Redireciona para o painel após o sucesso

    return jsonify(success=True, ticket={

        'id': ticket_id,

        'name': name,

        'ticket_number': ticket_number,

        'created_at': created_at,

        'category': category,

        'cpf': cpf,

        'data_nasc': data_nasc  # Inclui a data de nascimento na resposta

    })

















@app.route('/complete_registration', methods=['GET', 'POST'])

@login_required

def complete_registration():

    if request.method == 'POST':

        data = request.form

        cpf = data['cpf']

        name = data['name']

        cep = data.get('cep')

        rua = data.get('rua')

        numero = data.get('numero')

        complemento = data.get('complemento')

        bairro = data.get('bairro')

        cidade = data.get('cidade')

        telefones = data.get('telefones')



        db = get_sql_server_connection()

        cursor = db.cursor()

        

        cursor.execute('''

            INSERT INTO tickets (cpf, name, cep, rua, numero, complemento, bairro, cidade, telefones)

            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)

        ''', (cpf, name, cep, rua, numero, complemento, bairro, cidade, telefones))

        

        db.commit()

        cursor.close()



        return redirect(url_for('painel'))



    cpf = request.args.get('cpf')

    name = request.args.get('name')

    return render_template('complete_registration.html', cpf=cpf, name=name)



@app.route('/call_with_alert/<int:id>', methods=['POST'])

@login_required

def call_with_alert(id):

    db = get_sql_server_connection()

    cursor = db.cursor()

    called_at = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')



    cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

    ticket = cursor.fetchone()

    if not ticket:

        cursor.close()

        return jsonify(success=False, message="Ticket não encontrado"), 404



    cursor.execute(

        'UPDATE tickets SET status = ?, called_at = ? WHERE id = ?',

        ('CHAMADO', called_at, id)

    )

    db.commit()



    cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

    updated_ticket = cursor.fetchone()

    

    formatted_time = (

        datetime.strptime(updated_ticket[5], '%Y-%m-%d %H:%M:%S').strftime('%H:%M:%S')

        if updated_ticket[5] else datetime.now(BRASILIA_TZ).strftime('%H:%M:%S')

    )



    socketio.emit('update_display', {

        'current_ticket_number': updated_ticket[3],

        'current_guiche': updated_ticket[9],

        'current_name': first_and_second_name(updated_ticket[1]),

        'current_time': formatted_time,

        'called_tickets': [],

        'play_audio': True

    }, namespace='/')



    cursor.close()

    

    return jsonify(success=True, ticket={

        'id': updated_ticket[0],

        'ticket_number': updated_ticket[3],

        'name': updated_ticket[1],

        'category': updated_ticket[2] if updated_ticket[2] else "Sem Categoria",  # 🔹 Garante que sempre tenha categoria

        'status': 'CHAMADO',

        'guiche': updated_ticket[9],

        'called_at': formatted_time

    })





@app.route('/send_tv/<int:id>', methods=['POST'])

@login_required

def send_tv(id):

    guiche = request.form.get('guiche')

    stage = request.form.get('stage', 'RH')  # Novo: Pega o setor, padrão RH

    if not guiche:

        return "Guichê é obrigatório", 400



    db = get_sql_server_connection()

    cursor = db.cursor()

    called_at = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')



    try:

        cursor.execute(

            'UPDATE dbo.tickets SET status = ?, called_at = ?, guiche = ? WHERE id = ?',

            ('CHAMADO', called_at, guiche, id)

        )

        db.commit()



        cursor.execute('SELECT * FROM dbo.tickets WHERE id = ?', (id,))

        ticket = cursor.fetchone()



        if ticket:

            ticket_dict = {

                'ticket_number': ticket[3],

                'guiche': ticket[9],

                'name': ticket[1],

                'called_at': ticket[5],

            }



            cursor.execute(

                'SELECT TOP 10 * FROM dbo.tickets WHERE status = ? ORDER BY priority DESC, called_at DESC',

                ('CHAMADO',)

            )

            called_tickets = cursor.fetchall()



            called_tickets_list = [

                {

                    'name': first_and_second_name(str(t[1])),

                    'ticket_number': t[3],

                    'guiche': t[9],

                    'called_at': t[5].split(' ')[1] if t[5] and ' ' in t[5] else t[5],

                }

                for t in called_tickets

            ]



            # >>>>>>> Alterado aqui:

            socketio.emit('update_display', {

                'current_ticket_number': ticket_dict['ticket_number'],

                'current_guiche': ticket_dict['guiche'],

                'current_name': first_and_second_name(str(ticket_dict['name'])),

                'current_time': ticket_dict['called_at'].split(' ')[1] if ticket_dict['called_at'] and ' ' in ticket_dict['called_at'] else ticket_dict['called_at'],

                'called_tickets': called_tickets_list,

                'stage': stage,           # <<< Aqui vai o setor!

                'play_audio': True        # <<< Sempre toca áudio ao enviar para TV

            }, namespace='/')



            socketio.emit('update_guiche', {'ticket_id': id, 'guiche': guiche}, namespace='/')



        return '', 204



    except Exception as e:

        db.rollback()

        return str(e), 500



    finally:

        cursor.close()

        db.close()
@app.route('/reposition_ticket/<int:id>', methods=['POST'])
@login_required
def reposition_ticket(id):

    db = get_sql_server_connection()

    cursor = db.cursor()



    cursor.execute('UPDATE tickets SET status = ? WHERE id = ?', ('ESPERA', id))

    db.commit()

    cursor.close()



    socketio.emit('update_queue', {'data': 'ticket repositioned'}, namespace='/')

    return redirect(url_for('painel'))



@app.route('/reset_indicators', methods=['POST'])
@login_required
def reset_indicators():

    db = get_sql_server_connection()

    cursor = db.cursor()



    cursor.execute('DELETE FROM tickets WHERE status = ?', ('CONCLUIDO',))

    db.commit()

    cursor.close()



    socketio.emit('update_queue', {'data': 'indicators reset'}, namespace='/')

    return redirect(url_for('painel'))



# Função para determinar as datas de início e fim com base no período

def get_dates_for_period(period):

    today = datetime.now().strftime('%Y-%m-%d')

    if period == 'HOJE':

        return today, today

    elif period == '3DIAS':

        return (datetime.now() - timedelta(days=3)).strftime('%Y-%m-%d'), today

    elif period == 'SEMANA':

        return (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'), today

    elif period == 'MES':

        return (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'), today

    elif period == 'ANO':

        return (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d'), today

    else:

        return today, today



# Função para obter a quantidade de tickets por categoria

def get_ticket_counts_by_category(db, start_date, end_date, category='ALL'):

    cursor = db.cursor()

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']

    

    if category != 'ALL':

        cursor.execute('SELECT COUNT(*) FROM tickets WHERE category = ? AND created_at BETWEEN ? AND ?', 

                       (category, start_date, end_date))

        counts = {category: cursor.fetchone()[0]}

    else:

        counts = {}

        for cat in categories:

            cursor.execute('SELECT COUNT(*) FROM tickets WHERE category = ? AND created_at BETWEEN ? AND ?', 

                           (cat, start_date, end_date))

            counts[cat] = cursor.fetchone()[0]

    

    cursor.close()

    return counts



# Função para calcular o desempenho dos recrutadores

def get_recruiter_performance(db, start_date, end_date):

    cursor = db.cursor()

    cursor.execute('SELECT DISTINCT recruiter FROM tickets')

    recruiters = cursor.fetchall()

    

    performance = {}

    for recruiter in recruiters:

        cursor.execute('SELECT COUNT(*) FROM tickets WHERE recruiter = ? AND created_at BETWEEN ? AND ?', 

                       (recruiter[0], start_date, end_date))

        performance[recruiter[0]] = cursor.fetchone()[0]

    

    cursor.close()

    return performance



# Função para calcular o tempo médio de espera e atendimento por categoria

def get_average_times_by_category(db, start_date, end_date, time_type):

    cursor = db.cursor()

    query = f'SELECT AVG({time_type}) FROM tickets WHERE created_at BETWEEN ? AND ? AND category = ?'

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']

    

    average_times = {}

    for category in categories:

        cursor.execute(query, (start_date, end_date, category))

        average_times[category] = cursor.fetchone()[0] or 0

    

    cursor.close()

    return average_times



# Função para comparar tickets emitidos e concluídos por semana/mês

def get_tickets_comparison(db, start_date, end_date):

    cursor = db.cursor()



    # Consulta para tickets emitidos, agrupados por semana usando DATEPART

    query = '''

        SELECT DATEPART(WEEK, created_at) as week, COUNT(*) 

        FROM tickets 

        WHERE created_at BETWEEN ? AND ? 

        GROUP BY DATEPART(WEEK, created_at)

    '''

    cursor.execute(query, (start_date, end_date))

    tickets_issued = cursor.fetchall()



    # Consulta para tickets concluídos, agrupados por semana usando DATEPART

    query_concluido = '''

        SELECT DATEPART(WEEK, created_at) as week, COUNT(*) 

        FROM tickets 

        WHERE status = ? AND created_at BETWEEN ? AND ? 

        GROUP BY DATEPART(WEEK, created_at)

    '''

    cursor.execute(query_concluido, ('CONCLUIDO', start_date, end_date))

    tickets_completed = cursor.fetchall()



    cursor.close()

    return {

        "tickets_issued": [row[1] for row in tickets_issued],

        "tickets_completed": [row[1] for row in tickets_completed]

    }



# Função auxiliar para garantir que nenhum valor no dicionário seja None

def safe_dict(data):

    if isinstance(data, dict):

        return {str(k) if k is not None else 'undefined': (v if v is not None else 0) for k, v in data.items()}

    return data





@app.route('/indicadores', methods=['GET'])

@login_required

def indicadores():

    db = get_sql_server_connection()

    cursor = db.cursor()

    date = request.args.get('date')

    period = request.args.get('period', 'HOJE')

    category = request.args.get('category', 'ALL')

    

    # Definindo categorias

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']



    # Filtrando por data ou período

    if date:

        start_date, end_date = date, date

        period_display = f"Data: {date}"

    else:

        start_date, end_date = get_dates_for_period(period)

        period_display = f"Período: {period}"



    # Consultar contagem total de tickets e total concluído

    cursor.execute('SELECT COUNT(*) FROM tickets WHERE created_at BETWEEN ? AND ?', (start_date, end_date))

    total_tickets = cursor.fetchone()[0] or 0

    

    cursor.execute('SELECT COUNT(*) FROM tickets WHERE status = ? AND created_at BETWEEN ? AND ?', ('CONCLUIDO', start_date, end_date))

    total_concluido = cursor.fetchone()[0] or 0



    # Consultar contagem de tickets por categoria

    counts = {}

    if category != 'ALL':

        cursor.execute('SELECT COUNT(*) FROM tickets WHERE category = ? AND created_at BETWEEN ? AND ?', (category, start_date, end_date))

        counts[category] = cursor.fetchone()[0] or 0

    else:

        for cat in categories:

            cursor.execute('SELECT COUNT(*) FROM tickets WHERE category = ? AND created_at BETWEEN ? AND ?', (cat, start_date, end_date))

            counts[cat] = cursor.fetchone()[0]



    # Consultar tickets chamados e concluídos no intervalo de datas

    cursor.execute('SELECT * FROM tickets WHERE status = ? AND created_at BETWEEN ? AND ?', ('CHAMADO', start_date, end_date))

    called_tickets = cursor.fetchall()

    

    cursor.execute('SELECT * FROM tickets WHERE status = ? AND created_at BETWEEN ? AND ?', ('CONCLUIDO', start_date, end_date))

    concluded_tickets = cursor.fetchall()



    # Inicializar dicionários para tempos de espera e atendimento

    wait_times = {cat: [] for cat in categories}

    service_times = {cat: [] for cat in categories}



    # Preparar os dados históricos

    historical_data = []



    # Calcular tempo de espera e de atendimento por categoria

    for ticket in called_tickets + concluded_tickets:

        created_at = datetime.strptime(ticket[7], '%Y-%m-%d %H:%M:%S')  # Supondo 'created_at' na coluna 7

        called_at = datetime.strptime(ticket[8], '%Y-%m-%d %H:%M:%S') if ticket[8] else None  # Supondo 'called_at' na coluna 8

        concluded_at = datetime.strptime(ticket[9], '%Y-%m-%d %H:%M:%S') if ticket[9] else None  # Supondo 'concluded_at' na coluna 9



        if called_at:

            wait_time = (called_at - created_at).total_seconds()

            if wait_time > 0:

                wait_times[ticket[3]].append(wait_time)  # Supondo 'category' na coluna 3



        if concluded_at and called_at:

            service_time = (concluded_at - called_at).total_seconds()

            if service_time > 0:

                service_times[ticket[3]].append(service_time)



        # Dados históricos

        historical_data.append({

            'ticket_number': ticket[1],  # Supondo 'ticket_number' na coluna 1

            'name': ticket[2],  # Supondo 'name' na coluna 2

            'category': ticket[3],  # Supondo 'category' na coluna 3

            'created_at': format_brazilian_date(ticket[7]),  # 'created_at' na coluna 7

            'called_at': format_brazilian_date(ticket[8]) if called_at else 'Não chamado',

            'concluded_at': format_brazilian_date(ticket[9]) if concluded_at else 'Não concluído',

            'wait_time': format_time(wait_time) if called_at else 'N/A',

            'service_time': format_time(service_time) if concluded_at else 'N/A',

            'total_time': format_time(wait_time + service_time) if concluded_at else 'N/A'

        })



    # Calcular tempos médios

    average_wait_times = {cat: calculate_average_time(wait_times[cat]) for cat in categories}

    average_service_times = {cat: calculate_average_time(service_times[cat]) for cat in categories}



    # Comparação de tickets emitidos e concluídos

    tickets_issued = [total_tickets]

    tickets_completed = [total_concluido]



    # Desempenho de recrutadores

    recruiter_performance = get_recruiter_performance(db, start_date, end_date)



    # Construir resposta para o front-end (JS)

    response_data = {

        'total_tickets': total_tickets,

        'total_concluido': total_concluido,

        'counts': counts,

        'recruiter_performance': safe_dict(recruiter_performance),

        'average_wait_times': safe_dict(average_wait_times),

        'average_service_times': safe_dict(average_service_times),

        'period_display': period_display,

        'historical_data': historical_data,

        'tickets_issued': tickets_issued,

        'tickets_completed': tickets_completed

    }



    cursor.close()



    # Se for requisição AJAX, retorna JSON, caso contrário, renderiza HTML

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

        return jsonify(response_data)

    else:

        return render_template(

            'indicadores.html',

            total_tickets=total_tickets,

            total_concluido=total_concluido,

            counts=counts,

            average_wait_times=average_wait_times,

            average_service_times=average_service_times,

            tickets_issued=tickets_issued,

            tickets_completed=tickets_completed,

            comparison_labels=categories,

            historical_data=historical_data

        )
from flask import request, jsonify
import io
@app.route('/indicadores/data')
@login_required
def indicadores_data():

    import pandas as pd

    from flask import jsonify

    from datetime import datetime, timedelta



    def process_data_for_df(data, columns, force_row_with_zeros=False):

        """

        Processa os dados do SQL Server para garantir que sejam serializáveis para JSON.

        Converte objetos Row em listas/dicionários e lida com valores nulos.

        

        Args:

            data: Dados retornados pela consulta SQL

            columns: Nomes das colunas

            force_row_with_zeros: Se True, retorna uma linha de zeros quando não há dados

            

        Returns:

            Lista de tuplas ou dicionários prontos para conversão em DataFrame

        """

        # Se não há dados ou é None, retorna uma linha de zeros ou lista vazia

        if data is None or len(data) == 0:

            if force_row_with_zeros:

                return [[0] * len(columns)]

            return []



        # Verifica se cada linha tem o número correto de colunas

        valid_data = []

        for row in data:

            # Converte objeto Row em lista

            if hasattr(row, '_fields'):  # Verifica se é um objeto Row

                row_data = list(row)

            elif isinstance(row, tuple) and len(row) == 1:

                # Caso especial: tupla com apenas um elemento

                row_data = [row[0]]

            else:

                row_data = list(row)

                

            # Garante que valores None sejam convertidos para 0 em campos numéricos

            row_data = [0 if val is None else val for val in row_data]

            

            # Verifica se o número de colunas está correto

            if len(row_data) == len(columns):

                valid_data.append(row_data)

            else:

                # Se o número de colunas não bater, preenche com zeros ou trunca

                if len(row_data) < len(columns):

                    row_data.extend([0] * (len(columns) - len(row_data)))

                else:

                    row_data = row_data[:len(columns)]

                valid_data.append(row_data)



        # Se após a validação não restarem dados válidos

        if not valid_data and force_row_with_zeros:

            return [[0] * len(columns)]



        return valid_data





    def ajusta_shape(data, columns, force_row_with_zeros=False):

        """Função de compatibilidade para manter código existente funcionando"""

        return process_data_for_df(data, columns, force_row_with_zeros)



    tipo = request.args.get('tipo', 'quantitativo')

    period = request.args.get('period', 'HOJE')

    date_start = request.args.get('date_start')

    date_end = request.args.get('date_end')

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Lógica de período

    if period == 'PERSONALIZADO' and date_start and date_end:

        start_date, end_date = date_start, date_end

    else:

        start_date, end_date = get_dates_for_period(period)



    # Se for só 1 dia, end_date = start_date + 1

    if start_date == end_date:

        end_date = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")



    if tipo == 'quantitativo':

        query = """

            SELECT 

                COUNT(*) AS total_registros,

                SUM(CASE WHEN situacao = 'Aprovado RH' THEN 1 ELSE 0 END) AS aprovados_rh,

                SUM(CASE WHEN situacao = 'Reprovado RH' THEN 1 ELSE 0 END) AS reprovados_rh,

                SUM(CASE WHEN situacao = 'Aprovado Sindicância' THEN 1 ELSE 0 END) AS aprovados_sindicancia,

                SUM(CASE WHEN situacao = 'Reprovado Sindicância' THEN 1 ELSE 0 END) AS reprovados_sindicancia,

                SUM(CASE WHEN situacao = 'Aprovado Gerência' THEN 1 ELSE 0 END) AS aprovados_gerencia,

                SUM(CASE WHEN situacao = 'Reprovado Gerência' THEN 1 ELSE 0 END) AS reprovados_gerencia,

                SUM(CASE WHEN situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) AS admitidos,

                SUM(CASE WHEN LTRIM(RTRIM(ISNULL(situacao, ''))) = '' OR situacao = 'Não Avaliado' THEN 1 ELSE 0 END) AS nao_avaliado,

                SUM(CASE WHEN situacao = 'Em Verificação' THEN 1 ELSE 0 END) AS em_verificacao,

                SUM(CASE WHEN situacao = 'Em Conversa' THEN 1 ELSE 0 END) AS em_conversa,

                MIN(created_at) AS created_at,

                SUM(CASE 

                        WHEN situacao NOT IN (

                            'Aprovado RH', 'Reprovado RH', 

                            'Aprovado Sindicância', 'Reprovado Sindicância', 

                            'Aprovado Gerência', 'Reprovado Gerência', 

                            'Admitido', 'Não Avaliado', 'Em Verificação', 'Em Conversa'

                        )

                        AND LTRIM(RTRIM(ISNULL(situacao, ''))) <> ''

                    THEN 1 ELSE 0 

                END) AS outros

            FROM registration_form

            WHERE CONVERT(DATETIME, created_at, 120) >= ? AND CONVERT(DATETIME, created_at, 120) < ?

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = ajusta_shape(data, columns, force_row_with_zeros=True)

        df = pd.DataFrame(data, columns=columns)



        # Converte para datetime só os que são válidos!

        df['created_at_dt'] = pd.to_datetime(df['created_at'], errors='coerce')



        # Filtra no Pandas

        mask = (df['created_at_dt'] >= pd.to_datetime(start_date)) & (df['created_at_dt'] < pd.to_datetime(end_date))

        df = df.loc[mask]



        result = df.to_dict(orient='records')



    elif tipo == 'diario':

        query = """

            SELECT 

                CAST(r.created_at AS DATE) AS data_ordenacao,

                CONVERT(VARCHAR, CAST(r.created_at AS DATE), 103) AS data_formatada,

                COUNT(*) AS total_entrevistados,

                SUM(CASE WHEN r.situacao LIKE 'Aprovado%' THEN 1 ELSE 0 END) AS total_aprovados,

                SUM(CASE WHEN r.situacao LIKE 'Reprovado%' THEN 1 ELSE 0 END) AS total_reprovados,

                SUM(CASE WHEN r.situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) AS total_admitidos,

                SUM(CASE WHEN LTRIM(RTRIM(ISNULL(r.situacao, ''))) = '' OR r.situacao = 'Não Avaliado' THEN 1 ELSE 0 END) AS total_sem_avaliacao,

                SUM(CASE WHEN r.situacao = 'Em Verificação' THEN 1 ELSE 0 END) AS total_em_verificacao,

                SUM(CASE WHEN r.situacao = 'Em Conversa' THEN 1 ELSE 0 END) AS total_em_conversa,

                SUM(CASE 

                        WHEN r.situacao NOT LIKE 'Aprovado%' 

                             AND r.situacao NOT LIKE 'Reprovado%' 

                             AND r.situacao NOT LIKE 'Admitido%'

                             AND LTRIM(RTRIM(ISNULL(r.situacao, ''))) <> ''

                             AND r.situacao NOT IN ('Não Avaliado', 'Em Verificação', 'Em Conversa')

                        THEN 1 ELSE 0 

                    END) AS total_outros

            FROM registration_form r

            WHERE r.created_at >= ? AND r.created_at < ?

            GROUP BY CAST(r.created_at AS DATE)

            ORDER BY data_ordenacao

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'semana':

        query = """

            SELECT 

                CASE DATENAME(WEEKDAY, created_at)

                    WHEN 'Monday' THEN 'Segunda-feira'

                    WHEN 'Tuesday' THEN 'Terça-feira'

                    WHEN 'Wednesday' THEN 'Quarta-feira'

                    WHEN 'Thursday' THEN 'Quinta-feira'

                    WHEN 'Friday' THEN 'Sexta-feira'

                    WHEN 'Saturday' THEN 'Sábado'

                    WHEN 'Sunday' THEN 'Domingo'

                END AS dia_da_semana,

                COUNT(*) AS total_entrevistados,

                SUM(CASE WHEN situacao LIKE 'Aprovado%' THEN 1 ELSE 0 END) AS total_aprovados,

                SUM(CASE WHEN situacao LIKE 'Reprovado%' THEN 1 ELSE 0 END) AS total_reprovados

            FROM registration_form

            WHERE created_at >= ? AND created_at < ?

            GROUP BY DATENAME(WEEKDAY, created_at)

            ORDER BY 

                CASE DATENAME(WEEKDAY, created_at)

                    WHEN 'Monday' THEN 1

                    WHEN 'Tuesday' THEN 2

                    WHEN 'Wednesday' THEN 3

                    WHEN 'Thursday' THEN 4

                    WHEN 'Friday' THEN 5

                    WHEN 'Saturday' THEN 6

                    WHEN 'Sunday' THEN 7

                END

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'categoria':

        query = """

            SELECT 

                t.category,

                COUNT(*) AS total_entrevistados,

                SUM(CASE WHEN r.situacao LIKE 'Aprovado%' THEN 1 ELSE 0 END) AS total_aprovados,

                SUM(CASE WHEN r.situacao LIKE 'Reprovado%' THEN 1 ELSE 0 END) AS total_reprovados,

                SUM(CASE WHEN r.situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) AS total_admitidos

            FROM tickets t

            LEFT JOIN registration_form r ON t.cpf = r.cpf

            WHERE t.created_at >= ? AND t.created_at < ?

            GROUP BY t.category

            ORDER BY total_entrevistados DESC

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'aprovados_x_reprovados':

        query = """

            SELECT 

                situacao, 

                COUNT(*) as total

            FROM registration_form

            WHERE (situacao LIKE 'Aprovado%' OR situacao LIKE 'Reprovado%')

              AND created_at >= ? AND created_at < ?

            GROUP BY situacao

            ORDER BY situacao

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = ajusta_shape(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'lista_aprovados':

        query = """

            SELECT 

                nome_completo,

                cpf,

                situacao,

                CONVERT(VARCHAR(10), created_at, 103) AS data,

                CASE MONTH(created_at)

                    WHEN 1 THEN 'Janeiro'

                    WHEN 2 THEN 'Fevereiro'

                    WHEN 3 THEN 'Março'

                    WHEN 4 THEN 'Abril'

                    WHEN 5 THEN 'Maio'

                    WHEN 6 THEN 'Junho'

                    WHEN 7 THEN 'Julho'

                    WHEN 8 THEN 'Agosto'

                    WHEN 9 THEN 'Setembro'

                    WHEN 10 THEN 'Outubro'

                    WHEN 11 THEN 'Novembro'

                    WHEN 12 THEN 'Dezembro'

                END AS mes_extenso

            FROM registration_form

            WHERE situacao LIKE 'Aprovado%' AND created_at >= ? AND created_at < ?

            ORDER BY created_at DESC

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = ajusta_shape(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')
    elif tipo == 'lista_reprovados':

        query = """

            SELECT 

                nome_completo,

                cpf,

                situacao,

                CONVERT(VARCHAR(10), created_at, 103) AS data,

                CASE MONTH(created_at)

                    WHEN 1 THEN 'Janeiro'

                    WHEN 2 THEN 'Fevereiro'

                    WHEN 3 THEN 'Março'

                    WHEN 4 THEN 'Abril'

                    WHEN 5 THEN 'Maio'

                    WHEN 6 THEN 'Junho'

                    WHEN 7 THEN 'Julho'

                    WHEN 8 THEN 'Agosto'

                    WHEN 9 THEN 'Setembro'

                    WHEN 10 THEN 'Outubro'

                    WHEN 11 THEN 'Novembro'

                    WHEN 12 THEN 'Dezembro'

                END AS mes_extenso

            FROM registration_form

            WHERE situacao LIKE 'Reprovado%' AND created_at >= ? AND created_at < ?

            ORDER BY created_at DESC

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'ranking_recrutadores':

        query = """

            SELECT 

                recrutador,

                COUNT(*) AS total_entrevistados,

                SUM(CASE WHEN situacao LIKE 'Aprovado%' THEN 1 ELSE 0 END) AS total_aprovados,

                SUM(CASE WHEN situacao LIKE 'Reprovado%' THEN 1 ELSE 0 END) AS total_reprovados,

                SUM(CASE WHEN situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) AS total_admitidos,

                SUM(CASE WHEN LTRIM(RTRIM(ISNULL(situacao, ''))) = '' OR situacao = 'Não Avaliado' THEN 1 ELSE 0 END) AS total_sem_avaliacao,

                SUM(CASE WHEN situacao = 'Em Verificação' THEN 1 ELSE 0 END) AS total_em_verificacao,

                SUM(CASE WHEN situacao = 'Em Conversa' THEN 1 ELSE 0 END) AS total_em_conversa,

                SUM(CASE 

                        WHEN situacao NOT LIKE 'Aprovado%' 

                             AND situacao NOT LIKE 'Reprovado%' 

                             AND situacao NOT LIKE 'Admitido%'

                             AND LTRIM(RTRIM(ISNULL(situacao, ''))) <> ''

                             AND situacao NOT IN ('Não Avaliado', 'Em Verificação', 'Em Conversa')

                        THEN 1 ELSE 0 

                    END) AS total_outros

            FROM registration_form

            WHERE created_at >= ? AND created_at < ?

            GROUP BY recrutador

            ORDER BY total_aprovados DESC

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'tempo_espera_atendimento':

        query = """

            SELECT 

                t.category,

                AVG(CASE WHEN t.called_at IS NOT NULL THEN DATEDIFF(second, t.created_at, t.called_at) ELSE NULL END) / 60.0 AS tempo_medio_espera_min,

                AVG(CASE WHEN t.called_at IS NOT NULL AND t.concluded_at IS NOT NULL THEN DATEDIFF(second, t.called_at, t.concluded_at) ELSE NULL END) / 60.0 AS tempo_medio_atendimento_min

            FROM tickets t

            WHERE t.created_at >= ? AND t.created_at < ?

            GROUP BY t.category

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'mensal_categoria':

        query = """

            SELECT 

                YEAR(t.created_at) AS ano,

                MONTH(t.created_at) AS mes,

                t.category,

                COUNT(*) AS total_entrevistados,

                SUM(CASE WHEN r.situacao LIKE 'Aprovado%' THEN 1 ELSE 0 END) AS aprovados,

                SUM(CASE WHEN r.situacao LIKE 'Reprovado%' THEN 1 ELSE 0 END) AS reprovados,

                SUM(CASE WHEN r.situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) AS admitidos

            FROM tickets t

            LEFT JOIN registration_form r ON t.cpf = r.cpf

            WHERE t.created_at >= ? AND t.created_at < ?

            GROUP BY YEAR(t.created_at), MONTH(t.created_at), t.category

            ORDER BY ano DESC, mes DESC, t.category

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    elif tipo == 'linha_tempo':

        query = """

            SELECT 

                CONVERT(VARCHAR, created_at, 103) as data,

                COUNT(*) as total_entrevistados,

                SUM(CASE WHEN situacao LIKE 'Aprovado%' THEN 1 ELSE 0 END) as aprovados,

                SUM(CASE WHEN situacao LIKE 'Reprovado%' THEN 1 ELSE 0 END) as reprovados,

                SUM(CASE WHEN situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) as admitidos

            FROM registration_form

            WHERE created_at >= ? AND created_at < ?

            GROUP BY CONVERT(VARCHAR, created_at, 103)

            ORDER BY MIN(created_at)

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        data = process_data_for_df(data, columns)

        df = pd.DataFrame(data, columns=columns)

        result = df.to_dict(orient='records')



    else:

        result = []



    cursor.close()

    db.close()

    return jsonify({'data': result})
def exportar_relatorio_situacao():

    from io import BytesIO

    import pandas as pd

    import pyodbc

    from datetime import datetime



    data_inicial = request.args.get('data_inicial')

    data_final = request.args.get('data_final')

    

    # Convertendo as datas para o formato correto para SQL

    data_inicial_obj = datetime.strptime(data_inicial, '%Y-%m-%d')

    data_final_obj = datetime.strptime(data_final, '%Y-%m-%d')

    

    # Consulta SQL usando os parâmetros de data

    query = """

    SELECT

        nome_completo AS NOME,

        cpf AS CPF,

        recrutador AS RECRUTADOR,

        avaliacao_rh AS AVALIACAO_RH,

        sindicancia AS SINDICANCIA,

        avaliacao_gerencia AS AVALIACAO_GERENCIA,

        admitido AS ADMITIDO,

        situacao AS SITUAÇÃO,

        CONVERT(VARCHAR(10), created_at, 103) AS DATA,

        CASE MONTH(created_at)

            WHEN 1 THEN 'Janeiro'

            WHEN 2 THEN 'Fevereiro'

            WHEN 3 THEN 'Março'

            WHEN 4 THEN 'Abril'

            WHEN 5 THEN 'Maio'

            WHEN 6 THEN 'Junho'

            WHEN 7 THEN 'Julho'

            WHEN 8 THEN 'Agosto'

            WHEN 9 THEN 'Setembro'

            WHEN 10 THEN 'Outubro'

            WHEN 11 THEN 'Novembro'

            WHEN 12 THEN 'Dezembro'

        END AS MES_EXTENSO

    FROM [dbo].[registration_form]

    WHERE created_at BETWEEN ? AND ?

    ORDER BY created_at;

    """

    

    # Executar a consulta e obter os resultados

    conn = get_sql_server_connection()

    

    try:

        # Converter as datas para o formato que o SQL Server espera

        data_inicial_sql = data_inicial_obj.strftime('%Y-%m-%d')

        data_final_sql = data_final_obj.strftime('%Y-%m-%d 23:59:59')

        

        # Usar cursor em vez de pandas diretamente para evitar o aviso

        cursor = conn.cursor()

        cursor.execute(query, [data_inicial_sql, data_final_sql])

        

        # Obter as colunas

        columns = [column[0] for column in cursor.description]

        

        # Obter os dados

        data = cursor.fetchall()

        

        # Criar o DataFrame manualmente

        df = pd.DataFrame([list(row) for row in data], columns=columns)

        

        # Verificar se o DataFrame tem dados

        if df.empty:

            return jsonify({"message": "Nenhum dado encontrado para o período selecionado"}), 404

        

        # Gerar arquivo Excel na memória

        output = BytesIO()

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:

            df.to_excel(writer, sheet_name='Relatório de Situação', index=False)

            

            # Formatar a planilha

            workbook = writer.book

            worksheet = writer.sheets['Relatório de Situação']

            

            # Formato de cabeçalho

            header_format = workbook.add_format({

                'bold': True,

                'text_wrap': True,

                'valign': 'top',

                'fg_color': '#D7E4BC',

                'border': 1

            })

            

            # Aplicar formato aos cabeçalhos

            for col_num, value in enumerate(df.columns.values):

                worksheet.write(0, col_num, value, header_format)

                

            # Ajustar largura das colunas

            for i, col in enumerate(df.columns):

                column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2

                worksheet.set_column(i, i, column_width)

        

        # Preparar para download

        output.seek(0)

        nome_arquivo = f"Relatorio_Situacao_{data_inicial_obj.strftime('%d-%m-%Y')}_a_{data_final_obj.strftime('%d-%m-%Y')}.xlsx"

        

        return send_file(

            output,

            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

            as_attachment=True,

            download_name=nome_arquivo

        )

    

    except Exception as e:

        app.logger.error(f"Erro ao gerar relatório de situação: {str(e)}")

        return jsonify({"error": str(e)}), 500

    finally:

        conn.close()



@app.route('/indicadores/export')

@login_required

def indicadores_export():

    tipo = request.args.get('tipo', 'quantitativo')

    period = request.args.get('period', 'HOJE')

    date_start = request.args.get('date_start')

    date_end = request.args.get('date_end')

    db = get_sql_server_connection()

    cursor = db.cursor()



    if period == 'PERSONALIZADO' and date_start and date_end:

        start_date, end_date = date_start, date_end

    else:

        start_date, end_date = get_dates_for_period(period)



    if start_date == end_date:

        from datetime import datetime, timedelta

        end_date = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")



    if tipo == 'quantitativo':

        query = """

            SELECT 

                COUNT(*) AS total_registros,

                SUM(CASE WHEN situacao = 'Aprovado RH' THEN 1 ELSE 0 END) AS aprovados_rh,

                SUM(CASE WHEN situacao = 'Reprovado RH' THEN 1 ELSE 0 END) AS reprovados_rh,

                SUM(CASE WHEN situacao = 'Aprovado Sindicância' THEN 1 ELSE 0 END) AS aprovados_sindicancia,

                SUM(CASE WHEN situacao = 'Reprovado Sindicância' THEN 1 ELSE 0 END) AS reprovados_sindicancia,

                SUM(CASE WHEN situacao = 'Aprovado Gerência' THEN 1 ELSE 0 END) AS aprovados_gerencia,

                SUM(CASE WHEN situacao = 'Reprovado Gerência' THEN 1 ELSE 0 END) AS reprovados_gerencia,

                SUM(CASE WHEN situacao LIKE 'Admitido%' THEN 1 ELSE 0 END) AS admitidos,

                SUM(CASE WHEN LTRIM(RTRIM(ISNULL(situacao, ''))) = '' OR situacao = 'Não Avaliado' THEN 1 ELSE 0 END) AS nao_avaliado,

                SUM(CASE WHEN situacao = 'Em Verificação' THEN 1 ELSE 0 END) AS em_verificacao,

                SUM(CASE WHEN situacao = 'Em Conversa' THEN 1 ELSE 0 END) AS em_conversa,

                SUM(CASE 

                        WHEN situacao NOT IN (

                            'Aprovado RH', 'Reprovado RH', 

                            'Aprovado Sindicância', 'Reprovado Sindicância', 

                            'Aprovado Gerência', 'Reprovado Gerência', 

                            'Admitido', 'Não Avaliado', 'Em Verificação', 'Em Conversa'

                        )

                        AND LTRIM(RTRIM(ISNULL(situacao, ''))) <> ''

                    THEN 1 ELSE 0 

                END) AS outros

            FROM registration_form

            WHERE created_at >= ? AND created_at < ?

        """

        cursor.execute(query, (start_date, end_date))

        data = cursor.fetchall()

        columns = [desc[0] for desc in cursor.description]

        df = pd.DataFrame(data, columns=columns)

    else:

        df = pd.DataFrame()



    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:

        df.to_excel(writer, index=False, sheet_name='Relatorio')

    output.seek(0)

    cursor.close()

    db.close()

    return send_file(output, as_attachment=True,

                     download_name=f"relatorio_{tipo}.xlsx",

                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')







# Rota para pegar os dados de um dia específico

@app.route('/api/get-day-data')
def get_day_data():

    date = request.args.get('date')  # Data selecionada no calendário

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Puxa os dados desse dia específico

    cursor.execute("""

        SELECT COUNT(*) as total

        FROM tickets

        WHERE CONVERT(DATE, created_at) = ?

    """, (date,))

    tickets = cursor.fetchone()



    cursor.execute("""

        SELECT COUNT(*) as total

        FROM tickets

        WHERE CONVERT(DATE, concluded_at) = ?

    """, (date,))

    completed_tickets = cursor.fetchone()



    # Exemplo de dados para o desempenho dos recrutadores, tempo de espera, etc.

    cursor.execute("""

        SELECT guiche, COUNT(*) as total

        FROM tickets

        WHERE CONVERT(DATE, concluded_at) = ?

        GROUP BY guiche

    """, (date,))

    recruiter_performance = cursor.fetchall()



    cursor.execute("""

        SELECT AVG(DATEDIFF(MINUTE, created_at, called_at)) as avg_wait_time

        FROM tickets

        WHERE CONVERT(DATE, created_at) = ?

    """, (date,))

    average_wait_time = cursor.fetchone()



    cursor.execute("""

        SELECT AVG(DATEDIFF(MINUTE, called_at, concluded_at)) as avg_service_time

        FROM tickets

        WHERE CONVERT(DATE, created_at) = ?

    """, (date,))

    average_service_time = cursor.fetchone()



    cursor.close()

    db.close()



    # Preparar os dados para os gráficos

    data = {

        'tickets': tickets[0],

        'completedTickets': completed_tickets[0],

        'recruiterPerformance': [dict(guiche=row[0], total=row[1]) for row in recruiter_performance],

        'averageWaitTime': average_wait_time[0] if average_wait_time[0] is not None else 0,

        'averageServiceTime': average_service_time[0] if average_service_time[0] is not None else 0,

        # Adicione mais dados para os outros gráficos conforme necessário

    }



    return jsonify(data)



# Rota para os eventos do calendário

@app.route('/api/get-calendar-events')

def get_calendar_events():

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Agrupa os tickets por data e conta quantos foram criados por dia

    cursor.execute("""

        SELECT CONVERT(DATE, created_at) as date, COUNT(*) as total

        FROM tickets

        GROUP BY CONVERT(DATE, created_at)

    """)

    events = cursor.fetchall()

    cursor.close()

    db.close()



    # Converter os eventos para o formato FullCalendar com cores baseadas na quantidade de tickets

    calendar_events = []

    for event in events:

        total = event[1]

        color = ""



        # Definir cores com base na quantidade de tickets

        if total <= 20:

            color = 'green'  # Poucos tickets

        elif 20 < total <= 30:

            color = '#d4a003'  # Quantidade média

        else:

            color = 'red'  # Muitos tickets



        calendar_events.append({

            'title': f'{total} tickets',

            'start': event[0].strftime('%Y-%m-%d'),

            'color': color  # Cor do evento com base na quantidade

        })



    return jsonify(calendar_events)
# Rota para os indicadores diários
@app.route('/api/get-indicators')
def get_indicators():

    date = request.args.get('date')

    category_filter = request.args.get('category')

    period = request.args.get('period')

    

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Categorias disponíveis

    categories = ['Admissão', 'Demissão', 'Entrevista', 'Agendado', 'Treinamento', 'Documentação', 'Outros']

    

    # Aplicar filtro de categoria se necessário

    filtered_categories = [category_filter] if category_filter and category_filter in categories else categories

    

    # Definir intervalo de datas com base no período

    start_date = date

    end_date = date  # Por padrão, mesmo dia

    

    if period:

        today = datetime.now().strftime('%Y-%m-%d')

        if period == '3DIAS':

            # Últimos 3 dias a partir de hoje

            start_date = (datetime.now() - timedelta(days=3)).strftime('%Y-%m-%d')

            end_date = today

        elif period == 'SEMANA':

            # Última semana a partir de hoje

            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

            end_date = today

        elif period == 'MES':

            # Último mês a partir de hoje

            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

            end_date = today

        elif period == 'ANO':

            # Último ano a partir de hoje

            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')

            end_date = today



    # 1. Tickets Emitidos por Categoria (created_at no dia selecionado)

    tickets_issued = []

    for category in categories:

        if category in filtered_categories or not category_filter or category_filter not in categories:

            if period:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE category = ? AND CONVERT(DATE, created_at) BETWEEN ? AND ?",

                    (category, start_date, end_date)

                ).fetchone()[0]

            else:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE category = ? AND CONVERT(DATE, created_at) = ?",

                    (category, date)

                ).fetchone()[0]

            tickets_issued.append(count)

        else:

            tickets_issued.append(0)



    # 2. Tickets Concluídos por Categoria (concluded_at no dia selecionado)

    tickets_completed = []

    for category in categories:

        if category in filtered_categories or not category_filter or category_filter not in categories:

            if period:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE category = ? AND CONVERT(DATE, concluded_at) BETWEEN ? AND ?",

                    (category, start_date, end_date)

                ).fetchone()[0]

            else:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE category = ? AND CONVERT(DATE, concluded_at) = ?",

                    (category, date)

                ).fetchone()[0]

            tickets_completed.append(count)

        else:

            tickets_completed.append(0)



    # 3. Recrutadores e Performance por Recrutador (guichê) - Total de tickets atendidos por guichê no dia

    # Se houver filtro de categoria, filtra os recruters também

    recruiter_query = "SELECT DISTINCT guiche FROM tickets WHERE guiche IS NOT NULL"

    if category_filter and category_filter in categories:

        recruiter_query += " AND category = ?"

        cursor.execute(recruiter_query, (category_filter,))

    else:

        cursor.execute(recruiter_query)

    

    recruiters = cursor.fetchall()

    

    recruiter_performance = []

    for recruiter in recruiters:

        if category_filter and category_filter in categories:

            if period:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE guiche = ? AND category = ? AND CONVERT(DATE, concluded_at) BETWEEN ? AND ?",

                    (recruiter[0], category_filter, start_date, end_date)

                ).fetchone()[0]

            else:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE guiche = ? AND category = ? AND CONVERT(DATE, concluded_at) = ?",

                    (recruiter[0], category_filter, date)

                ).fetchone()[0]

        else:

            if period:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE guiche = ? AND CONVERT(DATE, concluded_at) BETWEEN ? AND ?",

                    (recruiter[0], start_date, end_date)

                ).fetchone()[0]

            else:

                count = cursor.execute(

                    "SELECT COUNT(*) FROM tickets WHERE guiche = ? AND CONVERT(DATE, concluded_at) = ?",

                    (recruiter[0], date)

                ).fetchone()[0]

        recruiter_performance.append(count)



    # 4. Tempo Médio de Espera por Categoria (chamada - criação)

    average_wait_times = []

    for category in categories:

        if category in filtered_categories or not category_filter or category_filter not in categories:

            if period:

                avg_time = cursor.execute(

                    "SELECT AVG(DATEDIFF(MINUTE, created_at, called_at)) AS avg_wait_time "

                    "FROM tickets WHERE category = ? AND CONVERT(DATE, created_at) BETWEEN ? AND ? AND called_at IS NOT NULL",

                    (category, start_date, end_date)

                ).fetchone()[0] or 0

            else:

                avg_time = cursor.execute(

                    "SELECT AVG(DATEDIFF(MINUTE, created_at, called_at)) AS avg_wait_time "

                    "FROM tickets WHERE category = ? AND CONVERT(DATE, created_at) = ? AND called_at IS NOT NULL",

                    (category, date)

                ).fetchone()[0] or 0

            average_wait_times.append(avg_time)

        else:

            average_wait_times.append(0)



    # 5. Tempo Médio de Atendimento por Categoria (conclusão - chamada)

    average_service_times = []

    for category in categories:

        if category in filtered_categories or not category_filter or category_filter not in categories:

            if period:

                avg_time = cursor.execute(

                    "SELECT AVG(DATEDIFF(MINUTE, called_at, concluded_at)) AS avg_service_time "

                    "FROM tickets WHERE category = ? AND CONVERT(DATE, called_at) BETWEEN ? AND ? AND concluded_at IS NOT NULL",

                    (category, start_date, end_date)

                ).fetchone()[0] or 0

            else:

                avg_time = cursor.execute(

                    "SELECT AVG(DATEDIFF(MINUTE, called_at, concluded_at)) AS avg_service_time "

                    "FROM tickets WHERE category = ? AND CONVERT(DATE, called_at) = ? AND concluded_at IS NOT NULL",

                    (category, date)

                ).fetchone()[0] or 0

            average_service_times.append(avg_time)

        else:

            average_service_times.append(0)



    # 6. Tempo de Espera vs. Tempo de Atendimento (dados para gráfico comparativo)

    wait_times = average_wait_times  # Mesmos dados do tempo médio de espera

    service_times = average_service_times  # Mesmos dados do tempo médio de atendimento



    cursor.close()

    db.close()



    # Retornar os dados em JSON para os gráficos

    return jsonify({

        'categories': categories,

        'ticketsIssued': tickets_issued,

        'ticketsCompleted': tickets_completed,

        'recruiters': [r[0] for r in recruiters],

        'recruiterPerformance': recruiter_performance,

        'averageWaitTimes': average_wait_times,

        'averageServiceTimes': average_service_times,

        'waitTimes': wait_times,

        'serviceTimes': service_times,

        'period': period,

        'startDate': start_date,

        'endDate': end_date

    })



from datetime import datetime



@app.route('/display')

def display():

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Consulta para o ticket atual

    cursor.execute(

        'SELECT TOP 1 * FROM tickets WHERE status = ? ORDER BY priority DESC, called_at DESC',

        ('CHAMADO',)

    )

    current_ticket = cursor.fetchone()



    # Lidar com o ticket atual de acordo com os índices das colunas

    current_ticket_number = current_ticket[2] if current_ticket else "N/A"  # 'ticket_number'

    current_guiche = current_ticket[8] if current_ticket else "N/A"  # 'guiche'

    current_name = current_ticket[0] if current_ticket else "N/A"  # 'name'



    # Validar e formatar 'called_at'

    if current_ticket and current_ticket[4]:  # 'called_at'

        current_time = datetime.strptime(current_ticket[4], '%Y-%m-%d %H:%M:%S').strftime('%H:%M:%S')

    else:

        current_time = 'N/A'



    # Consulta para os últimos 10 tickets chamados

    cursor.execute(

        'SELECT TOP 10 * FROM tickets WHERE status = ? ORDER BY priority DESC, called_at DESC',

        ('CHAMADO',)

    )

    called_tickets = cursor.fetchall()



    # Construir a lista de tickets chamados

    called_tickets_list = []

    for ticket in called_tickets:

        if ticket[6] != 'CONCLUIDO':  # 'status'

            called_tickets_list.append({

                'name': first_and_second_name(ticket[0]),  # 'name'

                'ticket_number': ticket[2],  # 'ticket_number'

                'guiche': ticket[8],  # 'guiche'

                'called_at': datetime.strptime(ticket[4], '%Y-%m-%d %H:%M:%S').strftime('%H:%M:%S') if ticket[4] else 'N/A'

            })



    cursor.close()

    db.close()



    # Renderizar o template com os dados ajustados

    return render_template(

        'display.html',

        current_ticket_number=current_ticket_number,

        current_guiche=current_guiche,

        current_name=current_name,

        current_time=current_time,

        called_tickets=called_tickets_list

    )







@app.route('/display_stage/<stage>')

@login_required

def display_stage(stage):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Em andamento

    cursor.execute('SELECT * FROM tickets WHERE stage != ? AND status != ? ORDER BY priority DESC, updated_at DESC', (stage, 'CONCLUIDO'))

    em_andamento = cursor.fetchall()

    em_andamento_list = [{'name': ticket[2], 'stage': ticket[5], 'status': ticket[7], 'updated_at': format_brazilian_date(ticket[12])} for ticket in em_andamento]



    # No setor

    cursor.execute('SELECT * FROM tickets WHERE stage = ? AND status != ? ORDER BY priority DESC, updated_at DESC', (stage, 'CONCLUIDO'))

    no_setor = cursor.fetchall()

    no_setor_list = [{'name': ticket[2], 'stage': ticket[5], 'status': ticket[7], 'updated_at': format_brazilian_date(ticket[12])} for ticket in no_setor]



    # Concluídos

    cursor.execute('SELECT * FROM tickets WHERE stage = ? AND status = ? ORDER BY priority DESC, updated_at DESC', (stage, 'CONCLUIDO'))

    concluidos = cursor.fetchall()

    concluidos_list = [{'name': ticket[2], 'stage': ticket[5], 'status': ticket[7], 'updated_at': format_brazilian_date(ticket[12])} for ticket in concluidos]



    cursor.close()

    db.close()



    return render_template('display_stage.html', em_andamento=em_andamento_list, no_setor=no_setor_list, concluidos=concluidos_list, stage=stage)



@app.route('/historico_completo')

@login_required

def historico_completo():

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Consulta para todos os tickets

    cursor.execute('SELECT * FROM tickets ORDER BY updated_at DESC')

    tickets = cursor.fetchall()

    tickets_list = [

        {

            'id': ticket[0],

            'name': ticket[2],

            'category': ticket[4],

            'ticket_number': ticket[3],

            'created_at': format_brazilian_date(ticket[7]),

            'called_at': format_brazilian_date(ticket[8]) if ticket[8] else 'Não chamado',

            'concluded_at': format_brazilian_date(ticket[9]) if ticket[9] else 'Não concluído',

            'status': ticket[7],

            'guiche': ticket[10] if ticket[10] else 'Não definido',

            'recruiter': ticket[11] if ticket[11] else 'Não definido',

            'cpf': ticket[13],

            'cep': ticket[14],

            'rua': ticket[15],

            'numero': ticket[16],

            'complemento': ticket[17],

            'bairro': ticket[18],

            'cidade': ticket[19],

            'telefones': ticket[20]

        }

        for ticket in tickets

    ]



    cursor.close()

    db.close()



    return render_template('historico_completo.html', tickets=tickets_list)
def save_registration_form_from_ticket(ticket):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Captura os dados do ticket e mapeia para a estrutura da tabela registration_form

    form_data = {

        'cpf': ticket['cpf'].upper() if ticket['cpf'] else '',

        'nome_completo': ticket['name'].upper() if ticket['name'] else '',

        'cargo_pretendido': ticket['category'].upper() if ticket['category'] else '',

        'cep': ticket['cep'] if ticket['cep'] else '',

        'endereco': ticket['rua'].upper() if ticket['rua'] else '',

        'numero': ticket['numero'] if ticket['numero'] else '',

        'complemento': ticket['complemento'].upper() if ticket['complemento'] else '',

        'bairro': ticket['bairro'].upper() if ticket['bairro'] else '',

        'cidade': ticket['cidade'].upper() if ticket['cidade'] else '',

        'telefone': ticket['telefones'] if ticket['telefones'] else '',

        'situacao': 'Admitido',

        'avaliacao_rh': 'Aprovado',

        'admitido': 'Sim',

        'created_at': datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S'),

        'recrutador': current_user.username

    }



    # Verifica se já existe um registro para o CPF fornecido

    cursor.execute('SELECT id FROM registration_form WHERE cpf = ?', (form_data['cpf'],))

    existing_entry = cursor.fetchone()



    if existing_entry:

        # Atualiza registro existente

        update_fields = ', '.join([f"{key} = ?" for key in form_data.keys() if key not in ['created_at', 'recrutador']])

        values = [form_data[key] for key in form_data.keys() if key not in ['created_at', 'recrutador']]

        values.append(form_data['cpf'])  # Adiciona o CPF ao final para a cláusula WHERE



        cursor.execute(f'''

            UPDATE registration_form

            SET {update_fields}, last_updated = GETDATE()

            WHERE cpf = ?

        ''', values)

    else:

        # Insere novo registro

        columns = ', '.join(form_data.keys())

        placeholders = ', '.join(['?' for _ in form_data])

        values = list(form_data.values())

        

        cursor.execute(f'''

            INSERT INTO registration_form ({columns}, created_at, last_updated)

            VALUES ({placeholders}, GETDATE(), GETDATE())

        ''', values)



    db.commit()

    cursor.close()

    db.close()
@app.route('/conclude_ticket/<int:id>', methods=['POST'])
def conclude_ticket(id):

    db = get_sql_server_connection()

    cursor = db.cursor()

    concluded_at = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')



    try:

        # 📌 Buscar todas as informações necessárias do ticket

        cursor.execute('''

            SELECT id, ticket_number, name, category, cpf, guiche, created_at, called_at 

            FROM tickets WHERE id = ?

        ''', (id,))

        ticket = cursor.fetchone()



        if not ticket:

            return jsonify(success=False, message="Ticket não encontrado"), 404



        # ✅ Desempacotando os valores corretamente

        ticket_id, ticket_number, name, category, cpf, guiche, created_at, called_at = ticket



        # ✅ Substituir valores None por "Não informado"

        created_at = created_at if created_at else "Não informado"

        called_at = called_at if called_at else "Não chamado"



        # 📌 Atualizar o status do ticket para "CONCLUÍDO"

        cursor.execute(

            'UPDATE tickets SET status = ?, concluded_at = ? WHERE id = ?',

            ('CONCLUIDO', concluded_at, id)

        )



        # 📌 Atualizar ou criar ficha no `registration_form` se for "Admissão"

        if category == 'Admissão' and cpf:

            cpf_normalizado = re.sub(r'[.-]', '', cpf)



            cursor.execute('''

                SELECT id FROM registration_form 

                WHERE REPLACE(REPLACE(cpf, '.', ''), '-', '') = ?

            ''', (cpf_normalizado,))

            existing_form = cursor.fetchone()



            if existing_form:

                cursor.execute('''

                    UPDATE registration_form

                    SET admitido = 'Sim', situacao = 'Admitido', last_updated = GETDATE()

                    WHERE id = ?

                ''', (existing_form[0],))

            else:

            



                cursor.execute('''

                    INSERT INTO registration_form (

                        cpf, nome_completo, created_at, admitido, situacao, last_updated, recrutador

                    ) VALUES (?, ?, ?, ?, ?, GETDATE(), ?)

                ''', (cpf, name, concluded_at, 'Sim', 'Concluído', 'Sistema'))



        db.commit()  # ✅ Salvar todas as mudanças no banco de dados



        # 📌 Garantir que os dados enviados para o WebSocket estejam completos

        ticket_data = {

            'id': ticket_id,

            'ticket_number': ticket_number,  # ✅ Agora corretamente definido

            'name': name,

            'category': category if category else "Sem Categoria",

            'status': 'CONCLUIDO',

            'guiche': guiche if guiche else "Nenhum",  # 🔥 Adicionado o GUICHÊ

            'created_at': created_at,  # ✅ Criado em

            'called_at': called_at,  # ✅ Chamado em

            'concluded_at': concluded_at  # ✅ Concluído em

        }



        # 📌 Emitir evento WebSocket com os dados completos

        socketio.emit('update_ticket_concluded', ticket_data, namespace='/')



        return jsonify(success=True, ticket=ticket_data)



    except Exception as e:

        db.rollback()  # 🛑 Desfazer alterações em caso de erro

        return jsonify(success=False, message=f"Erro ao concluir ticket: {str(e)}"), 500



    finally:

        cursor.close()

        db.close()









@app.route('/sistema_rs')

@login_required

def sistema_rs():

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Consultas para diferentes categorias de candidatos

    cursor.execute('SELECT * FROM tickets WHERE category = ?', ('Admissão',))

    admission_candidates = cursor.fetchall()



    cursor.execute('SELECT * FROM tickets WHERE category = ?', ('Demissão',))

    dismissal_candidates = cursor.fetchall()



    cursor.execute('SELECT * FROM tickets WHERE category = ?', ('Entrevista',))

    interview_candidates = cursor.fetchall()



    cursor.close()

    db.close()



    return render_template('sistema_rs.html', 

                           admission_candidates=admission_candidates, 

                           dismissal_candidates=dismissal_candidates, 

                           interview_candidates=interview_candidates)



@app.route('/gestao_pessoas')

@login_required

def gestao_pessoas():

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Data de ontem

    yesterday = (datetime.now(BRASILIA_TZ) - timedelta(days=1)).strftime('%Y-%m-%d')



    # Função para carregar os tickets com a lógica de botão de conclusão

    def load_tickets(category, recruiter=None):

        query = '''

            SELECT *, 

                   CASE 

                       WHEN CONVERT(DATE, created_at) <= ? AND status != 'CONCLUIDO' THEN 1 

                       ELSE 0 

                   END AS show_conclude_button 

            FROM tickets 

            WHERE category = ? AND status != ? 

        '''

        params = [yesterday, category, 'CONCLUIDO']

        if recruiter:

            query += ' AND recruiter = ?'

            params.append(recruiter)



        query += ' ORDER BY created_at ASC'

        cursor.execute(query, params)

        return cursor.fetchall()



    # Carregando tickets por categoria

    admission_candidates = load_tickets('Admissão')

    dismissal_candidates = load_tickets('Demissão')

    interview_candidates = load_tickets('Entrevista')

    scheduled_candidates = load_tickets('Agendado', recruiter=current_user.username)



    cursor.close()

    db.close()



    return render_template('gestao_pessoas.html',

                           admission_candidates=admission_candidates,

                           dismissal_candidates=dismissal_candidates,

                           interview_candidates=interview_candidates,

                           scheduled_candidates=scheduled_candidates)



@app.route('/conclude_ticket_inline/<int:id>', methods=['POST'])

@login_required

def conclude_ticket_inline(id):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Atualiza o ticket como concluído

    concluded_at = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')

    cursor.execute('UPDATE tickets SET status = ?, concluded_at = ? WHERE id = ?', 

                   ('CONCLUIDO', concluded_at, id))

    db.commit()



    # Fecha a conexão

    cursor.close()

    db.close()



    # Retorna JSON para o front-end

    return jsonify({'success': True, 'ticket_id': id})



def format_time(datetime_string):

    """Valida e converte strings de data para o formato HH:MM ou retorna 'N/A' para valores inválidos."""

    try:

        # Tenta converter a string para um objeto datetime

        dt = datetime.strptime(datetime_string, '%Y-%m-%d %H:%M:%S')

        return dt.strftime('%H:%M')  # Retorna apenas a hora e os minutos

    except (ValueError, TypeError):

        return 'N/A'  # Retorna 'N/A' para valores inválidos



def calculate_average_time(time_differences):

    if not time_differences:

        return 0

    return sum(time_differences) / len(time_differences)



def generate_ticket_number(category):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Dicionário de prefixos por categoria

    category_prefix_map = {

        'Agendado': 'AG',

        'Admissão': 'AD',

        'Demissão': 'DE',

        'Entrevista': 'EN',

        'Treinamento': 'TR',

        'Documentação': 'DO',

        'Outros': 'OU'

    }



    # Obtém o prefixo da categoria ou usa a primeira letra se não estiver no mapa

    category_prefix = category_prefix_map.get(category, category[0].upper())



    # Conta quantos tickets já existem na categoria e incrementa 1

    cursor.execute('SELECT COUNT(*) FROM tickets WHERE category = ?', (category,))

    count = cursor.fetchone()[0] + 1



    # Formata o número do ticket com 3 dígitos, como AD001, DE002, etc.

    ticket_number = f"{category_prefix}{count:03d}"



    cursor.close()

    db.close()



    return ticket_number



def get_candidato_by_id(id):

    db = get_sql_server_connection()

    cursor = db.cursor()

    cursor.execute("SELECT * FROM registration_form WHERE id = ?", (id,))

    candidato = cursor.fetchone()

    cursor.close()

    db.close()

    return candidato



@app.route('/get_registration_data', methods=['POST'])

@login_required

def get_registration_data():

    db = get_sql_server_connection()

    cursor = db.cursor()

    selected_names = request.json.get('selected_names', [])



    # Montar a consulta com placeholders para os nomes selecionados

    placeholders = ', '.join('?' for _ in selected_names)

    query = f'''

        SELECT nome_completo, cpf, data_nasc

        FROM registration_form

        WHERE nome_completo IN ({placeholders})

    '''

    cursor.execute(query, selected_names)

    result = cursor.fetchall()



    def format_date(value):

        """Formata datas no formato DD/MM/YYYY"""

        if isinstance(value, date):

            return value.strftime('%d/%m/%Y')

        elif isinstance(value, str):

            try:

                parsed_date = datetime.strptime(value, '%Y-%m-%d')

                return parsed_date.strftime('%d/%m/%Y')

            except ValueError:

                return "Data inválida"

        return "Data não disponível"



    data = []

    for row in result:

        nome_completo = row[0]

        cpf = row[1]

        data_nasc = format_date(row[2]) 

        data.append({

            'nome_completo': nome_completo,

            'cpf': cpf,

            'data_nasc': data_nasc

        })



    cursor.close()

    db.close()



    return jsonify(data)
def export_excel(cpf):

    db = get_sql_server_connection()

    cursor = db.cursor()



    cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

    candidato = cursor.fetchone()



    if candidato is None:

        cursor.close()

        db.close()

        return "Candidato não encontrado", 404



    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')



    try:

        with pd.ExcelWriter(temp_file.name, engine='xlsxwriter') as writer:

            workbook = writer.book

            worksheet = workbook.add_worksheet('Ficha')



            title_format = workbook.add_format({

                'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#F2F2F2'

            })

            header_format = workbook.add_format({

                'bold': True, 'bg_color': '#AB1A18', 'font_color': '#FFFFFF', 'border': 1, 'align': 'center', 'valign': 'vcenter'

            })

            text_format = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'})



            worksheet.set_column('A:A', 30)

            worksheet.set_column('B:B', 50)



            fields = {

                'Nome Completo': candidato.nome_completo,

                'Estado Civil': candidato.estado_civil,

                'Telefone': candidato.telefone,

                'Data de Nascimento': candidato.data_nasc,

                'Idade': candidato.idade,

                'Gênero': candidato.genero,

                'Fumante': candidato.fumante,

                'Bebida': candidato.bebida,

                'Alergia': candidato.alergia,

                'Medicamento de Uso Constante': candidato.medicamento_constante,

                'Qual Medicamento': candidato.qual_medicamento,

                'Escolaridade': candidato.escolaridade,

                'Número de Filhos': candidato.numero_filhos,

                'Peso': candidato.peso,

                'Cor da Pele': candidato.cor_pele,

                'Tatuagem': candidato.tatuagem,

                'PCD (Pessoa com Deficiência)': candidato.pcd,

                'Perfil': candidato.perfil,

                'CEP': candidato.cep,

                'Endereço': candidato.endereco,

                'Número': candidato.numero,

                'Complemento': candidato.complemento,

                'Bairro': candidato.bairro,

                'Cidade': candidato.cidade,

                'Estado': candidato.estado_nasc,

                'CPF': candidato.cpf,

                'Cursos Realizados': candidato.cursos_realizados,

                'Conhecimento de Digitação': candidato.conhecimento_digitacao,

                'Empresa 1': candidato.empresa1,

                'Função 1': candidato.funcao1,

                'Motivo de Saída 1': candidato.motivo_saida1,

                'Tempo de Permanência 1': f"{candidato.tempo_permanencia1_anos} anos, {candidato.tempo_permanencia1_meses} meses",

                'Salário 1': candidato.salario1,

                'Descrição de Atividades 1': candidato.atividades_empresa1,

                'Empresa 2': candidato.empresa2,

                'Função 2': candidato.funcao2,

                'Motivo de Saída 2': candidato.motivo_saida2,

                'Tempo de Permanência 2': f"{candidato.tempo_permanencia2_anos} anos, {candidato.tempo_permanencia2_meses} meses",

                'Salário 2': candidato.salario2,

                'Descrição de Atividades 2': candidato.atividades_empresa2,

                'Empresa 3': candidato.empresa3,

                'Função 3': candidato.funcao3,

                'Motivo de Saída 3': candidato.motivo_saida3,

                'Tempo de Permanência 3': f"{candidato.tempo_permanencia3_anos} anos, {candidato.tempo_permanencia3_meses} meses",

                'Salário 3': candidato.salario3,

                'Descrição de Atividades 3': candidato.atividades_empresa3,

                'Empregos Informais': candidato.empregos_informais,

                'Regiões de Preferência': candidato.regioes_preferencia,

                'Disponibilidade de Horário': candidato.disponibilidade_horario,

                'Avaliação RH': candidato.avaliacao_rh,

                'Avaliação Gerência': candidato.avaliacao_gerencia,

                'Observações': candidato.observacoes,

            }



            worksheet.merge_range('A4:B4', 'Ficha do Candidato', title_format)

            row = 5

            for field, value in fields.items():

                worksheet.write(row, 0, field, header_format)

                worksheet.write(row, 1, value if value else '', text_format)

                row += 1



        cursor.close()

        db.close()



        return send_file(temp_file.name, as_attachment=True, download_name=f'candidato_{cpf}.xlsx')



    except Exception as e:

        cursor.close()

        db.close()

        return str(e), 500



@app.route('/sistema_os')

def sistema_os():

    return redirect('http://192.168.0.79:8080/')



@app.route('/get_data_nasc/<cpf>')

def get_data_nasc(cpf):

    # Conectar ao SQL Server e buscar o candidato

    db = get_sql_server_connection()

    cursor = db.cursor()

    

    cursor.execute("SELECT data_nasc FROM registration_form WHERE cpf = ?", (cpf,))

    candidato = cursor.fetchone()

    

    cursor.close()

    db.close()



    if candidato and candidato[0]:

        data_nasc = candidato[0].strftime('%d/%m/%Y')

        return jsonify({'data_nasc': data_nasc})

    

    return jsonify({'error': 'Candidato não encontrado'}), 404



   

@app.route('/view_registration/<cpf>')

@login_required

def view_registration(cpf):

    db = get_sql_server_connection()

    cursor = db.cursor()



    # Primeiro tenta buscar na tabela registration_form

    cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

    form_data = cursor.fetchone()

    

    if form_data:

        # Obtém os nomes das colunas

        column_names = [column[0] for column in cursor.description]

        # Converte os resultados para um dicionário usando zip

        form_data = dict(zip(column_names, form_data))

        form_data['situacao'] = determinar_situacao(form_data)

    else:

        # Se não encontrar na registration_form, busca no tickets

        cursor.execute('SELECT * FROM tickets WHERE cpf = ?', (cpf,))

        ticket = cursor.fetchone()

        if ticket:

            # Cria form_data com os dados do ticket

            form_data = {

                'cpf': ticket[12],

                'nome_completo': ticket[1],  # Nome do ticket

                'cep': ticket[13],

                'endereco': ticket[14],

                'numero': ticket[15],

                'complemento': ticket[16],

                'bairro': ticket[17],

                'cidade': ticket[18],

                'telefone': ticket[19],

                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),

                'recrutador': current_user.username,

                'situacao': None,

                'avaliacao_rh': '',

                'avaliacao_gerencia': ''

            }

        else:

            # Se não encontrar em nenhuma tabela, cria form_data básico

            form_data = {

                'cpf': cpf,

                'nome_completo': '',

                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),

                'recrutador': current_user.username,

                'situacao': None,

                'avaliacao_rh': '',

                'avaliacao_gerencia': ''

            }

    

    cursor.close()

    db.close()



    return render_template('view_registration.html', form_data=form_data, form=form_data)







@app.route('/view_form/<cpf>', methods=['GET', 'POST'])

@login_required

def view_form(cpf):

    db = get_sql_server_connection()

    cursor = db.cursor()



    tz_brasilia = pytz.timezone('America/Sao_Paulo')

    current_date = datetime.now(tz_brasilia).strftime('%Y-%m-%d %H:%M:%S')



    # Busca o registro com o CPF fornecido

    cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

    form = cursor.fetchone()



    if not form:

        # Caso o formulário não exista, busca no tickets

        cursor.execute('SELECT * FROM tickets WHERE cpf = ?', (cpf,))

        ticket = cursor.fetchone()

        if ticket:

            # Cria `form_data` com os dados do ticket

            form_data = {

                'cpf': ticket[12],

                'nome_completo': ticket[1],

                'cep': ticket[13],

                'endereco': ticket[14],

                'numero': ticket[15],

                'complemento': ticket[16],

                'bairro': ticket[17],

                'cidade': ticket[18],

                'telefone': ticket[19],

                'cargo_pretendido': [],

                'cargo_indicado': [],

                'regioes_preferencia': [],

                'created_at': current_date,

                # Removida a atribuição automática de recrutador

                'recrutador': '',

                'situacao': None,

                'avaliacao_rh': '',

                'avaliacao_gerencia': '',

                'data_nasc': ticket[23]

            }

        else:

            # Se não existir também no tickets, cria form_data com campos básicos

            form_data = {

                'cpf': cpf,

                'created_at': current_date,

                # Removida a atribuição automática de recrutador

                'recrutador': '',

                'situacao': None,

                'avaliacao_rh': '',

                'avaliacao_gerencia': ''

            }

    else:

        # Converte os dados do formulário existente em um dicionário

        form_data = dict(zip([column[0] for column in cursor.description], form))

        form_data['cargo_pretendido'] = form_data['cargo_pretendido'].split(',') if form_data.get('cargo_pretendido') else []

        form_data['cargo_indicado'] = form_data['cargo_indicado'].split(',') if form_data.get('cargo_indicado') else []

        form_data['regioes_preferencia'] = form_data['regioes_preferencia'].split(',') if form_data.get('regioes_preferencia') else []



        # Adiciona campos para filhos, se existentes

        for i in range(1, 11):

            idade_filho_key = f'idade_filho_{i}'

            form_data[idade_filho_key] = form_data.get(idade_filho_key, '')



    # Verifica permissão para alterar o formulário se estiver em "Em Conversa"

    if form and form_data['situacao'] == 'Em Conversa' and current_user.username != form_data['recrutador']:

        flash('Você não tem permissão para alterar esta ficha, pois está em conversa.', 'danger')

        return render_template('view_registration.html', form_data=form_data, situacao=form_data.get('situacao'), current_date=current_date)



    if request.method == 'POST':

        # Atualiza `form_data` com dados do formulário enviado

        form_data.update(request.form.to_dict())

        form_data['last_updated'] = datetime.now(tz_brasilia).strftime('%Y-%m-%d %H:%M:%S')



        admitido = form_data.get('admitido', '')



        avaliacao_rh = form_data.get('avaliacao_rh', '')

        avaliacao_gerencia = form_data.get('avaliacao_gerencia', '')



        # Define a situação baseada nas avaliações

        if admitido == 'Sim':

            form_data['situacao'] = 'Admitido'

        elif avaliacao_rh == 'Reprovado' or avaliacao_gerencia == 'Reprovado':

            form_data['situacao'] = 'Reprovado'

        elif avaliacao_rh == 'Aprovado' and avaliacao_gerencia == 'Aprovado':

            form_data['situacao'] = 'Aprovado'

        elif avaliacao_rh == 'Aprovado' and avaliacao_gerencia == 'Em Conversa':

            form_data['situacao'] = 'Em Conversa'

        else:

            form_data['situacao'] = 'Não Avaliado'



        # Insere ou atualiza o formulário no banco de dados

        if not form:

            columns = ', '.join(form_data.keys())

            placeholders = ', '.join(['?'] * len(form_data))

            cursor.execute(f'INSERT INTO registration_form ({columns}) VALUES ({placeholders})', list(form_data.values()))

            flash('Formulário criado com sucesso!', 'success')

        else:

            update_query = ', '.join([f"{key} = ?" for key in form_data.keys() if key not in ['created_at', 'recrutador']])

            values = [form_data[key] for key in form_data.keys() if key not in ['created_at', 'recrutador']] + [cpf]

            cursor.execute(f'UPDATE registration_form SET {update_query} WHERE cpf = ?', values)

            flash('Formulário atualizado com sucesso!', 'success')



        db.commit()

        cursor.close()

        db.close()

        return redirect(url_for('view_form', cpf=cpf))



    cursor.close()

    db.close()

    return render_template('view_registration.html', form_data=form_data, situacao=form_data.get('situacao'), current_date=current_date, cpf=cpf)



def get_candidate_by_cpf(cpf):

    """Busca um candidato pelo CPF no banco de dados."""

    db = get_sql_server_connection()

    cursor = db.cursor()

    try:

        cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

        result = cursor.fetchone()

        

        if result:

            # Obtém os nomes das colunas

            column_names = [column[0] for column in cursor.description]

            # Converte os resultados para um dicionário usando zip

            form_data = dict(zip(column_names, result))

            form_data['situacao'] = determinar_situacao(form_data)

            return form_data

        return None

    finally:

        cursor.close()

        db.close()



@app.route('/modal_view/<cpf>')

def modal_view(cpf):

    # Busca os dados do candidato pelo CPF

    form_data = get_candidate_by_cpf(cpf)

    if not form_data:

        return "Candidato não encontrado", 404



    # Retorna o template view_registration.html para o modal

    current_date = datetime.now().strftime('%d/%m/%Y')

    return render_template(

        'view_registration.html', 

        form_data=form_data, 

        situacao=form_data.get('situacao'), 

        current_date=current_date, 

        cpf=cpf

    )
def update_form(cpf):

    db = None

    cursor = None

    

    try:

        db = get_sql_server_connection()

        cursor = db.cursor()

        

        # Busca o formulário com base no CPF

        cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

        form = cursor.fetchone()

        

        if not form:

            flash("Formulário não encontrado.", 'danger')

            return redirect(url_for('view_form', cpf=cpf))

        

        # Obtém os nomes das colunas da tabela

        column_names = [column[0] for column in cursor.description]

        

        # Mapeia os resultados para um dicionário

        form = dict(zip(column_names, form))

        

        if request.method == 'POST':

            form_data = request.form.to_dict()

            

            # Adiciona timestamp de atualização

            form_data['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            

            # Obtém valores específicos

            admitido = request.form.get('admitido', 'Não')

            

            # Captura os campos de lista e junta-os em strings

            campos_lista = ['cargo_indicado', 'cargo_pretendido', 'regioes_preferencia', 'disponibilidade_horario', 'rota_trabalho']

            for campo in campos_lista:

                valores = request.form.getlist(campo)

                if valores:

                    # Garantir que os valores são strings antes de juntar

                    form_data[campo] = ','.join([str(item).strip() for item in valores])

                else:

                    form_data[campo] = ''

            

            # Trata o campo de número de filhos

            numero_filhos = form_data.get('numero_filhos', 0)

            try:

                numero_filhos = int(numero_filhos) if numero_filhos else 0

            except ValueError:

                numero_filhos = 0

            form_data['numero_filhos'] = str(numero_filhos)

            

            # Trata dados de filhos

            for i in range(1, numero_filhos + 1):

                idade_filho_key = f'idade_filho_{i}'

                idade_filho = request.form.get(idade_filho_key, '')

                # Garantir que o valor é uma string

                form_data[idade_filho_key] = str(idade_filho).strip() if idade_filho else ''

            

            # Determina a situação baseada nas regras de negócio

            form_data['situacao'] = determinar_situacao(form_data)

            

            # Salvar o currículo, se houver

            if 'curriculo' in request.files and request.files['curriculo'].filename != '':

                try:

                    curriculo = request.files['curriculo']

                    curriculo_filename = secure_filename(curriculo.filename)

                    UPLOAD_DIR = os.path.join('static', 'uploads')

                    if not os.path.exists(UPLOAD_DIR):

                        os.makedirs(UPLOAD_DIR)

                    curriculo_path = os.path.join(UPLOAD_DIR, curriculo_filename)

                    curriculo.save(curriculo_path)

                    form_data['curriculo'] = curriculo_filename

                except Exception as e:

                    print(f"Erro ao salvar currículo: {e}")

                    # Continua mesmo se falhar

            

            # Verifica as colunas existentes na tabela

            cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'registration_form'")

            valid_columns = [row[0] for row in cursor.fetchall()]

            

            # Filtra para incluir apenas colunas válidas

            form_data_filtered = {}

            for k, v in form_data.items():

                if k in valid_columns:

                    # Garantir tipos de dados corretos

                    if v is None:

                        form_data_filtered[k] = ''

                    elif isinstance(v, (list, tuple, dict)):

                        form_data_filtered[k] = str(v)

                    else:

                        form_data_filtered[k] = str(v)

            

            # Logs para depuração

            print(f"Dados recebidos para atualização: {cpf}")

            

            # Se não houver dados para atualizar, retorne um erro

            if not form_data_filtered:

                return jsonify({"success": False, "message": "Nenhum dado válido para atualizar"}), 400

            

            # Prepara a query de atualização

            update_fields = []

            update_values = []

            

            for key, value in form_data_filtered.items():

                # Garantir que todos os valores são strings para evitar erros de tipo

                if value is None:

                    value = ''

                elif not isinstance(value, str):

                    value = str(value)

                

                update_fields.append(f"{key} = ?")

                update_values.append(value)

                

            # Adiciona o CPF como condição WHERE

            update_values.append(cpf)

            

            # Query de atualização completa

            update_query = f"""

                UPDATE registration_form 

                SET {", ".join(update_fields)}

                WHERE cpf = ?

            """

            

            # Log da query para depuração

            print(f"Query de atualização: {update_query}")

            print(f"Valores para atualização: {update_values}")

            

            cursor.execute(update_query, update_values)

            

            # Verifica se a atualização foi bem-sucedida

            if cursor.rowcount == 0:

                print(f"Nenhuma linha foi atualizada para o CPF: {cpf}")

                

            # Confirmar o valor atualizado no banco

            cursor.execute('SELECT * FROM registration_form WHERE cpf = ?', (cpf,))

            atualizado = cursor.fetchone()

            if atualizado:

                print(f"Registro atualizado com sucesso para CPF: {cpf}")

            else:

                print(f"Registro não encontrado após atualização para CPF: {cpf}")

            

            # Registra o sucesso da operação

            try:

                salvar_ficha_log(form_data_filtered)

            except Exception as log_error:

                print(f"Erro ao salvar log em arquivo: {str(log_error)}")

            

            # Se admitido = "Sim", duplicar o registro no banco de admitidos

            if admitido == 'Sim':

                cursor.execute('SELECT nome_completo FROM registration_form WHERE cpf = ?', (cpf,))

                candidato = cursor.fetchone()

                if candidato:

                    try:

                        # Verifica se já existe na tabela de admitidos

                        cursor.execute('SELECT id FROM admitidos WHERE nome_completo = ?', (candidato[0],))

                        admitido_existente = cursor.fetchone()

                        

                        if not admitido_existente:

                            # Obter o próximo ID para a tabela admitidos

                            cursor.execute('SELECT ISNULL(MAX(id), 0) + 1 FROM admitidos')

                            novo_id_admitidos = cursor.fetchone()[0]

                            

                            cursor.execute('''

                                INSERT INTO admitidos (id, nome_completo, admitido, data_admissao)

                                VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                            ''', (novo_id_admitidos, candidato[0], admitido))

                    except Exception as e:

                        print(f"Erro ao registrar admitido: {e}")

                        # Continue mesmo se esta parte falhar

            

            # LOG DE SUCESSO NO BANCO

            cursor.execute('''

                INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                VALUES (?, ?, ?, CURRENT_TIMESTAMP)

            ''', (cpf, "SUCESSO", "Formulário atualizado com sucesso!"))

            

            db.commit()

            return jsonify({"success": True, "message": "Formulário atualizado com sucesso!"})

        

    except Exception as e:

        if db:

            try:

                db.rollback()

            except Exception as rollback_error:

                print(f"Erro ao fazer rollback: {str(rollback_error)}")

                

        erro_msg = str(e)

        print(f"Erro ao atualizar o formulário: {erro_msg}")

        

        # Captura detalhes técnicos adicionais para depuração

        import traceback

        traceback_str = traceback.format_exc()

        print(f"Traceback completo: {traceback_str}")

        

        # LOG DE ERRO NO BANCO

        try:

            if cursor and cpf:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "ERRO", erro_msg[:950]))  # Limita tamanho da mensagem para evitar erros

                db.commit()

        except Exception as log_error:

            print(f"Erro ao registrar erro no banco: {str(log_error)}")

        

        return jsonify({"success": False, "message": "Erro ao atualizar o formulário!", "error": erro_msg}), 400

        

    finally:

        if cursor:

            cursor.close()

        if db:

            db.close()

@app.route('/get_registration/<cpf>', methods=['GET'])

def get_registration(cpf):

    db = get_sql_server_connection()

    cursor = db.cursor()

    try:

        # Consulta os dados do candidato com base no CPF

        cursor.execute("SELECT * FROM dbo.registration_form WHERE cpf = ?", (cpf,))

        row = cursor.fetchone()

        

        # Verifica se o candidato foi encontrado

        if not row:

            return jsonify({"error": "Candidato não encontrado"}), 404

        

        # Converte os resultados para um dicionário

        columns = [column[0] for column in cursor.description]

        candidato = dict(zip(columns, row))

        

        return jsonify(candidato)

    except Exception as e:

        print(f"Erro ao buscar candidato: {e}")

        return jsonify({"error": "Erro ao buscar candidato"}), 500

    finally:

        cursor.close()

        db.close()



@app.route('/update_registration', methods=['POST'])

def update_registration():

    """Atualiza os dados de uma ficha no banco de dados."""

    db = None

    cursor = None

    

    try:

        db = get_sql_server_connection()

        cursor = db.cursor()

        

        # Obter os dados do formulário

        cpf = request.form.get('cpf', '').strip()

        nome_completo = request.form.get('nome_completo', '').upper().strip()

        telefone = request.form.get('telefone', '').strip()

        endereco = request.form.get('endereco', '').upper().strip()

        bairro = request.form.get('bairro', '').upper().strip()

        cidade = request.form.get('cidade', '').upper().strip()

        observacoes = request.form.get('observacoes', '').upper().strip()

        

        # Verificar se o CPF existe

        cursor.execute('SELECT 1 FROM registration_form WHERE cpf = ?', (cpf,))

        if not cursor.fetchone():

            flash('CPF não encontrado no banco de dados.', 'danger')

            return redirect('/view_registrations')

        

        # Preparar a query de atualização

        update_query = """

            UPDATE dbo.registration_form

            SET nome_completo = ?, telefone = ?, endereco = ?, bairro = ?, cidade = ?, 

                observacoes = ?, last_updated = CURRENT_TIMESTAMP

            WHERE cpf = ?

        """

        

        # Executar a query

        cursor.execute(update_query, (nome_completo, telefone, endereco, bairro, cidade, observacoes, cpf))

        

        # Verificar se houve atualização

        if cursor.rowcount == 0:

            print(f"Atualização sem efeito para CPF: {cpf}")

            flash('Nenhum registro foi atualizado.', 'warning')

        else:

            print(f"Registro atualizado com sucesso para CPF: {cpf}")

            flash('Registro atualizado com sucesso!', 'success')

            

            # Registrar o log de atualização

            try:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "ATUALIZAÇÃO", "Atualização de dados básicos realizada com sucesso"))

            except Exception as log_error:

                print(f"Erro ao registrar log: {str(log_error)}")

                

        # Confirmar as alterações

        db.commit()

        return redirect('/view_registrations')

        

    except Exception as e:

        # Registrar erro detalhado

        erro_msg = str(e)

        print(f"Erro em update_registration: {erro_msg}")

        

        import traceback

        traceback_str = traceback.format_exc()

        print(f"Traceback completo: {traceback_str}")

        

        # Fazer rollback se necessário

        if db:

            try:

                db.rollback()

            except Exception as rollback_error:

                print(f"Erro ao fazer rollback: {str(rollback_error)}")

                

        # Tentar registrar erro no banco

        try:

            if cursor and cpf:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "ERRO-ATUALIZAÇÃO", erro_msg[:950]))

                db.commit()

        except Exception as log_error:

            print(f"Erro ao registrar log de erro: {str(log_error)}")

            

        flash(f'Erro ao atualizar registro: {erro_msg}', 'danger')

        return redirect('/view_registrations')

        

    finally:

        if cursor:

            cursor.close()

        if db:

            db.close()
def auto_save_form(cpf):

    db = None

    cursor = None

    record_id = None  # Variável definida aqui para evitar o erro

    

    try:

        # Normaliza o CPF (remove tudo que não for número)

        cpf = re.sub(r'[^0-9]', '', str(cpf))

        print(f"[AUTO-SAVE] Recebida solicitação para CPF: {cpf}")

        print(f"[AUTO-SAVE] Dados recebidos: {request.form}")

        

        db = get_sql_server_connection()

        cursor = db.cursor()

        

        # Verifica se o CPF existe

        cursor.execute('SELECT id FROM registration_form WHERE cpf = ?', (cpf,))

        result = cursor.fetchone()

        

        if not result:

            print(f"[AUTO-SAVE] CPF {cpf} não encontrado no banco de dados. Criando novo registro.")

            

            # Obter as colunas existentes na tabela

            cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'registration_form'")

            valid_columns = [row[0] for row in cursor.fetchall()]

            

            # Capturar os dados básicos para criação do registro

            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            

            # Dados iniciais obrigatórios

            base_data = {

                'cpf': cpf,

                'created_at': current_time,

                'last_updated': current_time,

                # Removida a atribuição automática de recrutador

                'recrutador': ''

            }

            

            # Pegar dados do formulário em submission

            form_data = request.form.to_dict()

            

            # Combinar com os dados base

            complete_data = {**base_data, **form_data}



            situacao_atualizada = determinar_situacao(complete_data)

            complete_data['situacao'] = situacao_atualizada



            # Filtrar para incluir apenas colunas válidas

            valid_data = {k: v for k, v in complete_data.items() if k in valid_columns}

            

            # Preparar a query de inserção

            columns = ', '.join(valid_data.keys())

            placeholders = ', '.join(['?' for _ in valid_data])

            values = list(valid_data.values())

            

            # Executar a inserção

            cursor.execute(f"INSERT INTO registration_form ({columns}) VALUES ({placeholders})", values)

            

            # Verificar se foi inserido corretamente

            if cursor.rowcount == 0:

                print(f"[AUTO-SAVE] ERRO: Não foi possível inserir o registro para CPF {cpf}")

                return jsonify(success=False, message="Não foi possível criar o registro"), 500

            else:

                print(f"[AUTO-SAVE] Registro criado com sucesso para CPF {cpf}")

                

            # Commit da transação

            db.commit()

            

            # Registrar log de criação

            try:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "AUTO-SAVE-CREATE", f"Registro criado automaticamente via auto-save"))

                db.commit()

            except Exception as log_error:

                print(f"[AUTO-SAVE] Erro ao registrar log de criação: {str(log_error)}")

        else:

            record_id = result[0]

            print(f"[AUTO-SAVE] Registro encontrado com ID: {record_id}")

            

        # A partir daqui, o código continua normalmente para atualizar o registro

        

        # Obter as colunas existentes na tabela

        cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'registration_form'")

        valid_columns = [row[0] for row in cursor.fetchall()]

        print(f"[AUTO-SAVE] Colunas válidas: {valid_columns}")

        

        # Processar os dados do formulário - usando MultiDict para capturar múltiplos valores

        form_data = request.form.to_dict(flat=False)  # Captura arrays de valores

        print(f"[AUTO-SAVE] Dados do formulário recebidos (MultiDict): {len(form_data)} campos")



        # Converte para dicionário normal, preservando arrays para campos multi-valor

        processed_data = {}

        for key, values in form_data.items():

            if len(values) > 1:

                processed_data[key] = ','.join(values)

                print(f"[AUTO-SAVE] Campo multi-valor '{key}' com valores: {values} -> '{processed_data[key]}'")

            else:

                processed_data[key] = values[0]



        # >>>> AGORA, calcular a situação <<<<

        situacao_atualizada = determinar_situacao(processed_data)

        processed_data['situacao'] = situacao_atualizada

        # >>>> FIM DA MODIFICAÇÃO <<<<



        # Filtrar para incluir apenas colunas existentes na tabela

        form_data_filtered = {k: (v if v is not None else '') 

                           for k, v in processed_data.items() 

                           if k in valid_columns}

        

        print(f"[AUTO-SAVE] Dados filtrados: {len(form_data_filtered)} campos válidos")

        print(f"[AUTO-SAVE] Campos que serão atualizados: {list(form_data_filtered.keys())}")

        

        # Se não há dados para atualizar, retorne sucesso vazio

        if not form_data_filtered:

            print(f"[AUTO-SAVE] Nenhum dado válido para atualizar")

            return jsonify(success=True, message="Nenhum dado para atualizar"), 200

        

        # Adicionar o timestamp de atualização

        form_data_filtered['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        

        # Preparar query de atualização

        update_fields = []

        update_values = []

        

        for key, value in form_data_filtered.items():

            update_fields.append(f"{key} = ?")

            update_values.append(value)

        

        # Adicionar CPF como condição WHERE

        update_values.append(cpf)

        

        # Query final

        update_query = f"UPDATE registration_form SET {', '.join(update_fields)} WHERE cpf = ?"

        print(f"[AUTO-SAVE] Query de atualização: {update_query}")

        

        # Executar a query

        cursor.execute(update_query, update_values)

        

        # Verifica se teve efeito

        if cursor.rowcount == 0:

            print(f"[AUTO-SAVE] Nenhum registro atualizado para CPF {cpf}")

            # Tente atualizar usando o ID em vez do CPF se a atualização não teve efeito

            update_query_by_id = f"UPDATE registration_form SET {', '.join(update_fields)} WHERE id = ?"

            update_values[-1] = record_id  # Substitui o CPF pelo ID

            cursor.execute(update_query_by_id, update_values)

            

            if cursor.rowcount == 0:

                print(f"[AUTO-SAVE] Falha na atualização com ID para CPF {cpf}")

                return jsonify(success=False, message="Falha ao atualizar registro"), 500

            else:

                print(f"[AUTO-SAVE] Registro atualizado com sucesso usando ID para CPF {cpf}")

        else:

            print(f"[AUTO-SAVE] Registro atualizado com sucesso para CPF {cpf}")

        

        # Commit das alterações

        db.commit()

        

        # Tenta registrar o log de salvamento automático

        try:

            cursor.execute('''

                INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                VALUES (?, ?, ?, CURRENT_TIMESTAMP)

            ''', (cpf, "AUTO-SAVE", f"Salvamento automático realizado com sucesso: {len(form_data_filtered)} campos"))

            db.commit()

            print(f"[AUTO-SAVE] Log registrado com sucesso para CPF {cpf}")

        except Exception as log_error:

            print(f"[AUTO-SAVE] Erro ao registrar log de auto-save: {str(log_error)}")

        

        # Backup local dos dados em caso de falha futura

        try:

            error_dir = os.path.join('static', 'backups')

            if not os.path.exists(error_dir):

                os.makedirs(error_dir)

                

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            filename = f"{cpf}_{timestamp}.json"

            

            with open(os.path.join(error_dir, filename), 'w', encoding='utf-8') as f:

                json.dump(form_data, f, ensure_ascii=False, indent=4)

                

            print(f"[AUTO-SAVE] Backup local salvo em {filename}")

        except Exception as backup_error:

            print(f"[AUTO-SAVE] Erro ao salvar backup local: {str(backup_error)}")

        

        return jsonify(success=True, situacao=situacao_atualizada), 200



    except Exception as e:

        # Registrar o erro detalhado

        erro_msg = str(e)

        print(f"[AUTO-SAVE] Erro em auto_save_form: {erro_msg}")

        

        import traceback

        traceback_str = traceback.format_exc()

        print(f"[AUTO-SAVE] Traceback completo: {traceback_str}")

        

        # Fazer rollback se necessário

        if db:

            try:

                db.rollback()

            except Exception as rollback_error:

                print(f"[AUTO-SAVE] Erro ao fazer rollback: {str(rollback_error)}")

        

        # Tentar registrar o erro no banco

        try:

            if cursor and cpf:

                cursor.execute('''

                    INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)

                ''', (cpf, "ERRO-AUTO-SAVE", erro_msg[:950]))

                db.commit()

                print(f"[AUTO-SAVE] Log de erro registrado para CPF {cpf}")

        except Exception as log_error:

            print(f"[AUTO-SAVE] Erro ao registrar log de erro: {str(log_error)}")

        

        # Tenta salvar os dados de erro para recuperação posterior

        try:

            error_dir = os.path.join('static', 'errors')

            if not os.path.exists(error_dir):

                os.makedirs(error_dir)

                

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            filename = f"{cpf}_erro_{timestamp}.json"

            

            with open(os.path.join(error_dir, filename), 'w', encoding='utf-8') as f:

                erro_data = {

                    'cpf': cpf,

                    'tipo_erro': "AUTO-SAVE-ERROR",

                    'mensagem': erro_msg,

                    'timestamp': timestamp,

                    'dados_formulario': request.form.to_dict()

                }

                json.dump(erro_data, f, ensure_ascii=False, indent=4)

                

            print(f"[AUTO-SAVE] Dados de erro salvos para recuperação futura em {filename}")

        except Exception as e:

            print(f"[AUTO-SAVE] Erro ao salvar dados para recuperação: {str(e)}")

        

        return jsonify(success=False, message=f"Erro: {str(e)}"), 500



    finally:

        if cursor:

            cursor.close()

        if db:

            db.close()



@app.route('/set_recruiter', methods=['POST'])

@login_required

def set_recruiter():

    cpf = request.form.get('cpf')

    recrutador = request.form.get('recrutador')



    if not cpf or not recrutador:

        # Se for uma requisição AJAX, retornar JSON

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return jsonify(success=False, message='CPF ou recrutador não fornecido.'), 400

        # Caso contrário, redirecionar com mensagem flash

        flash('CPF ou recrutador não fornecido.', 'danger')

        return redirect(request.referrer)



    db = get_sql_server_connection()

    cursor = db.cursor()

    

    try:

        cursor.execute('UPDATE registration_form SET recrutador = ? WHERE cpf = ?', (recrutador, cpf))

        db.commit()



        cursor.close()

        db.close()

        

        # Se for uma requisição AJAX, retornar JSON

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return jsonify(success=True, message='Recrutador definido com sucesso!', recrutador=recrutador)

        

        # Caso contrário, redirecionar com mensagem flash

        flash('Recrutador definido com sucesso!', 'success')

        return redirect(request.referrer)



    except Exception as e:

        cursor.close()

        db.close()

        

        # Log do erro para o console do servidor

        print(f"Erro ao definir recrutador para CPF {cpf}: {str(e)}")

        

        # Se for uma requisição AJAX, retornar JSON

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return jsonify(success=False, message=f'Erro ao definir recrutador: {str(e)}'), 500

        

        # Caso contrário, redirecionar com mensagem flash

        flash('Erro ao definir recrutador.', 'danger')

        return redirect(request.referrer)
def create_ficha_manual():

    nome_completo = request.form.get('nome_completo', '').strip()

    cpf = request.form.get('cpf', '').replace('.', '').replace('-', '').strip()

    data_nasc = request.form.get('data_nasc', '').strip()



    db = get_sql_server_connection()

    cursor = db.cursor()



    try:

        # Verifica se o CPF já existe no banco de dados

        cursor.execute('SELECT cpf FROM registration_form WHERE cpf = ?', (cpf,))

        existing_candidate = cursor.fetchone()



        if existing_candidate:

            flash('Já existe uma ficha criada com este CPF!', 'warning')

            return redirect(url_for('banco_rs'))



        # Obtenha todas as colunas válidas da tabela

        cursor.execute('SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = ?', ('registration_form',))

        valid_columns = {row[0] for row in cursor.fetchall()}



        # Valida e converte a data de nascimento para o formato correto (YYYY-MM-DD)

        if data_nasc:

            try:

                if "/" in data_nasc:

                    data_nasc = datetime.strptime(data_nasc, '%d/%m/%Y').strftime('%Y-%m-%d')

                else:

                    data_nasc = datetime.strptime(data_nasc, '%Y-%m-%d').strftime('%Y-%m-%d')

            except ValueError:

                flash('Formato de data inválido. Use DD/MM/YYYY ou YYYY-MM-DD.', 'danger')

                return redirect(url_for('banco_rs'))

        else:

            data_nasc = None



        # Dados para inserção (Sem a coluna ID, pois ela é gerada automaticamente)

        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        form_data = {

            'nome_completo': nome_completo,

            'cpf': cpf,

            'data_nasc': data_nasc,

            'created_at': created_at,

            'estado_civil': '',

            'telefone': '',

            'endereco': '',

            'cep': '',

            'complemento': '',

            'bairro': '',

            'numero': '',

            'cidade': '',

            'estado_nasc': '',

            'cidade_nasc': '',

            'idade': '',

            'numero_filhos': '',

            'fumante': '',

            'bebida': '',

            'alergia': '',

            'medicamento': '',

            'genero': '',

            'peso': '',

            'cor_pele': '',

            'tatuagem': '',

            'perfil': '',

            'cargo_indicado': '',

            'empresa1': '',

            'funcao1': '',

            'data_admissao1': '',

            'data_saida1': '',

            'salario1': '',

            'empresa2': '',

            'funcao2': '',

            'data_admissao2': '',

            'data_saida2': '',

            'salario2': '',

            'empresa3': '',

            'funcao3': '',

            'data_admissao3': '',

            'data_saida3': '',

            'salario3': '',

            'observacoes': '',

            'empregos_informais': '',

            'avaliacao_rh': '',

            'assinatura_rh': '',

            'avaliacao_gerencia': '',

            'assinatura_gerencia': '',

            'rota_trabalho': '',

            'cursos': '',

            'curriculo': '',

            'situacao': '',

            'last_updated': ''

        }



        # Filtrar colunas válidas para evitar erros

        form_data = {k: v for k, v in form_data.items() if k in valid_columns}



        # Construa a query dinamicamente

        columns = ', '.join(form_data.keys())

        placeholders = ', '.join(['?' for _ in form_data])



        # Insere os dados (Sem o campo ID)

        cursor.execute(f'INSERT INTO registration_form ({columns}) VALUES ({placeholders})', tuple(form_data.values()))

        db.commit()



        flash('Ficha criada com sucesso!', 'success')

        return redirect(url_for('view_form', cpf=cpf))



    except Exception as e:

        print(f"Erro ao criar ficha manual: {e}")

        flash('Erro ao criar a ficha.', 'danger')

        return redirect(url_for('banco_rs'))



    finally:

        cursor.close()

        db.close()



@app.route('/admin/report', methods=['GET', 'POST'])

@login_required

@admin_required

def admin_report():

    if request.method == 'POST':

        start_date = request.form.get('start_date')

        end_date = request.form.get('end_date')



        db = get_sql_server_connection()

        cursor = db.cursor()



        # Consultas para as métricas

        cursor.execute('''

            SELECT COUNT(*) FROM tickets WHERE category = 'Entrevista' AND created_at BETWEEN ? AND ?

        ''', (start_date, end_date))

        total_interviews = cursor.fetchone()[0]



        cursor.execute('''

            SELECT COUNT(*) FROM tickets WHERE category = 'Agendado' AND created_at BETWEEN ? AND ?

        ''', (start_date, end_date))

        total_scheduled = cursor.fetchone()[0]



        cursor.execute('''

            SELECT COUNT(*) FROM registration_form WHERE avaliacao_rh = 'Aprovado' AND last_updated BETWEEN ? AND ?

        ''', (start_date, end_date))

        total_approved = cursor.fetchone()[0]



        cursor.execute('''

            SELECT COUNT(*) FROM registration_form WHERE avaliacao_rh = 'Reprovado' AND last_updated BETWEEN ? AND ?

        ''', (start_date, end_date))

        total_rejected = cursor.fetchone()[0]



        cursor.execute('''

            SELECT recruiter, COUNT(*) FROM tickets WHERE created_at BETWEEN ? AND ? GROUP BY recruiter

        ''', (start_date, end_date))

        recruiter_counts = cursor.fetchall()



        # Calcular o tempo médio de espera

        cursor.execute('''

            SELECT AVG(DATEDIFF(MINUTE, created_at, called_at)) FROM tickets WHERE created_at BETWEEN ? AND ?

        ''', (start_date, end_date))

        average_wait_time = cursor.fetchone()[0]



        # Calcular o tempo médio de atendimento

        cursor.execute('''

            SELECT AVG(DATEDIFF(MINUTE, called_at, concluded_at)) FROM tickets WHERE created_at BETWEEN ? AND ?

        ''', (start_date, end_date))

        average_service_time = cursor.fetchone()[0]



        cursor.close()

        db.close()



        return render_template('admin_report.html', 

                               total_interviews=total_interviews,

                               total_scheduled=total_scheduled,

                               total_approved=total_approved,

                               total_rejected=total_rejected,

                               recruiter_counts=recruiter_counts,

                               average_wait_time=average_wait_time,

                               average_service_time=average_service_time,

                               start_date=start_date,

                               end_date=end_date)



    return render_template('admin_report.html')





def _processar_arquivo_erro(filepath):

    """Processa um arquivo de erro e tenta salvar os dados no banco"""

    try:

        with open(filepath, 'r', encoding='utf-8') as f:

            error_data = json.load(f)

            

        cpf = error_data.get('cpf')

        dados_form = error_data.get('dados_formulario', {})

        

        # Verificar se já existe no banco

        db = get_sql_server_connection()

        cursor = db.cursor()

        

        try:

            cursor.execute('SELECT id FROM registration_form WHERE cpf = ?', (cpf,))

            existing = cursor.fetchone()

            

            if existing:

                # O registro já existe, então não precisa recuperar

                os.rename(filepath, filepath + '.processed')

                return {'success': True, 'cpf': cpf, 'message': 'Registro já existe no banco'}

                

            # Prepara para inserção no banco

            # Remove campos que podem causar problemas

            for field in ['id', 'created_at', 'last_updated']:

                if field in dados_form:

                    dados_form.pop(field)

                    

            # Obter colunas válidas

            cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'registration_form'")

            valid_columns = [row[0] for row in cursor.fetchall()]

            

            # Filtrar para colunas válidas

            dados_filtered = {k: v for k, v in dados_form.items() if k in valid_columns}

            

            # Preparar query

            columns = list(dados_filtered.keys())

            placeholders = ', '.join(['?' for _ in columns])

            values = [dados_filtered[col] for col in columns]

            

            # Inserir no banco

            cursor.execute(f'''

                INSERT INTO registration_form ({', '.join(columns)}, created_at, last_updated)

                VALUES ({placeholders}, GETDATE(), GETDATE())

            ''', values)

            

            # Registrar no log

            cursor.execute('''

                INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                VALUES (?, ?, ?, GETDATE())

            ''', (cpf, "RECUPERAÇÃO", f"Registro recuperado de arquivo de erro {os.path.basename(filepath)}"))

            

            db.commit()

            

            # Marcar arquivo como processado

            os.rename(filepath, filepath + '.processed')

            

            return {'success': True, 'cpf': cpf, 'message': 'Registro recuperado com sucesso'}

            

        except Exception as db_error:

            db.rollback()

            return {'success': False, 'cpf': cpf, 'message': f"Erro ao processar: {str(db_error)}"}

            

        finally:

            cursor.close()

            db.close()

            

    except Exception as e:

        return {'success': False, 'cpf': 'Desconhecido', 'message': f"Erro ao ler arquivo: {str(e)}"}



def _processar_arquivo_backup(filepath):

    """Processa um arquivo de backup e tenta salvar os dados no banco"""

    try:

        with open(filepath, 'r', encoding='utf-8') as f:

            backup_data = json.load(f)

            

        cpf = backup_data.get('cpf')

        

        # Se não tiver CPF, não é possível processar

        if not cpf:

            return {'success': False, 'cpf': 'Desconhecido', 'message': 'CPF não encontrado no arquivo de backup'}

        

        # Verificar se já existe no banco

        db = get_sql_server_connection()

        cursor = db.cursor()

        

        try:

            cursor.execute('SELECT id FROM registration_form WHERE cpf = ?', (cpf,))

            existing = cursor.fetchone()

            

            if existing:

                # O registro já existe, não precisa recuperar

                os.rename(filepath, filepath + '.processed')

                return {'success': True, 'cpf': cpf, 'message': 'Registro já existe no banco'}

                

            # Prepara para inserção no banco

            # Remove campos que podem causar problemas

            for field in ['id', 'created_at', 'last_updated']:

                if field in backup_data:

                    backup_data.pop(field)

                    

            # Obter colunas válidas

            cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'registration_form'")

            valid_columns = [row[0] for row in cursor.fetchall()]

            

            # Filtrar para colunas válidas

            dados_filtered = {k: v for k, v in backup_data.items() if k in valid_columns}

            

            # Preparar query

            columns = list(dados_filtered.keys())

            placeholders = ', '.join(['?' for _ in columns])

            values = [dados_filtered[col] for col in columns]

            

            # Inserir no banco

            cursor.execute(f'''

                INSERT INTO registration_form ({', '.join(columns)}, created_at, last_updated)

                VALUES ({placeholders}, GETDATE(), GETDATE())

            ''', values)

            

            # Registrar no log

            cursor.execute('''

                INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                VALUES (?, ?, ?, GETDATE())

            ''', (cpf, "RECUPERAÇÃO", f"Registro recuperado de arquivo de backup {os.path.basename(filepath)}"))

            

            db.commit()

            

            # Marcar arquivo como processado

            os.rename(filepath, filepath + '.processed')

            

            return {'success': True, 'cpf': cpf, 'message': 'Backup recuperado com sucesso'}

            

        except Exception as db_error:

            db.rollback()

            return {'success': False, 'cpf': cpf, 'message': f"Erro ao processar: {str(db_error)}"}

            

        finally:

            cursor.close()

            db.close()

            

    except Exception as e:

        return {'success': False, 'cpf': 'Desconhecido', 'message': f"Erro ao ler arquivo: {str(e)}"}



@app.route('/local_recovery')

@login_required

def local_recovery():

    """Interface para o usuário recuperar fichas salvas localmente no navegador."""

    return render_template('local_recovery.html')
@app.route('/admin/recovery', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_recovery():

    """Interface administrativa para recuperar fichas que falharam no salvamento."""

    error_dir = os.path.join('static', 'errors')

    backup_dir = os.path.join('static', 'backups')

    

    # Garante que os diretórios existem

    if not os.path.exists(error_dir):

        os.makedirs(error_dir)

    if not os.path.exists(backup_dir):

        os.makedirs(backup_dir)

    

    # Recuperar arquivos de erro e backup

    error_files = []

    backup_files = []

    

    # Lista os arquivos de erro

    for file in os.listdir(error_dir):

        if file.endswith('.json'):

            try:

                with open(os.path.join(error_dir, file), 'r', encoding='utf-8') as f:

                    data = json.load(f)

                    error_files.append({

                        'filename': file,

                        'cpf': data.get('cpf', 'Desconhecido'),

                        'tipo_erro': data.get('tipo_erro', 'Desconhecido'),

                        'timestamp': data.get('timestamp', 'Desconhecido'),

                        'mensagem': data.get('mensagem', '')[:150] + '...' if len(data.get('mensagem', '')) > 150 else data.get('mensagem', '')

                    })

            except Exception as e:

                error_files.append({

                    'filename': file,

                    'cpf': 'Erro ao ler',

                    'tipo_erro': 'Erro de Leitura',

                    'timestamp': '-',

                    'mensagem': f"Erro ao ler arquivo: {str(e)}"

                })

    

    # Lista os arquivos de backup

    for file in os.listdir(backup_dir):

        if file.endswith('.json'):

            parts = file.split('_')

            if len(parts) >= 2:

                cpf = parts[0]

                timestamp = '_'.join(parts[1:]).replace('.json', '')

                backup_files.append({

                    'filename': file,

                    'cpf': cpf,

                    'timestamp': timestamp

                })

    

    # Processa a tentativa de recuperação se for uma solicitação POST

    if request.method == 'POST':

        action = request.form.get('action')

        file_to_process = request.form.get('file')

        

        if action == 'recover_error':

            result = _processar_arquivo_erro(os.path.join(error_dir, file_to_process))

            if result.get('success'):

                flash(f"Registro recuperado com sucesso para o CPF: {result.get('cpf')}", 'success')

            else:

                flash(f"Falha ao recuperar registro: {result.get('message')}", 'danger')

                

        elif action == 'recover_backup':

            result = _processar_arquivo_backup(os.path.join(backup_dir, file_to_process))

            if result.get('success'):

                flash(f"Backup recuperado com sucesso para o CPF: {result.get('cpf')}", 'success')

            else:

                flash(f"Falha ao recuperar backup: {result.get('message')}", 'danger')

                

        elif action == 'recover_all':

            total_success = 0

            total_failure = 0

            

            # Processa todos os arquivos de erro

            for error_file in error_files:

                result = _processar_arquivo_erro(os.path.join(error_dir, error_file['filename']))

                if result.get('success'):

                    total_success += 1

                else:

                    total_failure += 1

                    

            # Processa todos os arquivos de backup que não foram processados nos erros

            for backup_file in backup_files:

                # Verifica se o CPF já foi processado com sucesso

                if backup_file['cpf'] not in [ef['cpf'] for ef in error_files if ef['cpf'] in [r.get('cpf') for r in [_processar_arquivo_erro(os.path.join(error_dir, ef['filename'])) for ef in error_files] if r.get('success')]]:

                    result = _processar_arquivo_backup(os.path.join(backup_dir, backup_file['filename']))

                    if result.get('success'):

                        total_success += 1

                    else:

                        total_failure += 1

                        

            if total_success > 0:

                flash(f"Recuperação em massa concluída. {total_success} registros recuperados com sucesso. {total_failure} falhas.", 'success' if total_failure == 0 else 'warning')

            else:

                flash(f"Falha na recuperação em massa. {total_failure} operações falharam.", 'danger')

                

        # Redireciona para atualizar a lista após o processamento

        return redirect(url_for('admin_recovery'))

    

    return render_template('admin_recovery.html', error_files=error_files, backup_files=backup_files)



@app.route('/sync_local_backup', methods=['POST'])

@login_required

def sync_local_backup():

    """

    Rota para sincronizar backups locais com o servidor.

    Esta rota recebe dados de backups armazenados em localStorage e os salva no servidor.

    """

    try:

        data = request.get_json()

        if not data or 'cpf' not in data or 'backup_data' not in data:

            return jsonify(success=False, message="Dados incompletos"), 400

            

        cpf = data['cpf']

        backup_data = data['backup_data']

        timestamp = data.get('timestamp', datetime.now().isoformat())

        

        # Criar diretório de backups se não existir

        backup_dir = os.path.join('static', 'backups')

        if not os.path.exists(backup_dir):

            os.makedirs(backup_dir)

            

        # Salvar o backup no servidor com timestamp

        filename = f"{cpf}_{datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%Y%m%d_%H%M%S')}.json"

        filepath = os.path.join(backup_dir, filename)

        

        with open(filepath, 'w', encoding='utf-8') as f:

            json.dump({

                'cpf': cpf,

                'data': backup_data,

                'timestamp': timestamp,

                'origin': 'client_backup',

                'username': current_user.username

            }, f, ensure_ascii=False, indent=4)

            

        # Registrar em log

        try:

            db = get_sql_server_connection()

            cursor = db.cursor()

            cursor.execute('''

                INSERT INTO log_salvamentos (cpf, acao, mensagem, data_hora)

                VALUES (?, ?, ?, CURRENT_TIMESTAMP)

            ''', (cpf, "BACKUP-SYNC", f"Backup local sincronizado: {filename}"))

            db.commit()

            cursor.close()

            db.close()

        except Exception as log_error:

            print(f"Erro ao registrar log de sincronização: {str(log_error)}")

            

        return jsonify(success=True, message="Backup sincronizado com sucesso"), 200

        

    except Exception as e:

        print(f"Erro ao sincronizar backup local: {str(e)}")

        return jsonify(success=False, message=f"Erro: {str(e)}"), 500

        

@app.route('/admin/backups', methods=['GET', 'POST'])

@login_required

@admin_required

def admin_backups():

    """

    Interface administrativa para gerenciar backups

    """

    # Diretório onde os backups são armazenados

    backup_dir = os.path.join('static', 'backups')

    if not os.path.exists(backup_dir):

        os.makedirs(backup_dir)

    

    # Listar todos os arquivos de backup

    backup_files = []

    for filename in os.listdir(backup_dir):

        if filename.endswith('.json'):

            try:

                filepath = os.path.join(backup_dir, filename)

                with open(filepath, 'r', encoding='utf-8') as f:

                    try:

                        backup_data = json.load(f)

                        # Extrair CPF do nome do arquivo (formato: cpf_timestamp.json)

                        cpf = filename.split('_')[0] if '_' in filename else 'desconhecido'

                        timestamp = backup_data.get('timestamp', 'desconhecido')

                        

                        # Formatar a data para exibição

                        try:

                            date_obj = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))

                            formatted_date = date_obj.strftime('%d/%m/%Y %H:%M:%S')

                        except:

                            formatted_date = timestamp

                        

                        backup_files.append({

                            'cpf': cpf,

                            'filename': filename,

                            'timestamp': timestamp,

                            'formatted_date': formatted_date,

                            'size': os.path.getsize(filepath),

                            'username': backup_data.get('username', 'desconhecido')

                        })

                    except json.JSONDecodeError:

                        # Arquivo inválido

                        backup_files.append({

                            'cpf': 'erro',

                            'filename': filename,

                            'timestamp': 'erro',

                            'formatted_date': 'Arquivo inválido',

                            'size': os.path.getsize(filepath),

                            'username': 'desconhecido'

                        })

            except Exception as e:

                print(f"Erro ao processar arquivo de backup {filename}: {e}")

    

    # Ordenar por data (mais recente primeiro)

    backup_files.sort(key=lambda x: x.get('timestamp', ''), reverse=True)

    

    # Processar ações no POST

    if request.method == 'POST':

        action = request.form.get('action')

        selected_file = request.form.get('filename')

        

        if action == 'restore' and selected_file:

            # Recuperar o backup

            try:

                filepath = os.path.join(backup_dir, selected_file)

                result = _processar_arquivo_backup(filepath)

                

                if result.get('success'):

                    flash(f"Backup recuperado com sucesso para CPF {result.get('cpf')}", 'success')

                else:

                    flash(f"Falha ao recuperar backup: {result.get('message')}", 'danger')

            except Exception as e:

                flash(f"Erro ao recuperar backup: {str(e)}", 'danger')

        

        elif action == 'delete' and selected_file:

            # Excluir o backup

            try:

                filepath = os.path.join(backup_dir, selected_file)

                os.remove(filepath)

                flash(f"Backup {selected_file} excluído com sucesso", 'success')

            except Exception as e:

                flash(f"Erro ao excluir backup: {str(e)}", 'danger')

        

        # Redireciona para atualizar a lista

        return redirect(url_for('admin_backups'))

    

    return render_template('admin_backups.html', backup_files=backup_files)



@app.route('/upload_curriculo', methods=['POST'])

@login_required

def upload_curriculo():

    try:

        if 'curriculo' not in request.files:

            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        

        file = request.files['curriculo']

        if file.filename == '':

            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        

        if file:

            # Garante que o nome do arquivo é seguro

            filename = secure_filename(file.filename)

            

            # Garante que o diretório de uploads existe

            upload_dir = os.path.join('static', 'uploads')

            if not os.path.exists(upload_dir):

                os.makedirs(upload_dir)

                print(f"Diretório de uploads criado: {upload_dir}")

            

            # Salva o arquivo na pasta de uploads

            file_path = os.path.join(upload_dir, filename)

            file.save(file_path)

            

            # Verifica se o arquivo foi salvo corretamente

            if os.path.exists(file_path):

                print(f"Arquivo salvo com sucesso: {file_path}")

                

                # ATUALIZA O BANCO DE DADOS com o nome do arquivo

                # Obtém o CPF do usuário logado ou de um campo hidden no formulário

                cpf = request.form.get('cpf')

                

                if cpf:

                    try:

                        db = get_sql_server_connection()

                        cursor = db.cursor()

                        

                        # Atualiza o campo curriculo na tabela registration_form

                        cursor.execute('''

                            UPDATE registration_form 

                            SET curriculo = ?, last_updated = CURRENT_TIMESTAMP

                            WHERE cpf = ?

                        ''', (filename, cpf))

                        

                        # Verifica se a atualização foi bem-sucedida

                        if cursor.rowcount > 0:

                            print(f"Banco de dados atualizado com sucesso para CPF: {cpf}")

                            db.commit()

                        else:

                            print(f"Nenhum registro encontrado para CPF: {cpf}")

                        

                        cursor.close()

                        db.close()

                        

                    except Exception as db_error:

                        print(f"Erro ao atualizar banco de dados: {str(db_error)}")

                        # Continua mesmo se falhar a atualização do banco

                

                return jsonify({

                    'success': True, 

                    'message': 'Currículo enviado com sucesso!',

                    'filename': filename

                })

            else:

                print(f"Erro: Arquivo não foi salvo em {file_path}")

                return jsonify({'success': False, 'message': 'Erro ao salvar o arquivo'})

            

    except Exception as e:

        print(f"Erro no upload do currículo: {str(e)}")

        return jsonify({'success': False, 'message': str(e)})



@app.route('/get_recruiter', methods=['GET'])

@login_required

def get_recruiter():

    cpf = request.args.get('cpf')



    if not cpf:

        return jsonify(success=False, message='CPF não fornecido.'), 400



    db = get_sql_server_connection()

    cursor = db.cursor()

    

    try:

        cursor.execute('SELECT recrutador FROM registration_form WHERE cpf = ?', (cpf,))

        result = cursor.fetchone()



        cursor.close()

        db.close()

        

        if result and result[0]:

            return jsonify(success=True, recrutador=result[0])

        else:

            return jsonify(success=False, message='Recrutador não encontrado para este CPF.')

    except Exception as e:

        cursor.close()

        db.close()

        return jsonify(success=False, message=f'Erro ao buscar recrutador: {str(e)}')



@app.route('/send_to_dp/<int:id>', methods=['POST'])

@login_required

def send_to_dp(id):

                        db = get_sql_server_connection()

                        cursor = db.cursor()

                        try:

                            # Buscar o ticket (opcional, só para validar)

                            cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

                            ticket = cursor.fetchone()



                            if not ticket:

                                return jsonify(success=False, message="Ticket não encontrado"), 404



                            sent_to_dp_at = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')



                            # Atualizar ticket para enviar ao DP

                            cursor.execute('''

                                UPDATE tickets 

                                SET status = 'CHAMADO',

                                    dp_status = 'No DP',

                                    dp_start_time = ?,

                                    dp_process_start_time = NULL,

                                    dp_end_time = NULL,

                                    dp_notes = NULL,

                                    dp_responsible = NULL,

                                    dp_process_by = NULL,

                                    dp_completed_time = NULL,

                                    dp_completed_by = NULL,

                                    concluded_at = NULL

                                WHERE id = ?

                            ''', (sent_to_dp_at, id))



                            db.commit()



                            # Notificar painel DP em tempo real (SocketIO)

                            socketio.emit('new_dp_ticket', {

                                'id': id,

                                'ticket_number': ticket.ticket_number,

                                'name': ticket.name,

                                'category': ticket.category,

                                'dp_status': 'No DP',

                                'dp_start_time': sent_to_dp_at

                            }, namespace='/')



                            return jsonify(success=True, message="Ticket enviado para o DP com sucesso")

                

                        except Exception as e:

                            db.rollback()

                            return jsonify(success=False, message=f"Erro ao enviar ticket para o DP: {str(e)}"), 500


                        
                        finally:

                            cursor.close()

                            db.close()
@app.route('/painel_dp')
@login_required
def painel_dp():

    """

    Exibe o painel do Departamento Pessoal com tickets pendentes,

    em processamento e concluídos.

    """

    db = get_sql_server_connection()

    cursor = db.cursor()

    today = datetime.now().date()



    try:

        # Contagem de tickets por status

        dp_tickets_count = {

            'pending': 0,

            'in_progress': 0,

            'completed': 0

        }

        

        # Obter tickets pendentes (enviados para o DP mas ainda não iniciados)

        cursor.execute('''

            SELECT * FROM tickets 

            WHERE dp_status = 'No DP' 

            AND CONVERT(DATE, dp_start_time) = ? 

            ORDER BY dp_start_time

        ''', (today,))

        pending_tickets = cursor.fetchall()

        dp_tickets_count['pending'] = len(pending_tickets)

        

        # Obter tickets em processamento

        cursor.execute('''

            SELECT * FROM tickets 

            WHERE dp_status = 'Em Processamento' 

            AND CONVERT(DATE, dp_process_start_time) = ? 

            ORDER BY dp_process_start_time

        ''', (today,))

        in_progress_tickets = cursor.fetchall()

        dp_tickets_count['in_progress'] = len(in_progress_tickets)

        

        # Obter tickets concluídos hoje

        cursor.execute('''

            SELECT * FROM tickets 

            WHERE dp_status = 'Concluído' 

            AND CONVERT(DATE, dp_completed_time) = ? 

            ORDER BY dp_completed_time DESC

        ''', (today,))

        completed_tickets = cursor.fetchall()

        dp_tickets_count['completed'] = len(completed_tickets)

        

        # Formatar os tickets para o template

        def format_ticket(ticket):

            return {

                'id': ticket.id,

                'name': ticket.name,

                'category': ticket.category,

                'ticket_number': ticket.ticket_number,

                'cpf': ticket.cpf,

                'telefones': ticket.telefones,

                'rua': ticket.rua,

                'numero': ticket.numero,

                'complemento': ticket.complemento,

                'bairro': ticket.bairro,

                'cidade': ticket.cidade,

                'dp_status': ticket.dp_status,

                'dp_start_time': ticket.dp_start_time,

                'dp_process_start_time': ticket.dp_process_start_time,

                'dp_completed_time': ticket.dp_completed_time,

                'dp_notes': ticket.dp_notes

            }

        

        # Preparar dados para o template

        pending_tickets = [format_ticket(ticket) for ticket in pending_tickets]

        in_progress_tickets = [format_ticket(ticket) for ticket in in_progress_tickets]

        completed_tickets = [format_ticket(ticket) for ticket in completed_tickets]

        

        return render_template(

            'painel_dp.html',

            pending_tickets=pending_tickets,

            in_progress_tickets=in_progress_tickets,

            completed_tickets=completed_tickets,

            dp_tickets_count=dp_tickets_count

        )

    

    except Exception as e:

        flash(f'Erro ao carregar o painel do DP: {str(e)}', 'danger')

        return redirect(url_for('home'))

    

    finally:

        cursor.close()

        db.close()



@app.route('/start_dp_process/<int:id>', methods=['POST'])

@login_required

def start_dp_process(id):

    """

    Inicia o processamento de um ticket no Departamento Pessoal.

    Atualiza o status para 'Em Processamento' e registra o horário de início.

    """

    db = get_sql_server_connection()

    cursor = db.cursor()

    

    try:

        # Buscar o ticket

        cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

        ticket = cursor.fetchone()

        

        if not ticket:

            return jsonify(success=False, message="Ticket não encontrado"), 404

        

        # Verificar se o ticket está no status correto

        if ticket.dp_status != 'No DP':

            return jsonify(success=False, message="O ticket não está no status 'No DP'"), 400

        

        # Atualizar o status do ticket para 'Em Processamento'

        process_start_time = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')

        

        cursor.execute('''

            UPDATE tickets 

            SET dp_status = 'Em Processamento', 

                dp_process_start_time = ?,

                dp_process_by = ?

            WHERE id = ?

        ''', (process_start_time, current_user.name, id))

        

        db.commit()

        

        # Emitir evento para notificar em tempo real

        socketio.emit('dp_ticket_update', {

            'id': id,

            'dp_status': 'Em Processamento',

            'dp_process_start_time': process_start_time,

            'dp_process_by': current_user.name

        }, namespace='/')

        

        return jsonify(success=True, message="Processamento iniciado com sucesso")

        

    except Exception as e:

        db.rollback()

        return jsonify(success=False, message=f"Erro ao iniciar processamento: {str(e)}"), 500

        

    finally:

        cursor.close()

        db.close()



@app.route('/complete_dp_process/<int:id>', methods=['POST'])

@login_required

def complete_dp_process(id):

    """

    Conclui o processamento de um ticket no Departamento Pessoal.

    Atualiza o status para 'Concluído' e registra o horário de conclusão.

    """

    db = get_sql_server_connection()

    cursor = db.cursor()

    

    try:

        # Buscar o ticket

        cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

        ticket = cursor.fetchone()

        

        if not ticket:

            return jsonify(success=False, message="Ticket não encontrado"), 404

        

        # Verificar se o ticket está no status correto

        if ticket.dp_status != 'Em Processamento':

            return jsonify(success=False, message="O ticket não está em processamento"), 400

        

        # Atualizar o status do ticket para 'Concluído'

        completed_time = datetime.now(BRASILIA_TZ).strftime('%Y-%m-%d %H:%M:%S')

        

        cursor.execute('''

            UPDATE tickets 

            SET dp_status = 'Concluído', 

                dp_completed_time = ?,

                dp_completed_by = ?,

                status = 'CONCLUIDO',

                concluded_at = ?

            WHERE id = ?

        ''', (completed_time, current_user.id, completed_time, id))

        

        db.commit()

        

        # Emitir evento para notificar em tempo real

        socketio.emit('dp_ticket_update', {

            'id': id,

            'dp_status': 'Concluído',

            'dp_completed_time': completed_time,

            'dp_completed_by': current_user.name

        }, namespace='/')

        

        return jsonify(success=True, message="Processamento concluído com sucesso")

        

    except Exception as e:

        db.rollback()

        return jsonify(success=False, message=f"Erro ao concluir processamento: {str(e)}"), 500

        

    finally:

        cursor.close()

        db.close()



@app.route('/get_dp_ticket/<int:id>')

@login_required

def get_dp_ticket(id):

    """

    Obtém os detalhes de um ticket do Departamento Pessoal.

    """

    db = get_sql_server_connection()

    cursor = db.cursor()

    

    try:

        # Buscar o ticket

        cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

        ticket = cursor.fetchone()

        

        if not ticket:

            return jsonify(success=False, message="Ticket não encontrado"), 404

        

        # Formatar o ticket

        ticket_data = {

            'id': ticket.id,

            'name': ticket.name,

            'category': ticket.category,

            'ticket_number': ticket.ticket_number,

            'cpf': ticket.cpf,

            'telefones': ticket.telefones,

            'rua': ticket.rua,

            'numero': ticket.numero,

            'complemento': ticket.complemento,

            'bairro': ticket.bairro,

            'cidade': ticket.cidade,

            'dp_status': ticket.dp_status,

            'dp_start_time': ticket.dp_start_time,

            'dp_process_start_time': ticket.dp_process_start_time,

            'dp_completed_time': ticket.dp_completed_time,

            'dp_notes': ticket.dp_notes

        }

        

        return jsonify(success=True, ticket=ticket_data)

        

    except Exception as e:

        return jsonify(success=False, message=f"Erro ao buscar ticket: {str(e)}"), 500

        

    finally:

        cursor.close()

        db.close()



@app.route('/save_dp_notes/<int:id>', methods=['POST'])

@login_required

def save_dp_notes(id):

    """

    Salva as observações do Departamento Pessoal para um ticket.

    """

    db = get_sql_server_connection()

    cursor = db.cursor()

    

    try:

        # Obter as observações do corpo da requisição

        data = request.json

        notes = data.get('notes', '')

        

        # Buscar o ticket

        cursor.execute('SELECT * FROM tickets WHERE id = ?', (id,))

        ticket = cursor.fetchone()

        

        if not ticket:

            return jsonify(success=False, message="Ticket não encontrado"), 404

        

        # Atualizar as observações do ticket

        cursor.execute('''

            UPDATE tickets 

            SET dp_notes = ?

            WHERE id = ?

        ''', (notes, id))

        

        db.commit()

        

        return jsonify(success=True, message="Observações salvas com sucesso")

        

    except Exception as e:

        db.rollback()

        return jsonify(success=False, message=f"Erro ao salvar observações: {str(e)}"), 500

        

    finally:

        cursor.close()

        db.close()



@app.route('/display_tv')

def display_tv():

    """

    Exibe o painel de TV do Departamento Pessoal com os tickets

    em espera, em atendimento e concluídos, incluindo o operador.

    """

    db = get_sql_server_connection()

    cursor = db.cursor()

    today = datetime.now().date()

    current_date = datetime.now().strftime('%d/%m/%Y')



    try:

        # Tickets Em Espera ("No DP")

        cursor.execute('''

            SELECT * FROM tickets 

            WHERE dp_status = 'No DP' 

            AND CONVERT(DATE, dp_start_time) = ? 

            ORDER BY dp_start_time

        ''', (today,))

        pending_tickets = cursor.fetchall()



        # Tickets Em Atendimento ("Em Processamento")

        cursor.execute('''

            SELECT * FROM tickets 

            WHERE dp_status = 'Em Processamento' 

            AND CONVERT(DATE, dp_process_start_time) = ? 

            ORDER BY dp_process_start_time DESC

        ''', (today,))

        in_progress_tickets = cursor.fetchall()



        # Tickets Concluídos

        cursor.execute('''

            SELECT * FROM tickets 

            WHERE dp_status = 'Concluído' 

            AND CONVERT(DATE, dp_completed_time) = ? 

            ORDER BY dp_completed_time DESC

        ''', (today,))

        completed_tickets = cursor.fetchall()



        # Função auxiliar para buscar nomes dos usuários

        def format_ticket(ticket):

            # Operador que iniciou atendimento

            operator_name = 'N/A'

            if getattr(ticket, 'dp_process_by', None):

                cursor.execute('SELECT name FROM users WHERE id = ?', (ticket.dp_process_by,))

                user = cursor.fetchone()

                operator_name = user[0] if user else 'N/A'

            # Quem concluiu

            completer_name = 'N/A'

            if getattr(ticket, 'dp_completed_by', None):

                cursor.execute('SELECT name FROM users WHERE id = ?', (ticket.dp_completed_by,))

                completer = cursor.fetchone()

                completer_name = completer[0] if completer else 'N/A'



            return {

                'id': ticket.id,

                'name': ticket.name,

                'cpf': ticket.cpf,

                'ticket_number': ticket.ticket_number,

                'category': ticket.category,

                'dp_status': ticket.dp_status,

                'dp_start_time': format_brazilian_date(ticket.dp_start_time),

                'dp_process_start_time': format_brazilian_date(ticket.dp_process_start_time) if ticket.dp_process_start_time else 'N/A',

                'dp_process_by': ticket.dp_process_by or 'Não informado',  # <-- Certo aqui!

                'dp_completed_time': format_brazilian_date(ticket.dp_completed_time) if ticket.dp_completed_time else 'N/A',

                'dp_completed_by': completer_name,

            }



        pending_tickets = [format_ticket(ticket) for ticket in pending_tickets]

        in_progress_tickets = [format_ticket(ticket) for ticket in in_progress_tickets]

        completed_tickets = [format_ticket(ticket) for ticket in completed_tickets]



        return render_template(

            'display_tv.html',

            pending_tickets=pending_tickets,

            in_progress_tickets=in_progress_tickets,

            completed_tickets=completed_tickets,

            current_date=current_date

        )



    except Exception as e:

        print(f"Erro ao carregar display_tv: {e}")

        return render_template(

            'display_tv.html',

            pending_tickets=[],

            in_progress_tickets=[],

            completed_tickets=[],

            current_date=current_date

        )

    finally:

        if cursor:

            cursor.close()

        if db:

            db.close()
def api_tickets_dp():

    db = get_sql_server_connection()

    cursor = db.cursor()

    today = datetime.now().date()



    # Tickets em espera (somente hoje)

    cursor.execute("""

        SELECT * FROM tickets WHERE dp_status = 'No DP'

        AND CONVERT(DATE, dp_start_time) = ?

        ORDER BY dp_start_time

    """, (today,))

    pending_tickets = [ticket_to_dict(ticket) for ticket in cursor.fetchall()]



    # Tickets em atendimento (somente hoje)

    cursor.execute("""

        SELECT * FROM tickets WHERE dp_status LIKE '%Em Processamento%'

        AND CONVERT(DATE, dp_process_start_time) = ?

        ORDER BY dp_process_start_time DESC

    """, (today,))

    in_progress_tickets = [ticket_to_dict(ticket) for ticket in cursor.fetchall()]



    # Tickets concluídos (somente hoje)

    cursor.execute("""

        SELECT * FROM tickets WHERE dp_status = 'Concluído'

        AND CONVERT(DATE, dp_completed_time) = ?

        ORDER BY dp_completed_time DESC

    """, (today,))

    completed_tickets = [ticket_to_dict(ticket) for ticket in cursor.fetchall()]



    cursor.close()

    db.close()

                

    return jsonify({
        "pending_tickets": pending_tickets,
        "in_progress_tickets": in_progress_tickets,
        "completed_tickets": completed_tickets,
    })







if __name__ == "__main__":

    socketio.run(app, host='192.168.0.79', port=5051, debug=True)